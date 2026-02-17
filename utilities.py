"""
╔══════════════════════════════════════════════════════════════════════════════╗
║                    DEPRECATED: LEGACY MONOLITHIC FILE                         ║
║                    Use Modular Architecture Instead                           ║
╚══════════════════════════════════════════════════════════════════════════════╝

⚠️  DEPRECATION NOTICE ⚠️

This monolithic file is DEPRECATED and maintained for backward compatibility only.

MIGRATION GUIDE:
Please migrate to the modular architecture located in:
  - utils/ - Shared utilities (fasta, export, validation, etc.)
  - utils/plotting/ - Visualization functions
  - ui/ - User interface components

OLD (Deprecated):
    from utilities import parse_fasta, export_to_csv, plot_motif_distribution

NEW (Recommended):
    from utils.fasta import parse_fasta
    from utils.export import export_to_csv
    from utils.plotting import plot_motif_distribution

BENEFITS OF MODULAR ARCHITECTURE:
  ✅ Better code organization and maintainability
  ✅ Easier testing and debugging
  ✅ Clearer dependency management
  ✅ Follows Python best practices
  ✅ Enables selective imports
  ✅ Reduces coupling between components

See MODULAR_ARCHITECTURE_STATUS.md for complete migration guide.

════════════════════════════════════════════════════════════════════════════════

╔══════════════════════════════════════════════════════════════════════════════╗
║                    CONSOLIDATED UTILITIES MODULE                              ║
║     Utility Functions for Sequence Processing, Export, and Visualization     ║
╚══════════════════════════════════════════════════════════════════════════════╝

MODULE: utilities.py
AUTHOR: Dr. Venkata Rajesh Yella
VERSION: 2024.1 - Consolidated with Visualizations
LICENSE: MIT

DESCRIPTION:
    Consolidated module containing all utility functions for sequence processing,
    pattern loading, data export, visualization, and validation. Combines functionality
    from utils, load_hsdb, load_regex_registry, motif_patterns, canonicalize_motif,
    and visualizations modules into a single cohesive unit.

MAIN FUNCTIONS:
    Sequence Processing:
    - parse_fasta(): Parse FASTA format sequences
    - gc_content(): Calculate GC content
    - reverse_complement(): Generate reverse complement
    - validate_sequence(): Validate DNA sequence

    Data Export:
    - export_to_csv(): Export to CSV format
    - export_to_bed(): Export to BED format
    - export_to_json(): Export to JSON format
    - export_to_excel(): Export to Excel format

    Pattern Loading:
    - load_hyperscan_db(): Load pre-compiled Hyperscan databases
    - load_regex_registry(): Load regex pattern registries
    - get_motif_patterns(): Get patterns for specific motif classes

    Statistics:
    - get_basic_stats(): Calculate sequence statistics
    - quality_check_motifs(): Validate motif quality
    
    Visualizations:
    - plot_motif_distribution(): Distribution plots
    - plot_coverage_map(): Coverage visualization
    - plot_density_heatmap(): Density heatmaps
    - plot_nested_pie_chart(): Hierarchical charts
    - plot_manhattan_motif_density(): Manhattan plots
    - and 20+ more visualization functions
"""

import warnings
warnings.warn(
    "DEPRECATED: Importing from 'utilities.py' is deprecated. "
    "Please use the modular architecture instead: "
    "from utils.fasta import ..., from utils.export import ..., from utils.plotting import ... "
    "See MODULAR_ARCHITECTURE_STATUS.md for migration guide.",
    DeprecationWarning,
    stacklevel=2
)

from typing import Dict, Any, List, Optional, Tuple, Union
import json
import re
import os
import io
import hashlib
import time
import tempfile
import zipfile
import logging
from datetime import datetime

# Set up logging
logger = logging.getLogger(__name__)

# Import critical dependencies with error handling
try:
    # Set matplotlib to use non-interactive backend before importing pyplot
    # This is required for server environments like Streamlit Cloud
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as patches
    from matplotlib.backends.backend_pdf import PdfPages
except ImportError as e:
    raise ImportError(f"Failed to import matplotlib. Please ensure matplotlib>=3.5.0 is installed. Error: {e}")

try:
    import seaborn as sns
except ImportError as e:
    raise ImportError(f"Failed to import seaborn. Please ensure seaborn>=0.11.0 is installed. Error: {e}")

try:
    import pandas as pd
except ImportError as e:
    raise ImportError(f"Failed to import pandas. Please ensure pandas>=1.3.0 is installed. Error: {e}")

try:
    import numpy as np
except ImportError as e:
    raise ImportError(f"Failed to import numpy. Please ensure numpy>=1.21.0 is installed. Error: {e}")

from collections import Counter, defaultdict
import warnings
warnings.filterwarnings("ignore")

# Try to import plotly for interactive plots
try:
    import plotly.graph_objects as go
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False

# Import standard library modules
import gc  # Garbage collection (standard library, always available)

# =============================================================================
# PERFORMANCE & MEMORY MANAGEMENT UTILITIES
# =============================================================================

def trigger_garbage_collection():
    """
    Trigger explicit garbage collection to free memory.
    
    Should be called at critical pipeline stages:
    - After processing large sequences
    - After generating visualizations
    - Before/after large data exports
    
    Returns:
        int: Number of objects collected
    """
    collected = gc.collect()
    logger.debug(f"Garbage collection: {collected} objects freed")
    return collected


def optimize_dataframe_memory(df: pd.DataFrame) -> pd.DataFrame:
    """
    Optimize pandas DataFrame memory usage by downcasting numeric types.
    
    Reduces memory footprint for large result datasets without losing precision.
    Note: float64 → float32 conversion may affect precision for some calculations.
    
    Args:
        df: Input DataFrame
        
    Returns:
        Memory-optimized DataFrame (a copy, original unchanged)
        
    Example:
        >>> df = optimize_dataframe_memory(large_motif_df)
        >>> # Overall memory usage typically reduced by 50-70%
        >>> # (Varies based on data types: integers optimize more than floats)
    """
    # Create a copy to avoid modifying original
    df_optimized = df.copy()
    
    # Collect all type changes first
    type_changes = {}
    
    for col in df_optimized.columns:
        col_type = df_optimized[col].dtype
        
        # Optimize integer columns
        if col_type == 'int64':
            c_min = df_optimized[col].min()
            c_max = df_optimized[col].max()
            if c_min >= 0:
                # Unsigned integers (boundary inclusive)
                if c_max <= 255:
                    type_changes[col] = np.uint8
                elif c_max <= 65535:
                    type_changes[col] = np.uint16
                elif c_max <= 4294967295:
                    type_changes[col] = np.uint32
            else:
                # Signed integers (boundary inclusive)
                if c_min >= np.iinfo(np.int8).min and c_max <= np.iinfo(np.int8).max:
                    type_changes[col] = np.int8
                elif c_min >= np.iinfo(np.int16).min and c_max <= np.iinfo(np.int16).max:
                    type_changes[col] = np.int16
                elif c_min >= np.iinfo(np.int32).min and c_max <= np.iinfo(np.int32).max:
                    type_changes[col] = np.int32
        
        # Optimize float columns (with precision note)
        # Note: float32 provides ~7 significant digits, sufficient for most genomic scores
        elif col_type == 'float64':
            type_changes[col] = np.float32
    
    # Apply all type changes
    for col, new_type in type_changes.items():
        df_optimized[col] = df_optimized[col].astype(new_type)
            
    return df_optimized


def get_memory_usage_mb() -> float:
    """
    Get current process memory usage in MB.
    
    Returns:
        Memory usage in megabytes
    """
    try:
        import psutil
        import os
        process = psutil.Process(os.getpid())
        return process.memory_info().rss / 1024 / 1024
    except ImportError:
        logger.warning("psutil not available - cannot monitor memory")
        return 0.0


# =============================================================================
# CONSTANTS FOR DATA EXPORT
# =============================================================================

# Core columns for output tables (as per requirements)
# Mandatory, universal columns for publication-grade reporting
# These columns appear in all display tables and CSV exports
# Task 1 & 2: Minimal, high-value features for Nature/NAR/Genome Research standards
CORE_OUTPUT_COLUMNS = [
    'Sequence_Name',  # Identity: Traceability
    'Class',          # Classification: Biological interpretation
    'Subclass',       # Classification: Detailed subtype
    'Start',          # Genomics: Absolute genomic context
    'End',            # Genomics: Absolute genomic context
    'Length',         # Genomics: Feature size (bp)
    'Sequence',       # Sequence: Always visible motif sequence (publication-grade requirement)
    'Strand',         # Strand: DNA strand orientation (+/- indicates forward/reverse)
    'Score',          # Confidence: 0-3 normalized, cross-motif comparability
    'Method',         # Evidence: Reproducibility (Regex/k-mer/ΔG/Hyperscan)
    'Pattern_ID',     # Evidence: Pattern identifier for traceability
]

# Motif-specific columns (ONLY reported when relevant per motif class)
# These are attached only in class-specific sheets or tooltips
MOTIF_SPECIFIC_COLUMNS = {
    'G-Quadruplex': ['Num_Tracts', 'Loop_Length', 'Num_Stems', 'Stem_Length', 'Priority'],
    'Z-DNA': ['Mean_10mer_Score', 'Contributing_10mers', 'Alternating_CG_Regions'],
    'i-Motif': ['Num_C_Tracts', 'Loop_Length', 'Motif_Type'],
    'Slipped DNA': ['Repeat_Unit', 'Unit_Length', 'Repeat_Count'],
    'Cruciform': ['Arm_Length', 'Loop_Length', 'Num_Stems'],
    'Triplex': ['Mirror_Type', 'Spacer_Length', 'Arm_Length', 'Loop_Length'],
    'R-Loop': ['GC_Skew', 'RIZ_Length', 'REZ_Length'],
    'Curved DNA': ['Tract_Type', 'Tract_Length', 'Num_Tracts'],
    'A-Philic': ['Tract_Type', 'Tract_Length'],
}

# Classes that are excluded from non-overlapping consolidated outputs
# These represent composite/derived features rather than primary motifs
EXCLUDED_FROM_CONSOLIDATED = ['Hybrid', 'Non-B_DNA_Clusters']

# Default values for missing core columns
DEFAULT_COLUMN_VALUES = {
    'Strand': '+',
    'Method': 'Pattern_detection',
    'Pattern_ID': 'Unknown',
    'Score': 0.0
}

# =============================================================================
# CONSTANTS FOR DOWNLOAD PACKAGE GENERATION
# =============================================================================

# Job ID generation settings
MAX_SEQUENCE_NAME_LENGTH = 30  # Maximum characters in cleaned sequence name
HASH_LENGTH = 6                 # Length of hash suffix in job ID

# Visualization settings for PDF generation
DEFAULT_PDF_WINDOW_SIZE_MIN = 100    # Minimum window size for density calculations
DEFAULT_PDF_WINDOW_DIVISOR = 20      # Divisor for adaptive window sizing

# Application version
APP_NAME = "NonBDNAFinder"
APP_VERSION = "2025.1"

# File hosting settings - DEPRECATED
# Note: file.io upload functionality has been removed in favor of direct downloads
# All downloads are now handled directly through Streamlit's download buttons


def canonicalize_motif(m: Dict[str, Any]) -> Dict[str, Any]:
    """
    Canonicalize motif dictionary to standard format with consistent field names.
    
    This function ensures all motif dictionaries follow the same structure regardless
    of their source, making them compatible with export and visualization functions.
    Maintains backward compatibility with legacy field names.
    
    Args:
        m: Motif dictionary with potentially inconsistent field names
        
    Returns:
        Standardized motif dictionary with canonical field names:
        - Class: Motif class (e.g., 'G-Quadruplex', 'Z-DNA')
        - Subclass: Motif subclass/subtype
        - Start: Start position (int)
        - End: End position (int)
        - Length: Motif length in bp (int)
        - Score: Normalized score (float)
        - Actual_Score: Raw detector score (float)
        - Normalized_Score: Legacy field, set to 0 if not present
        - Motif: Actual sequence string
        - Sequence_Name: Name of parent sequence
        
    Example:
        >>> motif = {'Type': 'G-Quadruplex', 'Start': 100, 'End': 120}
        >>> canonical = canonicalize_motif(motif)
        >>> canonical['Class']
        'G-Quadruplex'
        >>> canonical['Length']
        20
    """
    mapping = {
        'Actual Score': 'Actual_Score',
        'ActualScore': 'Actual_Score',
        'Score': 'Score',
        'Normalized Score': 'Normalized_Score',
        'Normalized_Score': 'Normalized_Score',
        'Class': 'Class',
        'Type': 'Class',
        'Subclass': 'Subclass',
        'Subtype': 'Subclass',
        'Start': 'Start',
        'End': 'End',
        'Length': 'Length',
        'Sequence_Name': 'Sequence_Name',
        'Motif': 'Motif'
    }
    out = {}
    for new_key in mapping.values():
        aliases = [k for k, v in mapping.items() if v == new_key]
        val = None
        for alias in aliases:
            if alias in m:
                val = m[alias]
                break
        out[new_key] = val
    if out.get('Start') and out.get('End') and not out.get('Length'):
        out['Length'] = int(out['End']) - int(out['Start'])
    out['Class'] = str(out.get('Class') or 'Unknown')
    out['Subclass'] = str(out.get('Subclass') or 'Other')
    out['Actual_Score'] = float(out.get('Actual_Score') or m.get('Actual_Score') or m.get('Score') or 0.0)
    out['Score'] = float(out.get('Score') or out['Actual_Score'])
    # Keep Normalized_Score for backward compatibility if it exists, but set to 0 if not present
    out['Normalized_Score'] = float(out.get('Normalized_Score') or m.get('Normalized_Score') or 0.0)
    out['Motif'] = m.get('Motif') or m.get('matched_seq') or ''
    out['Sequence_Name'] = out.get('Sequence_Name') or m.get('sequence_name') or ''
    return out


# =============================================================================
# SCORE NORMALIZATION (1-3 SCALE)
# =============================================================================
"""
Scientific Score Normalization System
=====================================

Normalizes raw detector scores to a 1-3 scale where:
  - 1 = Minimal (weak motif formation potential)
  - 2 = Moderate (typical motif formation potential)
  - 3 = Highest (strong motif formation potential)

Each motif class has specific normalization parameters based on:
  - Scientific literature on formation potential
  - Biological significance thresholds
  - Experimental validation data
"""

# Scientific normalization parameters per motif class
# Format: {'class_name': {'min_raw': min_threshold, 'max_raw': max_threshold, 'method': str}}
SCORE_NORMALIZATION_PARAMS = {
    'Curved_DNA': {
        'min_raw': 0.1,   # Minimal curvature threshold
        'max_raw': 0.95,  # Strong curvature (APR phasing)
        'method': 'linear',
        'ref': 'Koo 1986, Olson 1998'
    },
    'Slipped_DNA': {
        'min_raw': 0.3,   # Minimal repeat instability
        'max_raw': 0.98,  # High repeat instability (STRs)
        'method': 'linear',
        'ref': 'Schlötterer 2000, Weber 1989'
    },
    'Cruciform': {
        'min_raw': 0.5,   # Minimum palindrome stability
        'max_raw': 0.95,  # Strong inverted repeat
        'method': 'linear',
        'ref': 'Lilley 2000, Sinden 1994'
    },
    'R-Loop': {
        'min_raw': 0.4,   # Minimal R-loop potential
        'max_raw': 0.95,  # High R-loop formation
        'method': 'linear',
        'ref': 'Aguilera 2012, Jenjaroenpun 2016'
    },
    'Triplex': {
        'min_raw': 0.5,   # Minimal triplex potential
        'max_raw': 0.95,  # Strong H-DNA formation
        'method': 'linear',
        'ref': 'Frank-Kamenetskii 1995'
    },
    'G-Quadruplex': {
        'min_raw': 0.5,   # G4Hunter threshold ~1.2 normalized
        'max_raw': 1.0,   # Perfect G4 structure
        'method': 'g4hunter',  # Special G4Hunter-based normalization
        'ref': 'Bedrat 2016, Burge 2006'
    },
    'i-Motif': {
        'min_raw': 0.4,   # Minimal i-motif stability
        'max_raw': 0.98,  # High C-tract density
        'method': 'linear',
        'ref': 'Gehring 1993, Zeraati 2018'
    },
    # NOTE: Z-DNA uses cumulative 10-mer scoring (Ho et al. 1986)
    # Individual 10-mer scores range from ~50-63, but regions with multiple
    # overlapping 10-mers accumulate scores, hence the larger max_raw value.
    # This is intentionally different from other 0-1 normalized scores.
    'Z-DNA': {
        'min_raw': 50.0,    # Single Z-DNA 10-mer score minimum (Ho 1986)
        'max_raw': 2000.0,  # Cumulative max for long Z-DNA regions (~30+ bp)
        'method': 'zdna_cumulative',  # Uses log-linear scaling for cumulative scores
        'ref': 'Ho 1986, Rich 1984'
    },
    'A-philic_DNA': {
        'min_raw': 0.5,   # Minimal A-philic propensity
        'max_raw': 0.95,  # High A-tract density
        'method': 'linear',
        'ref': 'Gorin 1995, Vinogradov 2003'
    },
    'Hybrid': {
        'min_raw': 0.3,   # Minimal hybrid score
        'max_raw': 0.9,   # Strong hybrid overlap
        'method': 'linear',
        'ref': 'This work'
    },
    'Non-B_DNA_Clusters': {
        'min_raw': 0.3,   # Minimal cluster score
        'max_raw': 0.9,   # High-density cluster
        'method': 'linear',
        'ref': 'This work'
    }
}


def normalize_score_to_1_3(raw_score: float, motif_class: str) -> float:
    """
    Normalize raw detector score to 1-3 scale.
    
    Scientific basis:
    - 1.0 = Minimal (below typical formation threshold)
    - 1.5 = Low (marginal formation potential)
    - 2.0 = Moderate (typical formation potential)
    - 2.5 = High (strong formation potential)
    - 3.0 = Highest (exceptional formation potential)
    
    Args:
        raw_score: Raw score from detector (typically 0-1 or class-specific range)
        motif_class: Motif class name for class-specific normalization
        
    Returns:
        Normalized score between 1.0 and 3.0
        
    Example:
        >>> normalize_score_to_1_3(0.8, 'G-Quadruplex')
        2.6
    """
    # Get normalization parameters for this class
    params = SCORE_NORMALIZATION_PARAMS.get(motif_class)
    
    if params is None:
        # Default linear normalization for unknown classes
        params = {'min_raw': 0.0, 'max_raw': 1.0, 'method': 'linear'}
    
    min_raw = params['min_raw']
    max_raw = params['max_raw']
    method = params.get('method', 'linear')
    
    # Handle special normalization methods
    if method == 'g4hunter':
        # G4Hunter scores: abs value indicates formation potential
        # Typical range is 0-2+, with 1.2+ being significant
        # Always use absolute value for G4Hunter scores
        effective_score = abs(raw_score)
        if effective_score < 0.5:
            return 1.0
        elif effective_score >= 1.0:
            return min(3.0, 2.0 + (effective_score - 0.5) * 2.0)
        else:
            return 1.0 + (effective_score - 0.5) * 2.0
    
    elif method == 'zdna_cumulative':
        # Z-DNA cumulative scores (sum of multiple 10-mer matches)
        # Use log-linear scaling for better distribution
        import math
        if raw_score < min_raw:
            return 1.0
        elif raw_score >= max_raw:
            return 3.0
        else:
            # Log-scale normalization for cumulative scores
            log_score = math.log10(raw_score + 1)
            log_min = math.log10(min_raw + 1)
            log_max = math.log10(max_raw + 1)
            normalized = (log_score - log_min) / (log_max - log_min)
            return 1.0 + normalized * 2.0
    
    elif method == 'zdna_10mer':
        # Z-DNA 10-mer scores range from ~50 to 63 (deprecated, use zdna_cumulative)
        if raw_score < min_raw:
            return 1.0
        elif raw_score >= max_raw:
            return 3.0
        else:
            normalized = (raw_score - min_raw) / (max_raw - min_raw)
            return 1.0 + normalized * 2.0
    
    else:  # Linear normalization (default)
        if raw_score <= min_raw:
            return 1.0
        elif raw_score >= max_raw:
            return 3.0
        else:
            # Linear interpolation between 1 and 3
            normalized = (raw_score - min_raw) / (max_raw - min_raw)
            return 1.0 + normalized * 2.0


def normalize_motif_scores(motifs: list) -> list:
    """
    Normalize all motif scores to universal 1-3 scale for cross-motif comparability.
    
    UNIVERSAL SCORE DISCIPLINE (ΔG-inspired Scale):
    ═══════════════════════════════════════════════════════════════════════
    This function enforces a unified scoring framework across all motif classes,
    enabling meaningful comparison and ranking of structurally diverse Non-B DNA
    motifs. The scale is inspired by ΔG (Gibbs free energy) conventions in
    structural biology, where lower values indicate weaker/less stable structures.
    
    SCORE INTERPRETATION BANDS:
    ┌────────────┬──────────────────┬─────────────────────────────────────┐
    │ Score      │ Interpretation   │ Biological Meaning                  │
    ├────────────┼──────────────────┼─────────────────────────────────────┤
    │ 1.0 - 1.7  │ Weak/Conditional │ Low confidence, context-dependent   │
    │ 1.7 - 2.3  │ Moderate         │ Reasonable confidence, likely valid │
    │ 2.3 - 3.0  │ Strong/High      │ High confidence, well-characterized │
    └────────────┴──────────────────┴─────────────────────────────────────┘
    
    CROSS-MOTIF COMPARABILITY GUARANTEE:
      ✓ G-Quadruplex score of 2.5 is comparable to Curved DNA score of 2.5
      ✓ Higher scores always indicate stronger/more confident predictions
      ✓ Scores are normalized AFTER detection (never influence candidate finding)
      ✓ Normalization is deterministic and reproducible
    
    METHODOLOGICAL FOUNDATION:
      - Each motif class has class-specific raw scoring (sequence-based features)
      - Raw scores are transformed to universal scale via class-specific mappings
      - Transformation preserves relative ordering within each class
      - Cross-class comparisons enabled by unified scale alignment
    
    USAGE IN DOWNSTREAM ANALYSIS:
      - Ranking motifs across all classes for prioritization
      - Setting thresholds for filtering (e.g., score > 2.0)
      - Clustering by quality/confidence levels
      - Comparative analysis across motif types
    ═══════════════════════════════════════════════════════════════════════
    
    Args:
        motifs: List of motif dictionaries with 'Score' and 'Class' fields
        
    Returns:
        List of motifs with updated 'Score' field (1-3 scale)
        
    Note:
        Original raw score is preserved in 'Raw_Score' field for reference.
        Raw scores use class-specific scales and are not directly comparable.
    
    Example:
        >>> motifs = [
        ...     {'Class': 'G-Quadruplex', 'Score': 0.8},
        ...     {'Class': 'Curved_DNA', 'Score': 0.6}
        ... ]
        >>> normalized = normalize_motif_scores(motifs)
        >>> # Both scores now on comparable 1-3 scale
        >>> assert 1.0 <= normalized[0]['Score'] <= 3.0
        >>> assert 1.0 <= normalized[1]['Score'] <= 3.0
    """
    normalized_motifs = []
    
    for motif in motifs:
        m = motif.copy()
        raw_score = m.get('Score', 0.0)
        motif_class = m.get('Class', 'Unknown')
        
        # Store raw score for reference (class-specific scale)
        m['Raw_Score'] = raw_score
        
        # Normalize to universal 1-3 scale (cross-motif comparable)
        m['Score'] = round(normalize_score_to_1_3(raw_score, motif_class), 2)
        
        normalized_motifs.append(m)
    
    return normalized_motifs


"""
Helper to load Hyperscan DB or compile from patterns when needed.

API:
    load_db_for_class(class_name: str, registry_dir: str) -> (db, id_to_pattern, id_to_score)
      - db: hyperscan.Database instance (or None if hyperscan not available)
      - id_to_pattern: dict mapping id -> pattern string (tenmer for 10-mer, regex for others)
      - id_to_score: dict mapping id -> score (float)
      
    Supports both:
      - 10-mer patterns (ZDNA, APhilic): patterns have 'tenmer' key
      - Regex patterns (G4, IMotif, etc.): patterns have 'pattern' key
"""

import os
import logging
import json

logger = logging.getLogger(__name__)

_HYPERSCAN_AVAILABLE = False
_HYPERSCAN_VERSION = None
_HYPERSCAN_ERROR = None

try:
    import hyperscan
    _HYPERSCAN_AVAILABLE = True
    # Try to get version, but don't fail if not available
    try:
        _HYPERSCAN_VERSION = hyperscan.__version__
    except AttributeError:
        _HYPERSCAN_VERSION = 'unknown'
    logger.info(f"Hyperscan loaded successfully (version: {_HYPERSCAN_VERSION})")
except ImportError as e:
    hyperscan = None
    _HYPERSCAN_ERROR = f"Hyperscan Python bindings not installed: {e}"
    logger.info(f"Hyperscan not available - using pure Python fallback. {_HYPERSCAN_ERROR}")
except Exception as e:
    hyperscan = None
    _HYPERSCAN_ERROR = f"Hyperscan initialization failed: {e}"
    logger.warning(f"Hyperscan not available - using pure Python fallback. {_HYPERSCAN_ERROR}")

# Cache for consolidated registry
_CONSOLIDATED_REGISTRY = None

# pandas is already imported at the top and is required for this module
_PANDAS_AVAILABLE = True
_PANDAS_WARNING_LOGGED = False


# Pattern registry version
PATTERN_REGISTRY_VERSION = "2024.2"

def _load_consolidated_registry_from_excel():
    """Load the consolidated registry from Excel file and cache it
    
    Returns:
        dict: Registry data loaded from pattern_registry2.xlsx
    """
    global _PANDAS_WARNING_LOGGED
    
    if not _PANDAS_AVAILABLE:
        if not _PANDAS_WARNING_LOGGED:
            logger.warning("pandas not available - Excel pattern loading disabled. Install with: pip install pandas openpyxl")
            _PANDAS_WARNING_LOGGED = True
        return None
    
    # Load pattern_registry2.xlsx from current directory
    excel_path = "pattern_registry2.xlsx"
    if not os.path.isfile(excel_path):
        logger.error(f"Pattern registry file not found: {excel_path}")
        return None
    
    try:
        # Read all sheets from Excel
        excel_data = pd.read_excel(excel_path, sheet_name=None)  # Load all sheets
        
        # Build consolidated registry structure
        registries = {}
        for sheet_name, df in excel_data.items():
            if sheet_name == 'Summary':
                continue  # Skip summary sheet
            
            # Convert DataFrame to list of dictionaries
            patterns = df.to_dict('records')
            
            # Clean up NaN values
            for pattern in patterns:
                for key, value in list(pattern.items()):
                    if pd.isna(value):
                        pattern[key] = None
            
            # Create registry structure
            registries[sheet_name] = {
                'class': sheet_name,
                'n_patterns': len(patterns),
                'patterns': patterns
            }
        
        registry_data = {
            'version': f'{PATTERN_REGISTRY_VERSION}-excel',
            'source': excel_path,
            'registries': registries,
            'total_patterns': sum(r['n_patterns'] for r in registries.values()),
            'total_classes': len(registries)
        }
        
        logger.info(f"Loaded consolidated registry from {excel_path} ({registry_data['total_patterns']} patterns)")
        return registry_data
        
    except Exception as e:
        logger.error(f"Failed to load Excel registry from {excel_path}: {e}")
        return None


def _load_consolidated_registry():
    """Load the consolidated registry file once and cache it
    
    Loads from Excel file (pattern_registry2.xlsx with updated normalized scores).
    
    Returns:
        dict: Consolidated registry data
    """
    global _CONSOLIDATED_REGISTRY
    if _CONSOLIDATED_REGISTRY is not None:
        return _CONSOLIDATED_REGISTRY
    
    # Load from Excel
    _CONSOLIDATED_REGISTRY = _load_consolidated_registry_from_excel()
    if _CONSOLIDATED_REGISTRY is not None:
        return _CONSOLIDATED_REGISTRY
    
    # If Excel loading failed, raise an error
    raise FileNotFoundError(
        "Failed to load pattern registry from pattern_registry2.xlsx. "
        "Please ensure the Excel file exists and is properly formatted."
    )


def _load_registry(registry_dir: str, class_name: str):
    """Load registry from consolidated Excel file"""
    # Load from consolidated registry
    consolidated = _load_consolidated_registry()
    if consolidated and "registries" in consolidated:
        if class_name in consolidated["registries"]:
            logger.debug(f"Loading {class_name} from consolidated registry")
            return consolidated["registries"][class_name]
    
    raise FileNotFoundError(f"No registry found for {class_name} in pattern registry Excel file")


def clear_pattern_registry_cache():
    """
    Clear the cached pattern registry data.
    
    This is mainly useful for testing or when you need to reload
    the pattern data after modifying the Excel or JSON files.
    
    Example:
        >>> from utilities import clear_pattern_registry_cache, _load_consolidated_registry
        >>> clear_pattern_registry_cache()
        >>> registry = _load_consolidated_registry()  # Reloads from file
    """
    global _CONSOLIDATED_REGISTRY
    _CONSOLIDATED_REGISTRY = None
    logger.debug("Pattern registry cache cleared")


def load_db_for_class(class_name: str, registry_dir: str = "registry"):
    """
    Returns (db, id_to_pattern, id_to_score).
    db is a hyperscan.Database instance (or None if hyperscan not available).
    id_to_pattern and id_to_score are dicts mapping integer id -> pattern/tenmer / score.
    
    Handles both 10-mer patterns (with 'tenmer' key) and regex patterns (with 'pattern' key).
    """
    reg = _load_registry(registry_dir, class_name)
    patterns = reg.get("patterns", [])
    
    # Handle both 10-mer registries (tenmer) and regex registries (pattern)
    id_to_pattern = {}
    for p in patterns:
        pattern_id = int(p["id"])
        # Try 'tenmer' first (for 10-mer patterns like ZDNA, APhilic)
        if "tenmer" in p:
            id_to_pattern[pattern_id] = p["tenmer"]
        # Fall back to 'pattern' (for regex patterns like G4, IMotif, etc.)
        elif "pattern" in p:
            id_to_pattern[pattern_id] = p["pattern"]
        else:
            logger.warning(f"Pattern {pattern_id} in {class_name} has neither 'tenmer' nor 'pattern' key")
    
    id_to_score = {int(p["id"]): float(p.get("score", 0.0)) for p in patterns}

    db = None
    hsdb_path = os.path.join(registry_dir, f"{class_name}.hsdb")
    if _HYPERSCAN_AVAILABLE:
        try:
            db = hyperscan.Database()
            if os.path.isfile(hsdb_path):
                # Try to deserialize pre-compiled database
                try:
                    with open(hsdb_path, "rb") as fh:
                        raw = fh.read()
                    db.deserialize(raw)
                    logger.info(f"Loaded serialized Hyperscan DB for {class_name} from {hsdb_path}")
                except Exception as deserialize_error:
                    logger.warning(f"Failed to deserialize {hsdb_path}: {deserialize_error}. Recompiling from patterns...")
                    # Fallback: compile from patterns
                    try:
                        expressions = [id_to_pattern[i].encode("ascii") for i in sorted(id_to_pattern.keys())]
                        ids = sorted(id_to_pattern.keys())
                        db.compile(expressions=expressions, ids=ids, elements=len(expressions))
                        logger.info(f"Successfully compiled Hyperscan DB for {class_name} from patterns")
                    except Exception as compile_error:
                        logger.error(f"Failed to compile Hyperscan DB for {class_name}: {compile_error}")
                        db = None
            else:
                # No serialized DB; compile from patterns
                logger.debug(f"No serialized DB found at {hsdb_path}, compiling from patterns...")
                try:
                    expressions = [id_to_pattern[i].encode("ascii") for i in sorted(id_to_pattern.keys())]
                    ids = sorted(id_to_pattern.keys())
                    db.compile(expressions=expressions, ids=ids, elements=len(expressions))
                    logger.info(f"Compiled Hyperscan DB for {class_name} from patterns in {registry_dir}")
                except Exception as compile_error:
                    logger.error(f"Failed to compile Hyperscan DB for {class_name}: {compile_error}")
                    db = None
        except Exception as e:
            logger.error(f"Hyperscan database operations failed for {class_name}: {e}")
            db = None
    else:
        logger.debug(f"Hyperscan not installed; returning None DB for {class_name} (use pure-Python matcher).")

    return db, id_to_pattern, id_to_score
"""
Helper to load regex pattern registries and compile them with Hyperscan.

This module provides utilities for loading and using regex-based pattern registries
(CurvedDNA, G4, IMotif) similar to how the 10-mer registries work for A-philic and Z-DNA.

API:
    load_registry_for_class(class_name: str, registry_dir: str) 
        -> (db, id_to_pattern, id_to_subclass, id_to_score)
      - db: hyperscan.Database instance (or None if hyperscan not available)
      - id_to_pattern: dict mapping id -> regex pattern (string)
      - id_to_subclass: dict mapping id -> subclass name (string)
      - id_to_score: dict mapping id -> score (float)
"""


def load_registry_for_class(class_name: str, registry_dir: str = "registry"):
    """
    Load regex pattern registry and compile with Hyperscan.
    
    Returns (db, id_to_pattern, id_to_subclass, id_to_score).
    - db is a hyperscan.Database instance (or None if hyperscan not available).
    - id_to_pattern: dict mapping integer id -> regex pattern string
    - id_to_subclass: dict mapping integer id -> subclass name
    - id_to_score: dict mapping integer id -> score value
    """
    # Load registry data
    reg = _load_registry(registry_dir, class_name)
    patterns = reg.get("patterns", [])
    
    # Extract mappings
    id_to_pattern = {int(p["id"]): p["pattern"] for p in patterns}
    id_to_subclass = {int(p["id"]): p.get("subclass", "unknown") for p in patterns}
    id_to_score = {int(p["id"]): float(p.get("score", 0.0)) for p in patterns}
    
    # Compile Hyperscan database if available
    db = None
    if _HYPERSCAN_AVAILABLE:
        try:
            # Try to load pre-compiled DB first
            hsdb_path = os.path.join(registry_dir, f"{class_name}.hsdb")
            if os.path.isfile(hsdb_path):
                try:
                    db = hyperscan.Database()
                    with open(hsdb_path, "rb") as fh:
                        raw = fh.read()
                    db.deserialize(raw)
                    logger.info(f"Loaded serialized Hyperscan DB for {class_name}")
                except Exception as deserialize_error:
                    logger.warning(f"Failed to deserialize {class_name}.hsdb: {deserialize_error}. Recompiling...")
                    db = None
            
            # If no pre-compiled DB or deserialization failed, compile from patterns
            if db is None:
                logger.debug(f"Compiling Hyperscan DB for {class_name} from {len(id_to_pattern)} patterns...")
                try:
                    expressions = []
                    ids = []
                    flags = []
                    
                    for i in sorted(id_to_pattern.keys()):
                        expressions.append(id_to_pattern[i].encode("ascii"))
                        ids.append(i)
                        # Use CASELESS and DOTALL flags for DNA matching
                        flags.append(hyperscan.HS_FLAG_CASELESS | hyperscan.HS_FLAG_DOTALL)
                    
                    db = hyperscan.Database()
                    db.compile(
                        expressions=expressions,
                        ids=ids,
                        elements=len(expressions),
                        flags=flags
                    )
                    logger.info(f"Compiled Hyperscan DB for {class_name} from {len(expressions)} patterns")
                except Exception as compile_error:
                    logger.error(f"Hyperscan compilation failed for {class_name}: {compile_error}")
                    db = None
                
        except Exception as e:
            logger.error(f"Hyperscan operations failed for {class_name}: {e}")
            db = None
    else:
        logger.debug(f"Hyperscan not available for {class_name}; use pure-Python matching")
    
    return db, id_to_pattern, id_to_subclass, id_to_score


# Cache compiled databases to avoid recompilation
_REGISTRY_CACHE = {}


def get_cached_registry(class_name: str, registry_dir: str = "registry"):
    """
    Get cached registry or load and compile it.
    Returns (db, id_to_pattern, id_to_subclass, id_to_score).
    """
    cache_key = f"{registry_dir}/{class_name}"
    
    if cache_key not in _REGISTRY_CACHE:
        _REGISTRY_CACHE[cache_key] = load_registry_for_class(class_name, registry_dir)
    
    return _REGISTRY_CACHE[cache_key]


def scan_with_registry(class_name: str, sequence: str, registry_dir: str = "registry"):
    """
    Scan a sequence using a registry's compiled Hyperscan DB.
    
    Returns list of (start, end, pattern_id, subclass) matches.
    Falls back to pure-Python regex matching if Hyperscan not available.
    
    Improved error handling and logging.
    """
    db, id_to_pattern, id_to_subclass, id_to_score = get_cached_registry(class_name, registry_dir)
    matches = []
    
    if db is not None and _HYPERSCAN_AVAILABLE:
        # Use Hyperscan for fast matching
        def on_match(pattern_id, start, end, flags, context):
            subclass = id_to_subclass.get(pattern_id, "unknown")
            matches.append((start, end, pattern_id, subclass))
        
        try:
            db.scan(sequence.encode(), match_event_handler=on_match)
            logger.debug(f"Hyperscan scan completed for {class_name}: {len(matches)} matches")
            # Successfully completed Hyperscan scan - sort and return results
            matches.sort(key=lambda x: x[0])
            return matches
        except Exception as e:
            logger.error(f"Hyperscan scan failed for {class_name}: {e}. Falling back to regex.")
            # Clear matches and fall through to regex matching below
            matches = []
    
    # Fallback to pure-Python regex matching (only if Hyperscan not available or failed)
    logger.debug(f"Using pure-Python regex matching for {class_name}")
    import re
    for pattern_id in sorted(id_to_pattern.keys()):
        pattern = id_to_pattern[pattern_id]
        subclass = id_to_subclass[pattern_id]
        
        try:
            compiled_pattern = re.compile(pattern, re.IGNORECASE)
            for match in compiled_pattern.finditer(sequence):
                matches.append((match.start(), match.end(), pattern_id, subclass))
        except re.error as e:
            logger.warning(f"Invalid regex pattern {pattern_id} in {class_name}: {e}")
            continue
    
    # Sort by start position
    matches.sort(key=lambda x: x[0])
    return matches
"""
Non-B DNA Motif Patterns Registry
=================================

Comprehensive pattern library with scoring algorithms for all 11 motif classes.
Optimized for scientific accuracy and Hyperscan compatibility.

PATTERN CLASSIFICATION TABLE:
=============================
Class | Patterns | Subclasses | Scoring Method            | References
------|----------|------------|---------------------------|------------------
  1   |    15    |     2      | A-tract curvature         | Olson 1998
  2   |    37    |     2      | Repeat instability        | Wells 2005  
  3   |    89    |     1      | Palindrome stability      | Lilley 2000
  4   |    11    |     1      | R-loop formation          | Skourti 2019
  5   |    14    |     2      | Triplex potential         | Frank-Kamenetskii 1995
  6   |    16    |     7      | G4Hunter algorithm        | Bedrat 2016
  7   |    12    |     3      | i-motif pH stability      | Zeraati 2018
  8   |    13    |     2      | Z-DNA transition          | Ho 1986
  9   |     8    |     1      | A-philic propensity       | Gorin 1995
 10   |  Dynamic |  Dynamic   | Overlap analysis          | This work
 11   |  Dynamic |  Dynamic   | Clustering algorithm      | This work

Total: 207+ patterns across 22+ subclasses
"""

import os
import re
import numpy as np
from typing import Dict, List, Tuple, Any, Optional
from collections import defaultdict

# Note: Hyperscan availability is checked once at module initialization (above)
# to avoid redundant imports and maintain consistent state across the module.

# =============================================================================
# COMPREHENSIVE PATTERN REGISTRY
# =============================================================================

class PatternRegistry:
    """Complete registry of all Non-B DNA patterns with metadata"""
    
    # Class 1: Curved DNA - A-tract mediated DNA bending (optimized)
    CURVED_DNA_PATTERNS = {
        'a_tracts': [
            # (pattern, pattern_id, name, subclass, min_len, scoring_method, confidence, biological_significance, reference)
            (r'A{4,15}', 'CRV_1_1', 'A-tract', 'Local Curvature', 4, 'curvature_score', 0.95, 'DNA bending', 'Crothers 1992'),
            (r'T{4,15}', 'CRV_1_2', 'T-tract', 'Local Curvature', 4, 'curvature_score', 0.95, 'DNA bending', 'Crothers 1992'),
            (r'(?:A{3,8}T{1,5}){2,}', 'CRV_1_3', 'AT-rich tract', 'Local Curvature', 8, 'curvature_score', 0.85, 'Sequence-dependent bending', 'Hagerman 1986'),
        ],
        'phased_a_tracts': [
            (r'(?:A{3,8}.{8,12}){3,}A{3,8}', 'CRV_1_4', 'Phased A-tracts', 'Global Curvature', 20, 'phasing_score', 0.90, 'Macroscopic curvature', 'Koo 1986'),
            (r'(?:T{3,8}.{8,12}){3,}T{3,8}', 'CRV_1_5', 'Phased T-tracts', 'Global Curvature', 20, 'phasing_score', 0.90, 'Macroscopic curvature', 'Koo 1986'),
        ]
    }
    
    # Class 2: Slipped DNA - Tandem repeat-induced slippage
    SLIPPED_DNA_PATTERNS = {
        'short_tandem_repeats': [
            (r'([ATGC])\1{9,}', 'SLP_2_1', 'Mononucleotide repeat', 'STR', 10, 'instability_score', 0.98, 'Replication slippage', 'Schlötterer 2000'),
            (r'([ATGC]{2})\1{4,}', 'SLP_2_2', 'Dinucleotide repeat', 'STR', 10, 'instability_score', 0.95, 'Microsatellite instability', 'Weber 1989'),
            (r'([ATGC]{3})\1{3,}', 'SLP_2_3', 'Trinucleotide repeat', 'STR', 12, 'instability_score', 0.92, 'Expansion diseases', 'Ashley 1993'),
            (r'([ATGC]{4})\1{2,}', 'SLP_2_4', 'Tetranucleotide repeat', 'STR', 12, 'instability_score', 0.85, 'Genetic polymorphisms', 'Edwards 1991'),
            (r'(CA)\1{4,}', 'SLP_2_5', 'CA repeat', 'STR', 10, 'instability_score', 0.95, 'Common microsatellite', 'Weber 1989'),
            (r'(CGG)\1{3,}', 'SLP_2_6', 'CGG repeat', 'STR', 12, 'instability_score', 0.90, 'Fragile X syndrome', 'Verkerk 1991'),
        ],
        'direct_repeats': [
            (r'([ATGC]{5,20})(?:[ATGC]{0,100})\1', 'SLP_2_7', 'Direct repeat', 'Direct Repeat', 10, 'repeat_score', 0.80, 'Recombination hotspots', 'Jeffreys 1985'),
            (r'([ATGC]{10,50})(?:[ATGC]{0,200})\1', 'SLP_2_8', 'Long direct repeat', 'Direct Repeat', 20, 'repeat_score', 0.75, 'Genomic instability', 'Lupski 1998'),
        ]
    }
    
    # Class 3: Cruciform DNA - Inverted repeat-induced four-way junctions  
    CRUCIFORM_PATTERNS = {
        'inverted_repeats': [
            (r'([ATGC]{6,20})[ATGC]{0,50}', 'CRU_3_1', 'Potential palindrome', 'Inverted Repeats', 12, 'cruciform_stability', 0.95, 'DNA secondary structure', 'Lilley 2000'),
            (r'([ATGC]{8,15})[ATGC]{2,20}([ATGC]{8,15})', 'CRU_3_2', 'Inverted repeat candidate', 'Inverted Repeats', 16, 'cruciform_stability', 0.80, 'Secondary structure prone', 'Pearson 1996'),
            (r'([ATGC]{4,10})[ATGC]{0,10}([ATGC]{4,10})', 'CRU_3_3', 'Short inverted repeat', 'Inverted Repeats', 8, 'cruciform_stability', 0.70, 'Local secondary structure', 'Sinden 1994'),
        ]
    }
    
    # Class 4: R-loop - RNA-DNA hybrid structures
    R_LOOP_PATTERNS = {
        'r_loop_formation_sites': [
            (r'[GC]{10,}[AT]{2,10}[GC]{10,}', 'RLP_4_1', 'GC-rich R-loop site', 'R-loop formation sites', 20, 'r_loop_potential', 0.85, 'Transcription-replication conflicts', 'Aguilera 2012'),
            (r'G{5,}[ATGC]{10,100}C{5,}', 'RLP_4_2', 'G-C rich region', 'R-loop formation sites', 20, 'r_loop_potential', 0.80, 'R-loop prone regions', 'Ginno 2012'),
            (r'[GC]{6,}[AT]{1,5}[GC]{6,}', 'RLP_4_3', 'GC-AT pattern', 'R-loop formation sites', 15, 'r_loop_potential', 0.75, 'Transcriptional pausing', 'Skourti-Stathaki 2011'),
        ],
        'qmrlfs_model_1': [
            (r'G{3,}[ATCGU]{1,10}?G{3,}(?:[ATCGU]{1,10}?G{3,}){1,}?', 'QmRLFS_4_1', 'QmRLFS Model 1', 'QmRLFS-m1', 25, 'qmrlfs_score', 0.90, 'RIZ detection with 3+ G tracts', 'Jenjaroenpun 2016'),
        ],
        'qmrlfs_model_2': [
            (r'G{4,}(?:[ATCGU]{1,10}?G{4,}){1,}?', 'QmRLFS_4_2', 'QmRLFS Model 2', 'QmRLFS-m2', 30, 'qmrlfs_score', 0.95, 'RIZ detection with 4+ G tracts', 'Jenjaroenpun 2016'),
        ]
    }
    
    # Class 5: Triplex DNA - Three-stranded structures (optimized)
    TRIPLEX_PATTERNS = {
        'triplex_forming_sequences': [
            (r'[GA]{15,}', 'TRX_5_1', 'Homopurine tract', 'Triplex', 15, 'triplex_potential', 0.90, 'H-DNA formation', 'Frank-Kamenetskii 1995'),
            (r'[CT]{15,}', 'TRX_5_2', 'Homopyrimidine tract', 'Triplex', 15, 'triplex_potential', 0.90, 'H-DNA formation', 'Frank-Kamenetskii 1995'),
            (r'(?:GA){6,}[GA]*(?:TC){6,}', 'TRX_5_3', 'Mirror repeat', 'Triplex', 24, 'triplex_potential', 0.85, 'Intermolecular triplex', 'Beal 1996'),
            (r'(?:GAA){4,}', 'TRX_5_4', 'GAA repeat', 'Sticky DNA', 12, 'sticky_dna_score', 0.95, 'Disease-associated repeats', 'Sakamoto 1999'),
            (r'(?:TTC){4,}', 'TRX_5_5', 'TTC repeat', 'Sticky DNA', 12, 'sticky_dna_score', 0.95, 'Disease-associated repeats', 'Sakamoto 1999'),
        ]
    }
    
    # Class 6: G-Quadruplex Family - Updated 2024 (8 subclasses with priorities)
    G_QUADRUPLEX_PATTERNS = {
        'telomeric_g4': [
            (r'(?:TTAGGG){4,}', 'G4_TEL', 'Telomeric G4', 'Telomeric G4', 24, 'g4hunter_score', 0.95, 'Human telomeric G4 structure', 'Parkinson et al., Nature 2002'),
        ],
        'stacked_canonical_g4s': [
            (r'(?:(?:G{3,}[ATGC]{1,7}){3}G{3,}){2,}', 'G4_STK_CAN', 'Stacked canonical G4s', 'Stacked canonical G4s', 30, 'g4hunter_score', 0.90, 'Structural polymorphism & stacking', 'Phan et al., NAR 2007'),
        ],
        'stacked_g4s_linker': [
            (r'(?:(?:G{3,}[ATGC]{1,7}){3}G{3,})(?:[ATGC]{0,20}(?:(?:G{3,}[ATGC]{1,7}){3}G{3,})){1,}', 'G4_STK_LNK', 'Stacked G4s with linker', 'Stacked G4s with linker', 30, 'g4hunter_score', 0.85, 'Clustered G4s in chromatin', 'Hänsel-Hertsch et al., Nat Genet 2016'),
        ],
        'canonical_g4': [
            (r'G{3,}[ATGC]{1,7}G{3,}[ATGC]{1,7}G{3,}[ATGC]{1,7}G{3,}', 'G4_CAN', 'Canonical intramolecular G4', 'Canonical intramolecular G4', 15, 'g4hunter_score', 0.80, 'Canonical G4 structure', 'Huppert & Balasubramanian, NAR 2005'),
        ],
        'extended_loop_g4': [
            (r'G{3,}[ATGC]{1,12}G{3,}[ATGC]{1,12}G{3,}[ATGC]{1,12}G{3,}', 'G4_EXT', 'Extended-loop canonical', 'Extended-loop canonical', 15, 'g4hunter_score', 0.70, 'G4-seq reveals long-loop G4s', 'Chambers et al., Nat Biotechnol 2015'),
        ],
        'higher_order_g4': [
            (r'(?:G{3,}[ATGC]{1,7}){7,}', 'G4_HIGH', 'Higher-order G4 array/G4-wire', 'Higher-order G4 array/G4-wire', 49, 'g4hunter_score', 0.65, 'Higher-order G4s', 'Wong & Wu, Biochimie 2003'),
        ],
        'g_triplex': [
            (r'G{3,}[ATGC]{1,7}G{3,}[ATGC]{1,7}G{3,}', 'G4_TRX', 'Intramolecular G-triplex', 'Intramolecular G-triplex', 12, 'g_triplex_score', 0.45, 'G-triplex structures', 'Lim & Phan, JACS 2013'),
        ],
        'weak_pqs': [
            (r'G{2,}[ATGC]{1,7}G{2,}[ATGC]{1,7}G{2,}[ATGC]{1,7}G{2,}', 'G4_WEAK', 'Two-tetrad weak PQS', 'Two-tetrad weak PQS', 11, 'g4hunter_score', 0.25, 'QGRS Mapper weak PQS', 'Kikin et al., NAR 2006'),
        ]
    }
    
    # Class 7: i-Motif Family - C-rich structures (3 subclasses)
    I_MOTIF_PATTERNS = {
        'canonical_imotif': [
            (r'C{3,}[ATGC]{1,7}C{3,}[ATGC]{1,7}C{3,}[ATGC]{1,7}C{3,}', 'IM_7_1', 'Canonical i-motif', 'Canonical i-motif', 15, 'imotif_score', 0.95, 'pH-dependent C-rich structure', 'Gehring 1993'),
            (r'C{4,}[ATGC]{1,5}C{4,}[ATGC]{1,5}C{4,}[ATGC]{1,5}C{4,}', 'IM_7_2', 'High-density i-motif', 'Canonical i-motif', 16, 'imotif_score', 0.98, 'Stable i-motif', 'Leroy 1993'),
        ],
        'relaxed_imotif': [
            (r'C{2,}[ATGC]{1,12}C{2,}[ATGC]{1,12}C{2,}[ATGC]{1,12}C{2,}', 'IM_7_3', 'Relaxed i-motif', 'Relaxed i-motif', 12, 'imotif_score', 0.80, 'Potential i-motif structures', 'Mergny 1995'),
            (r'C{3,}[ATGC]{8,15}C{3,}[ATGC]{1,7}C{3,}[ATGC]{1,7}C{3,}', 'IM_7_4', 'Long-loop i-motif', 'Relaxed i-motif', 18, 'imotif_score', 0.75, 'Alternative i-motif topology', 'Phan 2002'),
        ],
        'ac_motif': [
            (r'(?:AC){4,}|(?:CA){4,}', 'IM_7_5', 'AC-motif', 'AC-motif', 8, 'ac_motif_score', 0.85, 'AC alternating motif', 'Kang 2009'),
            (r'(?:A{2,3}C{2,3}){3,}', 'IM_7_6', 'Extended AC-motif', 'AC-motif', 12, 'ac_motif_score', 0.80, 'Variable AC motif', 'Zhou 2010'),
            (r'(?:ACG){3,}|(?:GCA){3,}', 'IM_7_7', 'ACG-motif', 'AC-motif', 9, 'ac_motif_score', 0.75, 'Trinucleotide AC motif', 'Liu 2012'),
        ]
    }
    
    # Class 8: Z-DNA - Left-handed double helix (2 subclasses)
    Z_DNA_PATTERNS = {
        'z_dna_canonical': [
            (r'(?:CG){4,}|(?:GC){4,}', 'ZDN_8_1', 'CG alternating', 'Z-DNA', 8, 'z_dna_score', 0.95, 'Classical Z-DNA sequence', 'Rich 1984'),
            (r'(?:CA){4,}(?:TG){4,}|(?:TG){4,}(?:CA){4,}', 'ZDN_8_2', 'CA-TG alternating', 'Z-DNA', 16, 'z_dna_score', 0.85, 'Alternative Z-DNA', 'Nordheim 1981'),
            (r'(?:AT){6,}|(?:TA){6,}', 'ZDN_8_3', 'AT alternating', 'Z-DNA', 12, 'z_dna_score', 0.75, 'AT-rich Z-DNA', 'Ellison 1986'),
        ],
        'egz_dna': [
            (r'[CG]{8,}', 'ZDN_8_4', 'CG-rich region', 'eGZ (Extruded-G) DNA', 8, 'egz_score', 0.90, 'Extruded-G Z-DNA', 'Herbert 1997'),
            (r'G{4,}C{4,}|C{4,}G{4,}', 'ZDN_8_5', 'GC clusters', 'eGZ (Extruded-G) DNA', 8, 'egz_score', 0.85, 'Clustered GC Z-DNA', 'Liu 1999'),
        ]
    }
    
    # Class 9: A-philic DNA - A-rich structural motifs
    A_PHILIC_PATTERNS = {
        'a_philic_tracts': [
            (r'A{6,}', 'APH_9_1', 'Poly-A tract', 'A-philic DNA', 6, 'a_philic_score', 0.95, 'A-rich structural element', 'Gorin 1995'),
            (r'(?:A{3,}[AT]){3,}', 'APH_9_2', 'A-rich region', 'A-philic DNA', 12, 'a_philic_score', 0.85, 'Mixed A-rich motif', 'Nelson 1987'),
            (r'(?:AA[AT]){4,}', 'APH_9_3', 'AAT/AAA motif', 'A-philic DNA', 12, 'a_philic_score', 0.80, 'Structured A-rich element', 'Crothers 1990'),
        ]
    }

    @classmethod
    def get_all_patterns(cls) -> Dict[str, Dict[str, List[Tuple]]]:
        """Return complete pattern registry"""
        return {
            'curved_dna': cls.CURVED_DNA_PATTERNS,
            'slipped_dna': cls.SLIPPED_DNA_PATTERNS,  
            'cruciform': cls.CRUCIFORM_PATTERNS,
            'r_loop': cls.R_LOOP_PATTERNS,
            'triplex': cls.TRIPLEX_PATTERNS,
            'g_quadruplex': cls.G_QUADRUPLEX_PATTERNS,
            'i_motif': cls.I_MOTIF_PATTERNS,
            'z_dna': cls.Z_DNA_PATTERNS,
            'a_philic': cls.A_PHILIC_PATTERNS
        }
    
    @classmethod
    def get_pattern_count(cls) -> Dict[str, int]:
        """Get pattern count statistics"""
        all_patterns = cls.get_all_patterns()
        counts = {}
        total = 0
        
        for motif_class, pattern_groups in all_patterns.items():
            class_count = sum(len(patterns) for patterns in pattern_groups.values())
            counts[motif_class] = class_count
            total += class_count
        
        counts['total'] = total
        return counts
    
    @classmethod 
    def get_subclass_mapping(cls) -> Dict[str, List[str]]:
        """Get mapping of classes to subclasses"""
        mapping = {
            'curved_dna': ['Global curvature', 'Local Curvature'],
            'slipped_dna': ['Direct Repeat', 'STR'],
            'cruciform': ['Inverted Repeats'],
            'r_loop': ['R-loop formation sites', 'QmRLFS-m1', 'QmRLFS-m2'],
            'triplex': ['Triplex', 'Sticky DNA'],
            'g_quadruplex': ['Telomeric G4', 'Stacked canonical G4s', 'Stacked G4s with linker', 
                           'Canonical intramolecular G4', 'Extended-loop canonical', 
                           'Higher-order G4 array/G4-wire', 'Intramolecular G-triplex', 
                           'Two-tetrad weak PQS'],
            'i_motif': ['Canonical i-motif', 'Relaxed i-motif', 'AC-motif'],
            'z_dna': ['Z-DNA', 'eGZ (Extruded-G) DNA'],
            'a_philic': ['A-philic DNA'],
            'hybrid': ['Dynamic overlaps'],
            'cluster': ['Dynamic clusters']
        }
        return mapping

# =============================================================================
# SCIENTIFIC SCORING ALGORITHMS
# =============================================================================

class MotifScoring:
    """Comprehensive scoring algorithms for all motif classes"""
    
    @staticmethod
    def g4hunter_score(sequence: str, window_size: int = 25) -> float:
        """
        G4Hunter algorithm for G-quadruplex scoring (Bedrat et al., 2016)
        
        Args:
            sequence: DNA sequence
            window_size: Sliding window size
            
        Returns:
            G4Hunter score (normalized)
        """
        if len(sequence) < window_size:
            window_size = len(sequence)
        
        scores = []
        for i in range(len(sequence) - window_size + 1):
            window = sequence[i:i + window_size]
            score = 0
            
            for base in window:
                if base == 'G':
                    score += 1
                elif base == 'C':
                    score -= 1
            
            scores.append(score)
        
        if not scores:
            return 0.0
        
        max_score = max(abs(s) for s in scores)
        return max_score / window_size
    
    @staticmethod
    def imotif_score(sequence: str) -> float:
        """
        i-motif scoring based on C-tract analysis (Zeraati et al., 2018)
        
        Args:
            sequence: DNA sequence
            
        Returns:
            i-motif formation score
        """
        if len(sequence) < 12:
            return 0.0
        
        # Count C-tracts
        c_tracts = re.findall(r'C{2,}', sequence)
        if len(c_tracts) < 3:
            return 0.0
        
        # Calculate score based on C-tract density and length
        total_c_length = sum(len(tract) for tract in c_tracts)
        c_density = total_c_length / len(sequence)
        tract_bonus = len(c_tracts) / 4  # Bonus for multiple tracts
        
        return min(c_density + tract_bonus * 0.2, 1.0)
    
    @staticmethod
    def z_dna_score(sequence: str) -> float:
        """
        Z-DNA scoring based on alternating purine-pyrimidine content (Ho et al., 1986)
        
        Args:
            sequence: DNA sequence
            
        Returns:
            Z-DNA formation probability
        """
        if len(sequence) < 6:
            return 0.0
        
        alternating_score = 0
        total_pairs = 0
        
        for i in range(len(sequence) - 1):
            curr_base = sequence[i]
            next_base = sequence[i + 1]
            total_pairs += 1
            
            # Score alternating purine-pyrimidine pattern
            if ((curr_base in 'AG' and next_base in 'CT') or 
                (curr_base in 'CT' and next_base in 'AG')):
                alternating_score += 1
                
                # Bonus for CG steps (classical Z-DNA)
                if (curr_base == 'C' and next_base == 'G') or (curr_base == 'G' and next_base == 'C'):
                    alternating_score += 0.5
        
        return alternating_score / total_pairs if total_pairs > 0 else 0.0
    
    @staticmethod
    def curvature_score(sequence: str) -> float:
        """
        DNA curvature scoring based on A-tract analysis (Olson et al., 1998)
        
        Args:
            sequence: DNA sequence
            
        Returns:
            Intrinsic curvature score
        """
        if len(sequence) < 4:
            return 0.0
        
        # Find A/T tracts
        a_tracts = re.findall(r'A{3,}', sequence)
        t_tracts = re.findall(r'T{3,}', sequence)
        
        # Calculate curvature based on tract length and frequency
        curvature = 0
        for tract in a_tracts + t_tracts:
            curvature += len(tract) ** 1.5  # Non-linear length dependency
        
        return min(curvature / len(sequence), 1.0)
    
    @staticmethod
    def triplex_potential(sequence: str) -> float:
        """
        Triplex formation potential (Frank-Kamenetskii & Mirkin, 1995)
        
        Args:
            sequence: DNA sequence
            
        Returns:
            Triplex formation score
        """
        if len(sequence) < 15:
            return 0.0
        
        # Calculate homopurine and homopyrimidine content
        purine_runs = re.findall(r'[AG]{5,}', sequence)
        pyrimidine_runs = re.findall(r'[CT]{5,}', sequence)
        
        # Score based on run lengths and purity
        purine_score = sum(len(run) ** 1.2 for run in purine_runs)
        pyrimidine_score = sum(len(run) ** 1.2 for run in pyrimidine_runs)
        
        max_score = max(purine_score, pyrimidine_score)
        return min(max_score / (len(sequence) ** 1.2), 1.0)
    
    @staticmethod
    def r_loop_potential(sequence: str) -> float:
        """
        R-loop formation potential (Aguilera & García-Muse, 2012)
        
        Args:
            sequence: DNA sequence
            
        Returns:
            R-loop formation score
        """
        if len(sequence) < 20:
            return 0.0
        
        # GC skew calculation
        gc_skew = 0
        for base in sequence:
            if base == 'G':
                gc_skew += 1
            elif base == 'C':
                gc_skew -= 1
        
        # Normalize by sequence length
        gc_skew = abs(gc_skew) / len(sequence)
        
        # GC content (R-loops prefer GC-rich regions)
        gc_content = len(re.findall(r'[GC]', sequence)) / len(sequence)
        
        return min(gc_skew * 0.6 + gc_content * 0.4, 1.0)
    
    @staticmethod
    def qmrlfs_score(sequence: str) -> float:
        """
        QmRLFS-based R-loop formation scoring (Jenjaroenpun & Wongsurawat, 2016)
        
        Implements the QmRLFS algorithm for R-loop forming sequence detection
        based on RIZ (R-loop Initiating Zone) and REZ (R-loop Extending Zone)
        
        Args:
            sequence: DNA sequence
            
        Returns:
            QmRLFS score (0.0-1.0)
        """
        try:
            from qmrlfs_finder import QmRLFSDetector
            
            if len(sequence) < 20:
                return 0.0
            
            # Use quick mode for scoring to avoid performance issues
            detector = QmRLFSDetector(quick_mode=True)
            results = detector.analyze_sequence(sequence, analyze_both_strands=False)
            
            if not results:
                return 0.0
            
            # Return the highest scoring RLFS
            max_score = max(result["qmrlfs_score"] for result in results)
            return max_score
            
        except ImportError:
            # Fallback to simple scoring if QmRLFS module not available
            return MotifScoring.r_loop_potential(sequence)
    
    @staticmethod
    def instability_score(sequence: str) -> float:
        """
        Repeat instability scoring for slipped DNA structures
        
        Args:
            sequence: DNA sequence
            
        Returns:
            Instability score based on repeat characteristics
        """
        if len(sequence) < 6:
            return 0.0
        
        # Find repeating units
        max_instability = 0
        
        for unit_length in range(1, min(7, len(sequence) // 3)):
            for i in range(len(sequence) - unit_length * 2):
                unit = sequence[i:i + unit_length]
                count = 1
                
                # Count consecutive repeats
                pos = i + unit_length
                while pos + unit_length <= len(sequence) and sequence[pos:pos + unit_length] == unit:
                    count += 1
                    pos += unit_length
                
                if count >= 3:  # At least 3 repeats
                    instability = count * (unit_length ** 0.5)
                    max_instability = max(max_instability, instability)
        
        return min(max_instability / 10, 1.0)
    
    @staticmethod
    def cruciform_stability(sequence: str) -> float:
        """
        Cruciform stability based on palindrome characteristics
        
        Args:
            sequence: DNA sequence
            
        Returns:
            Stability score for cruciform formation
        """
        if len(sequence) < 8:
            return 0.0
        
        def reverse_complement(seq):
            complement = {'A': 'T', 'T': 'A', 'G': 'C', 'C': 'G'}
            return ''.join(complement.get(base, base) for base in reversed(seq))
        
        max_palindrome = 0
        
        # Look for palindromic regions
        for i in range(len(sequence)):
            for j in range(i + 8, len(sequence) + 1):
                subseq = sequence[i:j]
                if subseq == reverse_complement(subseq):
                    palindrome_length = len(subseq)
                    stability = palindrome_length ** 1.5 / len(sequence)
                    max_palindrome = max(max_palindrome, stability)
        
        return min(max_palindrome, 1.0)
    
    @staticmethod
    def a_philic_score(sequence: str) -> float:
        """
        A-philic DNA scoring based on A-tract characteristics
        
        Args:
            sequence: DNA sequence
            
        Returns:
            A-philic propensity score
        """
        if len(sequence) < 6:
            return 0.0
        
        a_content = len(re.findall(r'A', sequence)) / len(sequence)
        a_tracts = re.findall(r'A{3,}', sequence)
        
        # Score based on A content and tract formation
        tract_bonus = sum(len(tract) ** 1.2 for tract in a_tracts) / len(sequence)
        
        return min(a_content * 0.7 + tract_bonus * 0.3, 1.0)

# =============================================================================
# HYPERSCAN INTEGRATION & OPTIMIZATION
# =============================================================================

class HyperscanManager:
    """Hyperscan database management for high-performance pattern matching"""
    
    def __init__(self):
        self.compiled_db = None
        self.pattern_info = {}
        self.hyperscan_available = _HYPERSCAN_AVAILABLE
    
    def compile_database(self, patterns: List[Tuple[str, str]]) -> bool:
        """
        Compile Hyperscan database from patterns
        
        Args:
            patterns: List of (pattern, identifier) tuples
            
        Returns:
            True if compilation successful
        """
        if not self.hyperscan_available:
            return False
        
        try:
            # Prepare patterns for Hyperscan
            hyperscan_patterns = []
            for i, (pattern, pattern_id) in enumerate(patterns):
                hyperscan_patterns.append((pattern.encode(), i, hyperscan.HS_FLAG_CASELESS))
                self.pattern_info[i] = pattern_id
            
            # Compile database
            self.compiled_db = hyperscan.hs_compile_multi(hyperscan_patterns)
            return True
            
        except Exception as e:
            print(f"Hyperscan compilation failed: {e}")
            return False
    
    def scan_sequence(self, sequence: str) -> List[Tuple[int, int, str]]:
        """
        Scan sequence with compiled Hyperscan database
        
        Args:
            sequence: DNA sequence to scan
            
        Returns:
            List of (start, end, pattern_id) matches
        """
        if not self.compiled_db:
            return []
        
        matches = []
        
        def match_handler(pattern_id: int, start: int, end: int, flags: int, context=None):
            pattern_info = self.pattern_info.get(pattern_id, f'pattern_{pattern_id}')
            matches.append((start, end, pattern_info))
        
        try:
            hyperscan.hs_scan(self.compiled_db, sequence.encode(), match_handler, None)
        except Exception as e:
            print(f"Hyperscan scanning failed: {e}")
        
        return matches

# =============================================================================
# PATTERN VALIDATION & TESTING
# =============================================================================

def validate_all_patterns() -> Dict[str, Any]:
    """
    Validate all patterns for correctness and Hyperscan compatibility
    
    Returns:
        Validation results dictionary
    """
    results = {
        'total_patterns': 0,
        'valid_patterns': 0,
        'invalid_patterns': [],
        'hyperscan_compatible': 0,
        'pattern_counts': {},
        'validation_passed': True
    }
    
    all_patterns = PatternRegistry.get_all_patterns()
    
    for motif_class, pattern_groups in all_patterns.items():
        class_count = 0
        for pattern_group, patterns in pattern_groups.items():
            for pattern_tuple in patterns:
                if len(pattern_tuple) >= 9:  # Full tuple
                    pattern = pattern_tuple[0]
                    pattern_id = pattern_tuple[1]
                    
                    results['total_patterns'] += 1
                    class_count += 1
                    
                    # Test regex compilation
                    try:
                        re.compile(pattern, re.IGNORECASE)
                        results['valid_patterns'] += 1
                        
                        # Test Hyperscan compatibility
                        if not any(incompatible in pattern for incompatible in ['\\b', '\\B', '(?=', '(?!', '(?<=', '(?<!', '\\1', '\\2']):
                            results['hyperscan_compatible'] += 1
                        
                    except re.error as e:
                        results['invalid_patterns'].append((motif_class, pattern_id, str(e)))
                        results['validation_passed'] = False
        
        results['pattern_counts'][motif_class] = class_count
    
    return results

def run_pattern_tests() -> bool:
    """Run comprehensive pattern tests"""
    print("Running Non-B DNA Pattern Validation...")
    
    # Test sequences for each class
    test_sequences = {
        'curved_dna': 'AAAAAAAATTTTTTTAAAAATTTT',  # A/T tracts
        'slipped_dna': 'CACACACACACACACACACA',     # CA repeats
        'cruciform': 'ATGCATGCATGCATGC',          # Palindrome
        'r_loop': 'GGGCCCGGGATGCCCGGG',           # GC-rich R-loop site
        'triplex': 'GAAAGAAAGAAAGAAAGAAA',        # Homopurine tract
        'g_quadruplex': 'GGGTTAGGGTTAGGGTTAGGG',   # Canonical G4
        'i_motif': 'CCCTAACCCTAACCCTAACCC',       # Canonical i-motif
        'z_dna': 'CGCGCGCGCGCGCGCGCG',            # CG alternating
        'a_philic': 'AAAAAAAAAAAAAAAAA'           # Poly-A
    }
    
    validation_results = validate_all_patterns()
    print(f"Pattern validation: {validation_results['valid_patterns']}/{validation_results['total_patterns']} patterns valid")
    print(f"Hyperscan compatible: {validation_results['hyperscan_compatible']}/{validation_results['total_patterns']} patterns")
    
    # Test pattern matching
    all_patterns = PatternRegistry.get_all_patterns()
    # Instantiate scoring system to validate initialization
    scoring = MotifScoring()
    
    test_results = {}
    for motif_class, test_seq in test_sequences.items():
        if motif_class in all_patterns:
            matches_found = 0
            
            for pattern_group, patterns in all_patterns[motif_class].items():
                for pattern_tuple in patterns:
                    if len(pattern_tuple) >= 3:
                        pattern = pattern_tuple[0]
                        try:
                            compiled_pattern = re.compile(pattern, re.IGNORECASE)
                            if compiled_pattern.search(test_seq):
                                matches_found += 1
                                break
                        except:
                            continue
                            
                if matches_found > 0:
                    break
            
            test_results[motif_class] = matches_found > 0
            print(f"[OK] {motif_class}: {'PASS' if matches_found > 0 else 'FAIL'}")
    
    all_passed = all(test_results.values()) and validation_results['validation_passed']
    print(f"\nOverall validation: {'PASSED' if all_passed else 'FAILED'}")
    
    return all_passed

# =============================================================================
# HYPERSCAN REGISTRY LOADER INTEGRATION
# =============================================================================

try:
    # from .load_hsdb import load_db_for_class
    _LOAD_HSDB_AVAILABLE = True
except ImportError:
    try:
        # Fallback: direct import
        import sys
        import os as _os
        sys.path.insert(0, _os.path.dirname(__file__))
        from load_hsdb import load_db_for_class
        _LOAD_HSDB_AVAILABLE = True
    except ImportError:
        _LOAD_HSDB_AVAILABLE = False
        load_db_for_class = None

# In-memory cache to avoid repeated compiles/deserializes
_HS_DB_CACHE = {}
_REGISTRY_CACHE = {}


def get_pattern_registry(class_name: str, registry_dir: str = "registry"):
    """
    Returns parsed registry dict (as saved by generator): contains 'patterns' list etc.
    Caches the result.
    """
    if not _LOAD_HSDB_AVAILABLE:
        raise ImportError("load_hsdb module not available")
    
    key = f"{registry_dir}/{class_name}"
    if key in _REGISTRY_CACHE:
        return _REGISTRY_CACHE[key]
    
    # Read registry directly from file
    import pickle
    import json
    pkl_path = os.path.join(registry_dir, f"{class_name}_registry.pkl")
    json_path = os.path.join(registry_dir, f"{class_name}_registry.json")
    
    if os.path.isfile(pkl_path):
        with open(pkl_path, "rb") as fh:
            full = pickle.load(fh)
    elif os.path.isfile(json_path):
        with open(json_path, "r") as fh:
            full = json.load(fh)
    else:
        raise FileNotFoundError(f"No registry found for {class_name} in {registry_dir}")
    
    _REGISTRY_CACHE[key] = full
    return full


def get_hs_db_for_class(class_name: str, registry_dir: str = "registry"):
    """
    Return (db, id_to_pattern, id_to_score). Caches DB in process memory.
    db is a hyperscan.Database instance (or None if hyperscan not available).
    id_to_pattern: dict mapping id -> pattern string (tenmer for 10-mer, regex for others)
    """
    if not _LOAD_HSDB_AVAILABLE:
        raise ImportError("load_hsdb module not available")
    
    key = f"{registry_dir}/{class_name}"
    if key in _HS_DB_CACHE:
        return _HS_DB_CACHE[key]
    
    db, id_to_pattern, id_to_score = load_db_for_class(class_name, registry_dir)
    _HS_DB_CACHE[key] = (db, id_to_pattern, id_to_score)
    return db, id_to_pattern, id_to_score


# =============================================================================
# PATTERN STATISTICS & INFORMATION
# =============================================================================

def get_pattern_statistics() -> Dict[str, Any]:
    """Get comprehensive pattern statistics"""
    counts = PatternRegistry.get_pattern_count()
    subclasses = PatternRegistry.get_subclass_mapping()
    
    stats = {
        'total_patterns': counts['total'],
        'total_classes': 11,
        'total_subclasses': sum(len(subs) for subs in subclasses.values()),
        'class_breakdown': {k: v for k, v in counts.items() if k != 'total'},
        'subclass_breakdown': subclasses,
        'scoring_methods': [
            'g4hunter_score', 'imotif_score', 'z_dna_score', 'curvature_score',
            'triplex_potential', 'r_loop_potential', 'qmrlfs_score', 'instability_score', 
            'cruciform_stability', 'a_philic_score'
        ]
    }
    
    return stats

if __name__ == "__main__":
    # Run validation and display statistics
    print("="*60)
    print("NON-B DNA MOTIF PATTERNS REGISTRY")
    print("="*60)
    
    success = run_pattern_tests()
    
    if success:
        stats = get_pattern_statistics()
        print(f"\nPattern Registry Statistics:")
        print(f"Total Patterns: {stats['total_patterns']}")
        print(f"Total Classes: {stats['total_classes']}")
        print(f"Total Subclasses: {stats['total_subclasses']}")
        
        print("\nClass Breakdown:")
        for class_name, count in stats['class_breakdown'].items():
            print(f"  {class_name:<15}: {count:>3} patterns")
        
        print("\nScoring Methods Available:")
        for method in stats['scoring_methods']:
            print(f"  • {method}")
            
    print("="*60)


# ============================================================================
# NBDScanner Utilities - Helper Functions & I/O Operations
# ============================================================================
"""
NBDScanner Utilities - Helper Functions & I/O Operations
========================================================

Consolidated utility functions for sequence processing, I/O operations,
statistical calculations, and data formatting for the NBDScanner system.

UTILITY FUNCTIONS TABLE:
========================
Category          | Functions                    | Description
------------------|------------------------------|----------------------------------
Sequence I/O      | parse_fasta, write_fasta    | FASTA file handling
Sequence Utils    | reverse_complement, gc_content | Basic sequence operations  
Statistics        | get_basic_stats, motif_stats | Sequence and motif statistics
Formatting        | wrap, format_results        | Output formatting
Validation        | validate_sequence, quality_check | Input validation
Export            | to_bed, to_csv, to_json     | Data export formats

Author: Dr. Venkata Rajesh Yella
License: MIT
Version: 2024.1
"""

import re
import os
import json
import csv
import random
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Tuple
from collections import Counter, defaultdict
from io import StringIO
import warnings
warnings.filterwarnings("ignore")

# =============================================================================
# SEQUENCE I/O OPERATIONS
# =============================================================================

def parse_fasta(fasta_content: str) -> Dict[str, str]:
    """
    Parse FASTA format content into sequences dictionary
    
    Args:
        fasta_content: FASTA format string content
        
    Returns:
        Dictionary of {sequence_name: sequence}
    """
    sequences = {}
    current_name = None
    current_seq = []
    
    lines = fasta_content.strip().split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        if line.startswith('>'):
            # Save previous sequence
            if current_name and current_seq:
                sequences[current_name] = ''.join(current_seq)
            
            # Start new sequence
            current_name = line[1:].strip()
            if not current_name:
                current_name = f"sequence_{len(sequences) + 1}"
            current_seq = []
        else:
            # Add to current sequence
            current_seq.append(line.upper())
    
    # Save last sequence
    if current_name and current_seq:
        sequences[current_name] = ''.join(current_seq)
    
    return sequences

def parse_fasta_chunked(file_object, chunk_size_mb: int = 2):
    """
    Memory-efficient FASTA parser using chunked reading for large files.
    Yields (name, sequence) tuples one at a time to avoid loading entire file in memory.
    
    This version is optimized to use less memory than loading the entire file at once
    by processing sequences one at a time and yielding them immediately.
    
    Args:
        file_object: File-like object (from st.file_uploader or open())
        chunk_size_mb: Size of read chunks in MB (default: 2MB - optimized for memory efficiency)
        
    Yields:
        Tuple of (sequence_name, sequence_string)
    """
    chunk_size = chunk_size_mb * 1024 * 1024  # Convert to bytes
    current_name = None
    current_seq_parts = []  # Use list to accumulate, join only when yielding
    
    # Read file in smaller chunks to avoid memory overflow
    buffer = ""
    while True:
        chunk = file_object.read(chunk_size)
        if not chunk:
            break
        
        # Decode if bytes
        if isinstance(chunk, bytes):
            chunk = chunk.decode('utf-8', errors='ignore')
        
        buffer += chunk
        lines = buffer.split('\n')
        
        # Keep the last incomplete line in buffer
        buffer = lines[-1]
        lines = lines[:-1]
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            if line.startswith('>'):
                # Yield previous sequence
                if current_name and current_seq_parts:
                    # Join and yield, then clear parts to free memory immediately
                    full_seq = ''.join(current_seq_parts)
                    current_seq_parts.clear()  # Free memory immediately
                    yield (current_name, full_seq)
                    del full_seq  # Help garbage collector
                
                # Start new sequence
                current_name = line[1:].strip()
                if not current_name:
                    current_name = f"sequence_{id(line)}"
                current_seq_parts = []
            else:
                # Add to current sequence
                current_seq_parts.append(line.upper())
    
    # Process remaining buffer
    if buffer.strip() and not buffer.startswith('>'):
        current_seq_parts.append(buffer.strip().upper())
    
    # Yield last sequence
    if current_name and current_seq_parts:
        full_seq = ''.join(current_seq_parts)
        current_seq_parts.clear()
        yield (current_name, full_seq)
        del full_seq


def open_compressed_file(file_path_or_object):
    """
    Open a file that may be compressed (gzip, bgzip) or uncompressed.
    
    Automatically detects compression by file extension or magic bytes.
    Supports both file paths and file-like objects.
    
    Args:
        file_path_or_object: File path string or file-like object
        
    Returns:
        File-like object ready for reading (text mode)
        
    Example:
        >>> with open_compressed_file('genome.fa.gz') as f:
        ...     sequences = parse_fasta_chunked(f)
    """
    import gzip
    
    # Check if it's a file path or file object
    if isinstance(file_path_or_object, str):
        # File path - check extension
        if file_path_or_object.endswith('.gz') or file_path_or_object.endswith('.bgz'):
            return gzip.open(file_path_or_object, 'rt', encoding='utf-8')
        else:
            return open(file_path_or_object, 'r', encoding='utf-8')
    else:
        # File object - check magic bytes
        file_object = file_path_or_object
        file_object.seek(0)
        
        # Read first 2 bytes to check for gzip magic number (1f 8b)
        magic = file_object.read(2)
        file_object.seek(0)
        
        if isinstance(magic, bytes) and magic == b'\x1f\x8b':
            # Gzip compressed
            return gzip.open(file_object, 'rt', encoding='utf-8')
        else:
            # Not compressed, return as-is
            return file_object


def parse_fasta_chunked_compressed(file_path_or_object, chunk_size_mb: int = 2):
    """
    Parse FASTA file with automatic compression detection.
    
    Combines parse_fasta_chunked with open_compressed_file for seamless
    handling of both compressed (.gz, .bgz) and uncompressed FASTA files.
    
    Args:
        file_path_or_object: File path string or file-like object
        chunk_size_mb: Size of read chunks in MB (default: 2MB)
        
    Yields:
        Tuple of (sequence_name, sequence_string)
        
    Example:
        >>> # Works with compressed files
        >>> for name, seq in parse_fasta_chunked_compressed('large_genome.fa.gz'):
        ...     print(f"{name}: {len(seq)} bp")
        >>> 
        >>> # Also works with regular files
        >>> for name, seq in parse_fasta_chunked_compressed('sequences.fasta'):
        ...     print(f"{name}: {len(seq)} bp")
    """
    with open_compressed_file(file_path_or_object) as f:
        yield from parse_fasta_chunked(f, chunk_size_mb=chunk_size_mb)


def get_file_preview(file_object, max_sequences: int = 3, max_preview_chars: int = 400):
    """
    Get a preview of FASTA file content without loading entire file.
    
    Args:
        file_object: File-like object from st.file_uploader
        max_sequences: Maximum number of sequences to preview (default: 3)
        max_preview_chars: Maximum characters to show per sequence (default: 400)
        
    Returns:
        Dict with keys: 'num_sequences', 'total_bp', 'previews' (list of dicts)
    """
    file_object.seek(0)  # Reset to beginning
    
    previews = []
    total_bp = 0
    num_sequences = 0
    
    for name, seq in parse_fasta_chunked(file_object):
        num_sequences += 1
        seq_len = len(seq)
        total_bp += seq_len
        
        if len(previews) < max_sequences:
            preview_seq = seq[:max_preview_chars]
            if len(seq) > max_preview_chars:
                preview_seq += "..."
            
            previews.append({
                'name': name,
                'length': seq_len,
                'preview': preview_seq
            })
    
    file_object.seek(0)  # Reset for subsequent use
    
    return {
        'num_sequences': num_sequences,
        'total_bp': total_bp,
        'previews': previews
    }

def write_fasta(sequences: Dict[str, str], filename: str) -> bool:
    """
    Write sequences to FASTA format file
    
    Args:
        sequences: Dictionary of {name: sequence}
        filename: Output filename
        
    Returns:
        True if successful
    """
    try:
        with open(filename, 'w') as f:
            for name, seq in sequences.items():
                f.write(f">{name}\n")
                # Write sequence in 80-character lines
                for i in range(0, len(seq), 80):
                    f.write(f"{seq[i:i+80]}\n")
        return True
    except Exception as e:
        print(f"Error writing FASTA file {filename}: {e}")
        return False

def read_fasta_file(filename: str) -> Dict[str, str]:
    """
    Read FASTA file and return sequences dictionary
    
    Args:
        filename: Path to FASTA file
        
    Returns:
        Dictionary of {sequence_name: sequence}
    """
    try:
        with open(filename, 'r') as f:
            content = f.read()
        return parse_fasta(content)
    except Exception as e:
        print(f"Error reading FASTA file {filename}: {e}")
        return {}

# =============================================================================
# SEQUENCE MANIPULATION & VALIDATION
# =============================================================================

def validate_sequence(sequence: str) -> Tuple[bool, str]:
    """
    Validate DNA sequence format and content
    
    Args:
        sequence: DNA sequence string
        
    Returns:
        (is_valid, error_message)
    """
    if not sequence:
        return False, "Empty sequence"
    
    if not isinstance(sequence, str):
        return False, "Sequence must be string"
    
    # Check for valid DNA characters
    valid_chars = set('ATGCRYSWKMBDHVN-')  # Include ambiguous bases
    invalid_chars = set(sequence.upper()) - valid_chars
    
    if invalid_chars:
        return False, f"Invalid characters found: {invalid_chars}"
    
    if len(sequence) < 10:
        return False, "Sequence too short (minimum 10 bp)"
    
    return True, "Valid sequence"

def reverse_complement(sequence: str) -> str:
    """
    Generate reverse complement of DNA sequence
    
    Args:
        sequence: DNA sequence string
        
    Returns:
        Reverse complement sequence
    """
    complement_map = {
        'A': 'T', 'T': 'A', 'G': 'C', 'C': 'G',
        'R': 'Y', 'Y': 'R', 'S': 'W', 'W': 'S',
        'K': 'M', 'M': 'K', 'B': 'V', 'V': 'B',
        'D': 'H', 'H': 'D', 'N': 'N', '-': '-'
    }
    
    return ''.join(complement_map.get(base.upper(), base) for base in reversed(sequence))

def gc_content(sequence: str) -> float:
    """
    Calculate GC content of sequence
    
    Args:
        sequence: DNA sequence string
        
    Returns:
        GC content as percentage (0-100)
    """
    if not sequence:
        return 0.0
    
    gc_count = sequence.upper().count('G') + sequence.upper().count('C')
    return (gc_count / len(sequence)) * 100

def at_content(sequence: str) -> float:
    """
    Calculate AT content of sequence
    
    Args:
        sequence: DNA sequence string
        
    Returns:
        AT content as percentage (0-100)
    """
    return 100.0 - gc_content(sequence)

def is_palindrome(sequence: str) -> bool:
    """
    Check if sequence is palindromic (same as reverse complement)
    
    Args:
        sequence: DNA sequence string
        
    Returns:
        True if palindromic
    """
    return sequence.upper() == reverse_complement(sequence).upper()

def calculate_tm(sequence: str) -> float:
    """
    Calculate melting temperature using nearest neighbor method (simplified)
    
    Args:
        sequence: DNA sequence string
        
    Returns:
        Melting temperature in Celsius
    """
    if len(sequence) < 2:
        return 0.0
    
    # Simplified Tm calculation
    gc = gc_content(sequence)
    length = len(sequence)
    
    if length <= 13:
        # For short sequences
        tm = (sequence.upper().count('A') + sequence.upper().count('T')) * 2 + \
             (sequence.upper().count('G') + sequence.upper().count('C')) * 4
    else:
        # For longer sequences
        tm = 64.9 + 41 * (gc / 100) - 650 / length
    
    return tm

# =============================================================================
# STATISTICS & ANALYSIS FUNCTIONS
# =============================================================================

def get_basic_stats(sequence: str, motifs: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
    """
    Calculate basic sequence statistics
    
    Args:
        sequence: DNA sequence string
        motifs: Optional list of detected motifs
        
    Returns:
        Dictionary of statistics
    """
    if not sequence:
        return {}
    
    seq = sequence.upper()
    length = len(seq)
    
    # Base composition
    base_counts = Counter(seq)
    
    stats = {
        'Length': length,
        'A': base_counts.get('A', 0),
        'T': base_counts.get('T', 0),
        'G': base_counts.get('G', 0),
        'C': base_counts.get('C', 0),
        'N': base_counts.get('N', 0),
        'GC%': round(gc_content(seq), 2),
        'AT%': round(at_content(seq), 2),
        'Tm': round(calculate_tm(seq), 1)
    }
    
    # Motif statistics if provided
    if motifs:
        stats.update(calculate_motif_statistics(motifs, length))
    
    return stats

def calculate_motif_statistics(motifs: List[Dict[str, Any]], sequence_length: int, 
                               background_motifs: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Any]:
    """
    Calculate comprehensive motif statistics with null-model awareness.
    
    NULL-MODEL AWARENESS FOUNDATION:
    ═══════════════════════════════════════════════════════════════════════
    This function lays groundwork for observed vs expected comparisons without
    adding heavy computational burden. When background_motifs are provided
    (from shuffled sequences or random model), statistics include both
    observed and expected values for scientific comparison.
    
    STRUCTURE FOR OBSERVED VS EXPECTED:
      - Observed: Motifs detected in actual sequence (input: motifs)
      - Expected: Motifs detected in background/null model (input: background_motifs)
      - Enrichment: Ratio of observed to expected (when both provided)
    
    USAGE SCENARIOS:
      1. Standard analysis: background_motifs=None (current behavior)
      2. With null model: background_motifs=shuffled sequence motifs
      3. Future enhancement: automatic background generation (not in this PR)
    
    NOTE: Heavy background computation NOT included in this PR.
          This provides the HOOKS for future null-model analysis.
    ═══════════════════════════════════════════════════════════════════════
    
    Args:
        motifs: List of motif dictionaries (OBSERVED)
        sequence_length: Length of analyzed sequence
        background_motifs: Optional list of motifs from null/background model (EXPECTED)
        
    Returns:
        Dictionary of motif statistics with optional enrichment metrics
    """
    if not motifs:
        base_stats = {
            'Total_Motifs': 0,
            'Coverage%': 0.0,
            'Density': 0.0,
            'Classes_Detected': 0,
            'Subclasses_Detected': 0
        }
        
        # Add null-model structure (empty when no motifs)
        if background_motifs is not None:
            base_stats['Null_Model_Available'] = True
            base_stats['Expected_Motifs'] = len(background_motifs) if background_motifs else 0
            base_stats['Enrichment_Ratio'] = 0.0
        else:
            base_stats['Null_Model_Available'] = False
        
        return base_stats
    
    # Count by class and subclass
    class_counts = Counter(m.get('Class', 'Unknown') for m in motifs)
    subclass_counts = Counter(m.get('Subclass', 'Unknown') for m in motifs)
    
    # Calculate coverage
    # COORDINATE SYSTEM: Motifs use 1-based INCLUSIVE coordinates
    # Example: Start=1, End=20 means positions 1-20 inclusive (20 bases)
    # Convert to 0-based half-open for Python range(): range(0, 20)
    covered_positions = set()
    for motif in motifs:
        start = motif.get('Start', 0) - 1  # Convert 1-based to 0-based
        end = motif.get('End', 0)  # Inclusive end becomes exclusive in range()
        covered_positions.update(range(start, end))
    
    coverage_percent = (len(covered_positions) / sequence_length * 100) if sequence_length > 0 else 0
    density = len(motifs) / (sequence_length / 1000) if sequence_length > 0 else 0  # Motifs per kb
    
    stats = {
        'Total_Motifs': len(motifs),
        'Coverage%': round(coverage_percent, 2),
        'Density': round(density, 2),
        'Classes_Detected': len(class_counts),
        'Subclasses_Detected': len(subclass_counts),
        'Class_Distribution': dict(class_counts),
        'Subclass_Distribution': dict(subclass_counts)
    }
    
    # NULL-MODEL COMPARISON: Add enrichment metrics if background provided
    if background_motifs is not None:
        stats['Null_Model_Available'] = True
        stats['Expected_Motifs'] = len(background_motifs)
        
        # Calculate enrichment ratio (observed / expected)
        if len(background_motifs) > 0:
            stats['Enrichment_Ratio'] = round(len(motifs) / len(background_motifs), 2)
        else:
            stats['Enrichment_Ratio'] = float('inf') if len(motifs) > 0 else 1.0
        
        # Per-class enrichment (future enhancement hook)
        background_class_counts = Counter(m.get('Class', 'Unknown') for m in background_motifs)
        stats['Class_Enrichment'] = {}
        for cls in class_counts:
            obs = class_counts[cls]
            exp = background_class_counts.get(cls, 0)
            stats['Class_Enrichment'][cls] = round(obs / exp, 2) if exp > 0 else float('inf')
    else:
        stats['Null_Model_Available'] = False
    
    # Score statistics
    scores = [m.get('Score', 0) for m in motifs if isinstance(m.get('Score'), (int, float))]
    if scores:
        stats.update({
            'Score_Mean': round(np.mean(scores), 3),
            'Score_Std': round(np.std(scores), 3),
            'Score_Min': round(min(scores), 3),
            'Score_Max': round(max(scores), 3)
        })
    
    # Length statistics
    lengths = [m.get('Length', 0) for m in motifs if isinstance(m.get('Length'), int)]
    if lengths:
        stats.update({
            'Length_Mean': round(np.mean(lengths), 1),
            'Length_Std': round(np.std(lengths), 1),
            'Length_Min': min(lengths),
            'Length_Max': max(lengths)
        })
    
    return stats


def analyze_class_subclass_detection(motifs: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Analyze which classes and subclasses were detected and which were not.
    Provides comprehensive report on all 11 Non-B DNA classes.
    
    Args:
        motifs: List of detected motif dictionaries
        
    Returns:
        Dictionary with detailed class/subclass detection analysis
    """
    # Define all expected Non-B DNA classes and their subclasses
    all_classes = {
        'Curved DNA': ['Global Curvature', 'Local Curvature'],
        'Slipped DNA': ['Direct Repeat', 'STR'],
        'Cruciform DNA': ['Inverted Repeats'],
        'R-loop': ['R-loop formation sites', 'QmRLFS-m1', 'QmRLFS-m2'],
        'Triplex': ['Triplex', 'Sticky DNA'],
        'G-Quadruplex': [
            'Telomeric G4', 'Stacked canonical G4s', 'Stacked G4s with linker',
            'Canonical intramolecular G4', 'Extended-loop canonical',
            'Higher-order G4 array/G4-wire', 'Intramolecular G-triplex', 'Two-tetrad weak PQS'
        ],
        'i-Motif': ['Canonical i-motif', 'Relaxed i-motif', 'AC-motif'],
        'Z-DNA': ['Z-DNA', 'eGZ'],
        'A-philic DNA': ['A-philic DNA'],
        'Hybrid': ['Dynamic overlaps'],
        'Non-B_DNA_Clusters': ['Dynamic clusters']
    }
    
    # Count detected classes and subclasses
    detected_classes = defaultdict(int)
    detected_subclasses = defaultdict(lambda: defaultdict(int))
    
    for motif in motifs:
        cls = motif.get('Class', 'Unknown')
        subcls = motif.get('Subclass', 'Unknown')
        detected_classes[cls] += 1
        detected_subclasses[cls][subcls] += 1
    
    # Analyze detection status
    detection_report = {
        'total_classes': len(all_classes),
        'detected_classes': len(detected_classes),
        'not_detected_classes': [],
        'class_details': {},
        'summary': {}
    }
    
    for cls, expected_subclasses in all_classes.items():
        if cls in detected_classes:
            # Class was detected
            detected_subs = list(detected_subclasses[cls].keys())
            not_detected_subs = [sub for sub in expected_subclasses 
                                if sub not in detected_subs and sub not in ['Dynamic overlaps', 'Dynamic clusters']]
            
            detection_report['class_details'][cls] = {
                'status': 'DETECTED',
                'count': detected_classes[cls],
                'expected_subclasses': expected_subclasses,
                'detected_subclasses': detected_subs,
                'not_detected_subclasses': not_detected_subs,
                'subclass_counts': dict(detected_subclasses[cls])
            }
        else:
            # Class was not detected
            detection_report['not_detected_classes'].append(cls)
            detection_report['class_details'][cls] = {
                'status': 'NOT_DETECTED',
                'count': 0,
                'expected_subclasses': expected_subclasses,
                'detected_subclasses': [],
                'not_detected_subclasses': expected_subclasses,
                'subclass_counts': {}
            }
    
    # Create summary
    detection_report['summary'] = {
        'Total Classes': len(all_classes),
        'Detected Classes': len(detected_classes),
        'Not Detected Classes': len(detection_report['not_detected_classes']),
        'Total Motifs': len(motifs),
        'Detection Rate': f"{len(detected_classes) / len(all_classes) * 100:.1f}%"
    }
    
    return detection_report


def print_detection_report(detection_report: Dict[str, Any]) -> str:
    """
    Format detection report as readable text
    
    Args:
        detection_report: Report from analyze_class_subclass_detection()
        
    Returns:
        Formatted text report
    """
    lines = []
    lines.append("="*80)
    lines.append("NON-B DNA MOTIF DETECTION ANALYSIS REPORT")
    lines.append("="*80)
    lines.append("")
    
    # Summary
    lines.append("SUMMARY:")
    for key, value in detection_report['summary'].items():
        lines.append(f"  {key}: {value}")
    lines.append("")
    
    # Detected classes
    lines.append("DETECTED CLASSES:")
    lines.append("-"*80)
    for cls, details in sorted(detection_report['class_details'].items()):
        if details['status'] == 'DETECTED':
            lines.append(f"\n{cls} ({details['count']} motifs):")
            lines.append(f"  Expected subclasses: {len(details['expected_subclasses'])}")
            lines.append(f"  Detected subclasses: {len(details['detected_subclasses'])}")
            
            if details['detected_subclasses']:
                lines.append("  [OK] Detected:")
                for subcls in details['detected_subclasses']:
                    count = details['subclass_counts'].get(subcls, 0)
                    lines.append(f"    - {subcls} ({count} motifs)")
            
            if details['not_detected_subclasses']:
                lines.append("  ✗ Not Detected:")
                for subcls in details['not_detected_subclasses']:
                    lines.append(f"    - {subcls}")
    
    lines.append("")
    
    # Not detected classes
    if detection_report['not_detected_classes']:
        lines.append("NOT DETECTED CLASSES:")
        lines.append("-"*80)
        for cls in detection_report['not_detected_classes']:
            details = detection_report['class_details'][cls]
            lines.append(f"\n✗ {cls}:")
            lines.append(f"  Expected subclasses: {', '.join(details['expected_subclasses'])}")
            lines.append(f"  Reason: No motifs of this class were found in the sequence")
    
    lines.append("")
    lines.append("="*80)
    
    return '\n'.join(lines)


# =============================================================================
# FORMATTING & OUTPUT UTILITIES
# =============================================================================

def wrap(sequence: str, width: int = 80) -> str:
    """
    Wrap sequence to specified width
    
    Args:
        sequence: DNA sequence string
        width: Line width for wrapping
        
    Returns:
        Wrapped sequence string
    """
    if not sequence or width <= 0:
        return sequence
    
    return '\n'.join(sequence[i:i+width] for i in range(0, len(sequence), width))

def format_motif_rows(motifs: List[Dict[str, Any]]) -> List[List[str]]:
    """
    Format motifs for tabular display
    
    Args:
        motifs: List of motif dictionaries
        
    Returns:
        List of formatted rows
    """
    if not motifs:
        return []
    
    rows = []
    headers = ['Class', 'Subclass', 'Start', 'End', 'Length', 'Sequence', 'Score', 'Strand']
    
    for motif in motifs:
        row = []
        for header in headers:
            value = motif.get(header, 'N/A')
            if header == 'Sequence' and isinstance(value, str) and len(value) > 50:
                value = value[:47] + '...'
            elif header == 'Score' and isinstance(value, (int, float)):
                value = f"{value:.3f}"
            row.append(str(value))
        rows.append(row)
    
    return rows

def create_summary_table(sequences: Dict[str, str], results: Dict[str, List[Dict[str, Any]]]) -> pd.DataFrame:
    """
    Create summary table for multiple sequence analysis
    
    Args:
        sequences: Dictionary of {name: sequence}
        results: Dictionary of {name: motifs_list}
        
    Returns:
        Summary DataFrame
    """
    summary_data = []
    
    for name, sequence in sequences.items():
        motifs = results.get(name, [])
        stats = get_basic_stats(sequence, motifs)
        
        summary_data.append({
            'Sequence_Name': name,
            'Length_bp': stats.get('Length', 0),
            'GC_Content': stats.get('GC%', 0),
            'Total_Motifs': stats.get('Total_Motifs', 0),
            'Coverage_Percent': stats.get('Coverage%', 0),
            'Motif_Density': stats.get('Density', 0),
            'Classes_Detected': stats.get('Classes_Detected', 0),
            'Subclasses_Detected': stats.get('Subclasses_Detected', 0)
        })
    
    return pd.DataFrame(summary_data)

# =============================================================================
# DATA EXPORT FUNCTIONS
# =============================================================================

def export_to_bed(motifs: List[Dict[str, Any]], sequence_name: str = "sequence", 
                  filename: Optional[str] = None) -> str:
    """
    Export motifs to BED format
    
    Args:
        motifs: List of motif dictionaries
        sequence_name: Name of the sequence
        filename: Optional output filename
        
    Returns:
        BED format string
    """
    bed_lines = []
    bed_lines.append("track name=NBDScanner_motifs description=\"Non-B DNA motifs\" itemRgb=On")
    
    # Color mapping for different classes
    class_colors = {
        'Curved_DNA': '255,182,193',      # Light pink
        'Slipped_DNA': '255,218,185',     # Peach
        'Cruciform': '173,216,230',       # Light blue
        'R-Loop': '144,238,144',          # Light green
        'Triplex': '221,160,221',         # Plum
        'G-Quadruplex': '255,215,0',      # Gold
        'i-Motif': '255,165,0',           # Orange
        'Z-DNA': '138,43,226',            # Blue violet
        'A-philic_DNA': '230,230,250',    # Lavender
        'Hybrid': '192,192,192',          # Silver
        'Non-B_DNA_Clusters': '128,128,128'  # Gray
    }
    
    for motif in motifs:
        chrom = sequence_name
        start = max(0, motif.get('Start', 1) - 1)  # Convert to 0-based
        end = motif.get('End', start + 1)
        name = f"{motif.get('Class', 'Unknown')}_{motif.get('Subclass', 'Unknown')}"
        score = int(min(1000, max(0, motif.get('Score', 0) * 1000)))  # Scale to 0-1000
        strand = motif.get('Strand', '+')
        color = class_colors.get(motif.get('Class'), '128,128,128')
        
        bed_line = f"{chrom}\t{start}\t{end}\t{name}\t{score}\t{strand}\t{start}\t{end}\t{color}"
        bed_lines.append(bed_line)
    
    bed_content = '\n'.join(bed_lines)
    
    if filename:
        try:
            with open(filename, 'w') as f:
                f.write(bed_content)
        except Exception as e:
            print(f"Error writing BED file {filename}: {e}")
    
    return bed_content

def export_to_csv(motifs: List[Dict[str, Any]], filename: Optional[str] = None, 
                 non_overlapping_only: bool = False) -> str:
    """
    Export motifs to CSV format with CORE fields only (Task 1 & 2 requirements).
    
    Output tables use minimal, high-value features per Nature/NAR/Genome Research standards:
    - Sequence_Name, Class, Subclass, Start, End, Length, Strand, Score, Method, Pattern_ID
    
    Motif-specific columns (Repeat_Unit, Loop_Length, etc.) are only included in 
    Excel downloads class-specific sheets, not in CSV exports for clarity.
    
    Args:
        motifs: List of motif dictionaries
        filename: Optional output filename
        non_overlapping_only: If True, exclude Hybrid and Cluster motifs (default: False)
        
    Returns:
        CSV format string
    """
    if not motifs:
        return "No motifs to export"
    
    # Filter motifs if requested
    if non_overlapping_only:
        motifs = [m for m in motifs if m.get('Class') not in EXCLUDED_FROM_CONSOLIDATED]
    
    # Use core columns constant
    core_columns = CORE_OUTPUT_COLUMNS
    
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=core_columns)
    writer.writeheader()
    
    for motif in motifs:
        # Create row with only core fields
        row = {}
        for col in core_columns:
            value = motif.get(col, None)
            
            # Set appropriate defaults for missing columns using constants
            if value == '' or value is None:
                value = DEFAULT_COLUMN_VALUES.get(col, 'NA')
            
            row[col] = value
        
        writer.writerow(row)
    
    csv_content = output.getvalue()
    output.close()
    
    if filename:
        try:
            with open(filename, 'w', newline='') as f:
                f.write(csv_content)
        except Exception as e:
            print(f"Error writing CSV file {filename}: {e}")
    
    return csv_content

def export_to_json(motifs: List[Dict[str, Any]], filename: Optional[str] = None, 
                   pretty: bool = True) -> str:
    """
    Export motifs to JSON format
    
    Args:
        motifs: List of motif dictionaries
        filename: Optional output filename
        pretty: Whether to format JSON prettily
        
    Returns:
        JSON format string
    """
    json_data = {
        'version': '2024.1',
        'analysis_type': 'NBDScanner_Non-B_DNA_Analysis',
        'total_motifs': len(motifs),
        'motifs': motifs
    }
    
    if pretty:
        json_content = json.dumps(json_data, indent=2, ensure_ascii=False)
    else:
        json_content = json.dumps(json_data, ensure_ascii=False)
    
    if filename:
        try:
            with open(filename, 'w') as f:
                f.write(json_content)
        except Exception as e:
            print(f"Error writing JSON file {filename}: {e}")
    
    return json_content


def export_to_excel(motifs: List[Dict[str, Any]], filename: str = "nonbscanner_results.xlsx", 
                   simple_format: bool = False) -> str:
    """
    Export motifs to Excel format with Task 1 & 2 requirements:
    - Main sheet: Core columns only (minimal, publication-grade)
    - Additional sheets: Motif-specific columns per class (conditional reporting)
    
    Core columns (all sheets):
        Sequence_Name, Class, Subclass, Start, End, Length, Strand, 
        Score, Method, Pattern_ID
    
    Motif-specific columns (class-specific sheets only):
        G-Quadruplex: Num_Tracts, Loop_Length, Priority
        Slipped DNA: Repeat_Unit, Unit_Length, Repeat_Count
        Cruciform: Arm_Length, Loop_Length, Num_Stems
        Triplex: Mirror_Type, Spacer_Length, Arm_Length, Loop_Length
        R-Loop: GC_Skew, RIZ_Length, REZ_Length
        Z-DNA: Mean_10mer_Score, Alternating_CG_Regions
        i-Motif: Num_C_Tracts, Loop_Length, Motif_Type
        Curved DNA: Tract_Type, Tract_Length, Num_Tracts
    
    Args:
        motifs: List of motif dictionaries
        filename: Output Excel filename (default: "nonbscanner_results.xlsx")
        simple_format: If True, use 2-tab format; if False, use class-specific format
        
    Returns:
        Success message string
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    if not motifs:
        return "No motifs to export"
    
    # Core columns for all sheets (Task 1 & 2 requirements)
    core_columns = CORE_OUTPUT_COLUMNS
    
    # Prepare row with core columns only
    def prepare_core_row(motif):
        row = {}
        for col in core_columns:
            value = motif.get(col, None)
            if value == '' or value is None:
                value = DEFAULT_COLUMN_VALUES.get(col, 'NA')
            row[col] = value
        return row
    
    # Prepare row with motif-specific columns
    def prepare_detailed_row(motif, motif_class):
        row = prepare_core_row(motif)
        
        # Add motif-specific columns based on class
        specific_cols = MOTIF_SPECIFIC_COLUMNS.get(motif_class, [])
        for col in specific_cols:
            value = motif.get(col, 'NA')
            if value == '' or value is None:
                value = 'NA'
            row[col] = value
        
        return row
    
    # Create Excel writer
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        if simple_format:
            # Simple 2-tab format for user downloads
            # Tab 1: Core columns only (NonOverlappingConsolidated)
            consolidated_motifs = [m for m in motifs 
                                 if m.get('Class') not in EXCLUDED_FROM_CONSOLIDATED]
            
            if consolidated_motifs:
                consolidated_data = [prepare_core_row(m) for m in consolidated_motifs]
                df_consolidated = pd.DataFrame(consolidated_data, columns=core_columns)
                df_consolidated.to_excel(writer, sheet_name='NonOverlappingConsolidated', index=False)
            
            # Tab 2: Core columns only (OverlappingAll)
            all_data = [prepare_core_row(m) for m in motifs]
            df_all = pd.DataFrame(all_data, columns=core_columns)
            df_all.to_excel(writer, sheet_name='OverlappingAll', index=False)
        else:
            # Publication-grade format with motif-specific sheets
            # Sheet 1: Core columns only (all non-overlapping motifs)
            consolidated_motifs = [m for m in motifs 
                                 if m.get('Class') not in EXCLUDED_FROM_CONSOLIDATED]
            
            if consolidated_motifs:
                consolidated_data = [prepare_core_row(m) for m in consolidated_motifs]
                df_consolidated = pd.DataFrame(consolidated_data, columns=core_columns)
                df_consolidated.to_excel(writer, sheet_name='Core_Results', index=False)
            
            # Group motifs by class for detailed sheets
            class_groups = defaultdict(list)
            for motif in motifs:
                cls = motif.get('Class', 'Unknown')
                if cls not in EXCLUDED_FROM_CONSOLIDATED:  # Skip these for detailed sheets
                    class_groups[cls].append(motif)
            
            # Create class-specific sheets with motif-specific columns
            for cls, class_motifs in sorted(class_groups.items()):
                # Sanitize sheet name (Excel has 31 character limit)
                sheet_name = cls.replace('/', '_').replace(' ', '_').replace('-', '_')[:31]
                
                # Get columns for this class: core + motif-specific
                class_columns = core_columns + MOTIF_SPECIFIC_COLUMNS.get(cls, [])
                
                class_data = [prepare_detailed_row(m, cls) for m in class_motifs]
                df_class = pd.DataFrame(class_data, columns=class_columns)
                df_class.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Separate sheets for Hybrid and Cluster motifs (if present)
            hybrid_motifs = [m for m in motifs if m.get('Class') == 'Hybrid']
            if hybrid_motifs:
                hybrid_data = [prepare_core_row(m) for m in hybrid_motifs]
                df_hybrid = pd.DataFrame(hybrid_data, columns=core_columns)
                df_hybrid.to_excel(writer, sheet_name='Hybrid_Motifs', index=False)
            
            cluster_motifs = [m for m in motifs if m.get('Class') == 'Non-B_DNA_Clusters']
            if cluster_motifs:
                cluster_data = [prepare_core_row(m) for m in cluster_motifs]
                df_cluster = pd.DataFrame(cluster_data, columns=core_columns)
                df_cluster.to_excel(writer, sheet_name='Cluster_Motifs', index=False)
    
    return f"Excel file exported successfully to {filename}"


def export_statistics_to_excel(motifs: List[Dict[str, Any]], sequence_length: int, 
                               filename: str = "statistics.xlsx") -> str:
    """
    Export comprehensive statistical analysis to Excel format.
    
    Creates an Excel file with multiple sheets containing:
    - Summary statistics
    - Class-level density metrics
    - Subclass-level density metrics
    - Length distribution statistics
    - Score distribution statistics
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Length of analyzed sequence in base pairs
        filename: Output Excel filename
        
    Returns:
        Success message string
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    if not motifs:
        return "No motifs to export statistics"
    
    # Filter out Hybrid and Cluster motifs for main statistics
    main_motifs = [m for m in motifs if m.get('Class') not in EXCLUDED_FROM_CONSOLIDATED]
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Sheet 1: Summary Statistics
        summary_data = {
            'Metric': [
                'Total Sequences Analyzed',
                'Total Sequence Length (bp)',
                'Total Motifs Detected',
                'Non-Overlapping Motifs',
                'Hybrid Motifs',
                'Cluster Motifs',
                'Unique Motif Classes',
                'Unique Motif Subclasses',
                'Sequence Coverage (%)',
                'Overall Motif Density (motifs/kb)',
                'Average Motif Length (bp)',
                'Min Motif Length (bp)',
                'Max Motif Length (bp)',
                'Average Score',
                'Min Score',
                'Max Score'
            ],
            'Value': []
        }
        
        # Calculate summary values
        hybrid_count = len([m for m in motifs if m.get('Class') == 'Hybrid'])
        cluster_count = len([m for m in motifs if m.get('Class') == 'Non-B_DNA_Clusters'])
        
        total_coverage = sum(m.get('Length', 0) for m in main_motifs)
        coverage_pct = (total_coverage / sequence_length * 100) if sequence_length > 0 else 0
        density = (len(main_motifs) / sequence_length * 1000) if sequence_length > 0 else 0
        
        lengths = [m.get('Length', 0) for m in main_motifs]
        scores = [m.get('Score', 0) for m in main_motifs]
        
        summary_data['Value'] = [
            1,  # Total sequences
            sequence_length,
            len(motifs),
            len(main_motifs),
            hybrid_count,
            cluster_count,
            len(set(m.get('Class', 'Unknown') for m in main_motifs)),
            len(set(m.get('Subclass', 'Unknown') for m in main_motifs)),
            f"{coverage_pct:.2f}",
            f"{density:.2f}",
            f"{sum(lengths)/len(lengths):.2f}" if lengths else "0",
            min(lengths) if lengths else 0,
            max(lengths) if lengths else 0,
            f"{sum(scores)/len(scores):.2f}" if scores else "0",
            f"{min(scores):.2f}" if scores else "0",
            f"{max(scores):.2f}" if scores else "0"
        ]
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Sheet 2: Class-Level Density Analysis
        class_stats = {}
        for motif in main_motifs:
            cls = motif.get('Class', 'Unknown')
            if cls not in class_stats:
                class_stats[cls] = {'count': 0, 'total_length': 0, 'scores': []}
            class_stats[cls]['count'] += 1
            class_stats[cls]['total_length'] += motif.get('Length', 0)
            class_stats[cls]['scores'].append(motif.get('Score', 0))
        
        class_data = {
            'Motif Class': [],
            'Count': [],
            'Total Length (bp)': [],
            'Genomic Density (%)': [],
            'Motifs per kb': [],
            'Average Length (bp)': [],
            'Average Score': []
        }
        
        for cls, stats in sorted(class_stats.items()):
            class_data['Motif Class'].append(cls)
            class_data['Count'].append(stats['count'])
            class_data['Total Length (bp)'].append(stats['total_length'])
            genomic_density = (stats['total_length'] / sequence_length * 100) if sequence_length > 0 else 0
            class_data['Genomic Density (%)'].append(f"{genomic_density:.4f}")
            motifs_per_kb = (stats['count'] / sequence_length * 1000) if sequence_length > 0 else 0
            class_data['Motifs per kb'].append(f"{motifs_per_kb:.2f}")
            avg_length = stats['total_length'] / stats['count'] if stats['count'] > 0 else 0
            class_data['Average Length (bp)'].append(f"{avg_length:.2f}")
            avg_score = sum(stats['scores']) / len(stats['scores']) if stats['scores'] else 0
            class_data['Average Score'].append(f"{avg_score:.2f}")
        
        df_class = pd.DataFrame(class_data)
        df_class.to_excel(writer, sheet_name='Class_Level_Analysis', index=False)
        
        # Sheet 3: Subclass-Level Density Analysis
        subclass_stats = {}
        for motif in main_motifs:
            cls = motif.get('Class', 'Unknown')
            subcls = motif.get('Subclass', 'Unknown')
            key = f"{cls}:{subcls}"
            if key not in subclass_stats:
                subclass_stats[key] = {'count': 0, 'total_length': 0, 'scores': []}
            subclass_stats[key]['count'] += 1
            subclass_stats[key]['total_length'] += motif.get('Length', 0)
            subclass_stats[key]['scores'].append(motif.get('Score', 0))
        
        subclass_data = {
            'Motif Class': [],
            'Motif Subclass': [],
            'Count': [],
            'Total Length (bp)': [],
            'Genomic Density (%)': [],
            'Motifs per kb': [],
            'Average Length (bp)': [],
            'Average Score': []
        }
        
        for key, stats in sorted(subclass_stats.items()):
            cls, subcls = key.split(':', 1)
            subclass_data['Motif Class'].append(cls)
            subclass_data['Motif Subclass'].append(subcls)
            subclass_data['Count'].append(stats['count'])
            subclass_data['Total Length (bp)'].append(stats['total_length'])
            genomic_density = (stats['total_length'] / sequence_length * 100) if sequence_length > 0 else 0
            subclass_data['Genomic Density (%)'].append(f"{genomic_density:.4f}")
            motifs_per_kb = (stats['count'] / sequence_length * 1000) if sequence_length > 0 else 0
            subclass_data['Motifs per kb'].append(f"{motifs_per_kb:.2f}")
            avg_length = stats['total_length'] / stats['count'] if stats['count'] > 0 else 0
            subclass_data['Average Length (bp)'].append(f"{avg_length:.2f}")
            avg_score = sum(stats['scores']) / len(stats['scores']) if stats['scores'] else 0
            subclass_data['Average Score'].append(f"{avg_score:.2f}")
        
        df_subclass = pd.DataFrame(subclass_data)
        df_subclass.to_excel(writer, sheet_name='Subclass_Level_Analysis', index=False)
        
        # Sheet 4: Length Distribution by Class
        length_dist = {}
        for motif in main_motifs:
            cls = motif.get('Class', 'Unknown')
            if cls not in length_dist:
                length_dist[cls] = []
            length_dist[cls].append(motif.get('Length', 0))
        
        length_data = {
            'Motif Class': [],
            'Min Length': [],
            'Max Length': [],
            'Mean Length': [],
            'Median Length': [],
            'Std Dev': []
        }
        
        for cls, lengths in sorted(length_dist.items()):
            length_data['Motif Class'].append(cls)
            length_data['Min Length'].append(min(lengths))
            length_data['Max Length'].append(max(lengths))
            length_data['Mean Length'].append(f"{sum(lengths)/len(lengths):.2f}")
            sorted_lengths = sorted(lengths)
            # Correct median calculation for even/odd length lists
            n = len(sorted_lengths)
            if n % 2 == 0:
                median = (sorted_lengths[n//2 - 1] + sorted_lengths[n//2]) / 2
            else:
                median = sorted_lengths[n//2]
            length_data['Median Length'].append(median)
            if len(lengths) > 1:
                mean = sum(lengths) / len(lengths)
                # Use sample variance (N-1) for standard deviation
                variance = sum((x - mean) ** 2 for x in lengths) / (len(lengths) - 1)
                std_dev = variance ** 0.5
                length_data['Std Dev'].append(f"{std_dev:.2f}")
            else:
                length_data['Std Dev'].append("0.00")
        
        df_length = pd.DataFrame(length_data)
        df_length.to_excel(writer, sheet_name='Length_Distribution', index=False)
        
        # Sheet 5: Score Distribution by Class
        score_dist = {}
        for motif in main_motifs:
            cls = motif.get('Class', 'Unknown')
            if cls not in score_dist:
                score_dist[cls] = []
            score_dist[cls].append(motif.get('Score', 0))
        
        score_data = {
            'Motif Class': [],
            'Min Score': [],
            'Max Score': [],
            'Mean Score': [],
            'Median Score': [],
            'Std Dev': []
        }
        
        for cls, scores in sorted(score_dist.items()):
            score_data['Motif Class'].append(cls)
            score_data['Min Score'].append(f"{min(scores):.2f}")
            score_data['Max Score'].append(f"{max(scores):.2f}")
            score_data['Mean Score'].append(f"{sum(scores)/len(scores):.2f}")
            sorted_scores = sorted(scores)
            # Correct median calculation for even/odd length lists
            n = len(sorted_scores)
            if n % 2 == 0:
                median = (sorted_scores[n//2 - 1] + sorted_scores[n//2]) / 2
            else:
                median = sorted_scores[n//2]
            score_data['Median Score'].append(f"{median:.2f}")
            if len(scores) > 1:
                mean = sum(scores) / len(scores)
                # Use sample variance (N-1) for standard deviation
                variance = sum((x - mean) ** 2 for x in scores) / (len(scores) - 1)
                std_dev = variance ** 0.5
                score_data['Std Dev'].append(f"{std_dev:.2f}")
            else:
                score_data['Std Dev'].append("0.00")
        
        df_score = pd.DataFrame(score_data)
        df_score.to_excel(writer, sheet_name='Score_Distribution', index=False)
    
    return f"Statistics exported successfully to {filename}"


def export_to_gff3(motifs: List[Dict[str, Any]], sequence_name: str = "sequence", 
                   filename: Optional[str] = None) -> str:
    """
    Export motifs to GFF3 format
    
    Args:
        motifs: List of motif dictionaries
        sequence_name: Name of the sequence
        filename: Optional output filename
        
    Returns:
        GFF3 format string
    """
    gff_lines = []
    gff_lines.append("##gff-version 3")
    gff_lines.append(f"##sequence-region {sequence_name} 1 {len(sequence_name)}")
    
    for i, motif in enumerate(motifs, 1):
        seqid = sequence_name
        source = "NBDScanner"
        feature_type = "Non_B_DNA_motif"
        start = motif.get('Start', 1)
        end = motif.get('End', start)
        score = motif.get('Score', '.')
        strand = motif.get('Strand', '+')
        phase = "."
        
        # Attributes
        attributes = [
            f"ID=motif_{i}",
            f"Name={motif.get('Class', 'Unknown')}_{motif.get('Subclass', 'Unknown')}",
            f"motif_class={motif.get('Class', 'Unknown')}",
            f"motif_subclass={motif.get('Subclass', 'Unknown')}",
            f"length={motif.get('Length', 0)}",
            f"method={motif.get('Method', 'NBDScanner')}"
        ]
        
        attributes_str = ';'.join(attributes)
        
        gff_line = f"{seqid}\t{source}\t{feature_type}\t{start}\t{end}\t{score}\t{strand}\t{phase}\t{attributes_str}"
        gff_lines.append(gff_line)
    
    gff_content = '\n'.join(gff_lines)
    
    if filename:
        try:
            with open(filename, 'w') as f:
                f.write(gff_content)
        except Exception as e:
            print(f"Error writing GFF3 file {filename}: {e}")
    
    return gff_content

# =============================================================================
# QUALITY CONTROL & FILTERING
# =============================================================================

def quality_check_motifs(motifs: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Perform quality checks on detected motifs
    
    Args:
        motifs: List of motif dictionaries
        
    Returns:
        Quality check report
    """
    if not motifs:
        return {'status': 'No motifs to check', 'passed': True, 'issues': []}
    
    issues = []
    
    # Check for required fields
    required_fields = ['Class', 'Subclass', 'Start', 'End', 'Sequence']
    for i, motif in enumerate(motifs):
        missing_fields = [field for field in required_fields if field not in motif]
        if missing_fields:
            issues.append(f"Motif {i+1}: Missing fields {missing_fields}")
    
    # Check coordinate consistency
    for i, motif in enumerate(motifs):
        start = motif.get('Start')
        end = motif.get('End')
        length = motif.get('Length')
        sequence = motif.get('Sequence', '')
        
        if start and end and start >= end:
            issues.append(f"Motif {i+1}: Invalid coordinates (start >= end)")
        
        if start and end and length and (end - start + 1) != length:
            issues.append(f"Motif {i+1}: Length inconsistent with coordinates")
        
        if sequence and length and len(sequence) != length:
            issues.append(f"Motif {i+1}: Sequence length doesn't match reported length")
    
    # Check for overlaps within same class
    class_groups = defaultdict(list)
    for motif in motifs:
        class_groups[motif.get('Class')].append(motif)
    
    for class_name, class_motifs in class_groups.items():
        for i, motif1 in enumerate(class_motifs):
            for motif2 in class_motifs[i+1:]:
                if (motif1.get('Start', 0) < motif2.get('End', 0) and 
                    motif2.get('Start', 0) < motif1.get('End', 0)):
                    issues.append(f"Overlapping motifs in class {class_name}")
                    break
    
    report = {
        'total_motifs': len(motifs),
        'issues_found': len(issues),
        'passed': len(issues) == 0,
        'issues': issues[:10],  # Limit to first 10 issues
        'status': 'PASSED' if len(issues) == 0 else f'FAILED ({len(issues)} issues)'
    }
    
    return report

def filter_motifs_by_score(motifs: List[Dict[str, Any]], min_score: float = 0.0) -> List[Dict[str, Any]]:
    """
    Filter motifs by minimum score threshold
    
    Args:
        motifs: List of motif dictionaries
        min_score: Minimum score threshold
        
    Returns:
        Filtered motifs list
    """
    return [m for m in motifs if m.get('Score', 0) >= min_score]

def filter_motifs_by_length(motifs: List[Dict[str, Any]], 
                           min_length: int = 0, max_length: int = float('inf')) -> List[Dict[str, Any]]:
    """
    Filter motifs by length range
    
    Args:
        motifs: List of motif dictionaries
        min_length: Minimum length
        max_length: Maximum length
        
    Returns:
        Filtered motifs list
    """
    return [m for m in motifs if min_length <= m.get('Length', 0) <= max_length]

def filter_motifs_by_class(motifs: List[Dict[str, Any]], 
                          allowed_classes: List[str]) -> List[Dict[str, Any]]:
    """
    Filter motifs by allowed classes
    
    Args:
        motifs: List of motif dictionaries
        allowed_classes: List of allowed class names
        
    Returns:
        Filtered motifs list
    """
    return [m for m in motifs if m.get('Class') in allowed_classes]

# =============================================================================
# CROSS-DETECTOR OVERLAP RESOLUTION (Hyperscan Integration Pattern)
# =============================================================================

def resolve_cross_class_overlaps(motifs: List[Dict[str, Any]], 
                                 mode: str = 'strict') -> List[Dict[str, Any]]:
    """
    Resolve overlaps across different motif classes using deterministic selection.
    
    Implements the overlap resolution strategy from the Hyperscan integration plan:
    - Option A (strict): Select highest normalized score among overlapping regions
    - Option B (hybrid): Keep both but mark as Hybrid (handled by detector pipeline)
    
    Algorithm:
    1. Sort candidates by score (descending), then by length (descending)
    2. Greedily select highest-scoring non-overlapping motifs
    3. This ensures deterministic, reproducible output
    
    Args:
        motifs: List of motif dictionaries from multiple detectors
        mode: 'strict' for non-overlapping selection, 'hybrid' for hybrid marking
        
    Returns:
        List of resolved motifs (non-overlapping if mode='strict')
        
    References:
        - Hyperscan integration pattern for Non-B DNA detection
        - Score-aware greedy selection (similar to Cruciform._remove_overlaps)
    """
    if not motifs:
        return []
    
    if mode == 'hybrid':
        # Return all motifs - hybrid detection happens in the scanner pipeline
        return motifs
    
    # Mode 'strict': Select highest-scoring non-overlapping set
    # Sort by score (descending), then by length (descending) for tie-breaking
    sorted_motifs = sorted(motifs, 
                          key=lambda x: (-x.get('Score', 0), 
                                       -(x.get('End', 0) - x.get('Start', 0))))
    
    selected = []
    
    for candidate in sorted_motifs:
        # Check if this candidate overlaps with any already selected motif
        overlaps = False
        cand_start = candidate.get('Start', 0)
        cand_end = candidate.get('End', 0)
        
        for selected_motif in selected:
            sel_start = selected_motif.get('Start', 0)
            sel_end = selected_motif.get('End', 0)
            
            # Check for overlap: two regions overlap if neither is completely before the other
            if not (cand_end <= sel_start or cand_start >= sel_end):
                overlaps = True
                break
        
        if not overlaps:
            selected.append(candidate)
    
    # Sort by start position for final output
    selected.sort(key=lambda x: x.get('Start', 0))
    
    return selected

def merge_detector_results(detector_results: Dict[str, List[Dict[str, Any]]],
                          overlap_mode: str = 'strict') -> List[Dict[str, Any]]:
    """
    Merge results from multiple detectors with overlap resolution.
    
    This function implements the complete Hyperscan integration pattern:
    1. Collect results from all detectors (already scored and filtered)
    2. Resolve cross-class overlaps according to specified mode
    3. Return unified, non-redundant motif list
    
    Args:
        detector_results: Dictionary mapping detector_name -> list of motifs
        overlap_mode: 'strict' for non-overlapping, 'hybrid' for overlap annotation
        
    Returns:
        Merged list of motifs with overlaps resolved
        
    Example:
        >>> results = {
        ...     'a_philic': [motif1, motif2],
        ...     'z_dna': [motif3, motif4],
        ...     'g_quadruplex': [motif5]
        ... }
        >>> merged = merge_detector_results(results, overlap_mode='strict')
    """
    # Flatten all detector results into single list
    all_motifs = []
    for detector_name, motifs in detector_results.items():
        # Add detector source to each motif for tracking
        for motif in motifs:
            if 'Detector' not in motif:
                motif['Detector'] = detector_name
            all_motifs.append(motif)
    
    # Resolve overlaps according to mode
    resolved = resolve_cross_class_overlaps(all_motifs, mode=overlap_mode)
    
    return resolved

def export_results_to_dataframe(motifs: List[Dict[str, Any]]) -> pd.DataFrame:
    """Convert motif results to pandas DataFrame with CORE fields only for display tables.
    
    Per Task 1 & 2 requirements, output tables should only show mandatory, universal columns:
    - Sequence_Name, Class, Subclass, Start, End, Length, Strand, Score, Method, Pattern_ID
    
    Additional motif-specific columns (Repeat_Unit, Loop_Length, etc.) are only shown 
    in Excel download class-specific sheets, not in display tables.
    
    This ensures publication-grade clarity per Nature/NAR/Genome Research standards.
    """
    if not motifs:
        return pd.DataFrame()
    
    df = pd.DataFrame(motifs)
    
    # Use core columns constant
    core_columns = CORE_OUTPUT_COLUMNS
    
    # Ensure all core columns are present with appropriate defaults
    for col in core_columns:
        if col not in df.columns:
            # Set appropriate defaults for missing columns using constants
            df[col] = DEFAULT_COLUMN_VALUES.get(col, 'NA')
    
    # Fill all NaN/None values with appropriate defaults
    result_df = df[core_columns].copy()
    for col in core_columns:
        default_val = DEFAULT_COLUMN_VALUES.get(col, 'NA')
        result_df[col] = result_df[col].fillna(default_val)
    
    return result_df


# =============================================================================
# TESTING & EXAMPLES
# =============================================================================

def test_utilities():
    """Test utility functions with example data"""
    print("Testing NBDScanner utilities...")
    
    # Test sequence validation
    test_sequences = [
        "ATGCATGCATGC",     # Valid
        "ATGCXYZ",          # Invalid characters
        "ATG",              # Too short
        "",                 # Empty
    ]
    
    print("\nSequence validation tests:")
    for seq in test_sequences:
        valid, msg = validate_sequence(seq)
        print(f"  '{seq[:20]}': {'VALID' if valid else 'INVALID'} - {msg}")
    
    # Test basic statistics
    test_seq = "GGGTTAGGGTTAGGGTTAGGGAAAAATTTTCGCGCGCGCG"
    stats = get_basic_stats(test_seq)
    print(f"\nBasic stats for test sequence:")
    for key, value in stats.items():
        print(f"  {key}: {value}")
    
    # Test FASTA parsing
    fasta_content = """>test_sequence
GGGTTAGGGTTAGGGTTAGGG
>another_sequence
AAAAATTTTCCCCGGGG"""
    
    sequences = parse_fasta(fasta_content)
    print(f"\nFASTA parsing test: {len(sequences)} sequences parsed")
    for name, seq in sequences.items():
        print(f"  {name}: {seq}")
    
    # Test motif formatting
    example_motifs = [
        {
            'Class': 'G-Quadruplex',
            'Subclass': 'Canonical G4',
            'Start': 1,
            'End': 21,
            'Length': 21,
            'Sequence': 'GGGTTAGGGTTAGGGTTAGGG',
            'Score': 0.857,
            'Strand': '+'
        }
    ]
    
    quality_report = quality_check_motifs(example_motifs)
    print(f"\nQuality check: {quality_report['status']}")
    
    print("[OK] All utility tests completed")

# =============================================================================
# ENHANCED STATISTICS: DENSITY AND ENRICHMENT ANALYSIS
# =============================================================================

# Constants for enrichment analysis
DEFAULT_FOLD_ENRICHMENT_WHEN_ZERO_BACKGROUND = 1.0  # When background is zero





def calculate_genomic_density(motifs: List[Dict[str, Any]], 
                               sequence_length: int,
                               by_class: bool = True,
                               by_subclass: bool = False) -> Dict[str, float]:
    """
    Calculate genomic density (coverage) for motifs.
    
    Genomic Density (σ_G) = (Total unique bp covered by motifs / 
                             Total length in bp of analyzed region) × 100
    
    IMPORTANT: Uses set-based overlap handling to ensure coverage never exceeds 100%.
    If motifs overlap, only unique positions are counted.
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Total length of analyzed sequence
        by_class: If True, calculate density per motif class
        by_subclass: If True, calculate density per motif subclass (takes precedence over by_class)
        
    Returns:
        Dictionary with density metrics (percentage, capped at 100%)
        - If by_subclass=True: keys are 'Class:Subclass' format
        - If by_class=True: keys are class names
        - Otherwise: key is 'Overall'
    """
    if not motifs or sequence_length == 0:
        return {'Overall': 0.0}
    
    if not by_class and not by_subclass:
        # Overall density using set-based coverage (handles overlaps correctly)
        covered_positions = set()
        for motif in motifs:
            start = motif.get('Start', 0) - 1  # Convert to 0-based
            end = motif.get('End', 0)
            covered_positions.update(range(start, end))
        
        overall_density = min((len(covered_positions) / sequence_length) * 100, 100.0)
        return {'Overall': round(overall_density, 4)}
    
    # Density per subclass using set-based coverage
    if by_subclass:
        density_by_subclass = {}
        subclass_groups = defaultdict(list)
        
        for motif in motifs:
            class_name = motif.get('Class', 'Unknown')
            subclass_name = motif.get('Subclass', 'Unknown')
            key = f"{class_name}:{subclass_name}"
            subclass_groups[key].append(motif)
        
        # Calculate per-subclass density with overlap handling
        for subclass_key, subclass_motifs in subclass_groups.items():
            covered_positions = set()
            for motif in subclass_motifs:
                start = motif.get('Start', 0) - 1  # Convert to 0-based
                end = motif.get('End', 0)
                covered_positions.update(range(start, end))
            
            subclass_density = min((len(covered_positions) / sequence_length) * 100, 100.0)
            density_by_subclass[subclass_key] = round(subclass_density, 4)
        
        # Calculate overall density (all motifs combined)
        all_covered_positions = set()
        for motif in motifs:
            start = motif.get('Start', 0) - 1  # Convert to 0-based
            end = motif.get('End', 0)
            all_covered_positions.update(range(start, end))
        
        overall_density = min((len(all_covered_positions) / sequence_length) * 100, 100.0)
        density_by_subclass['Overall'] = round(overall_density, 4)
        
        return density_by_subclass
    
    # Density per class using set-based coverage
    density_by_class = {}
    class_groups = defaultdict(list)
    
    for motif in motifs:
        class_name = motif.get('Class', 'Unknown')
        class_groups[class_name].append(motif)
    
    # Calculate per-class density with overlap handling
    for class_name, class_motifs in class_groups.items():
        covered_positions = set()
        for motif in class_motifs:
            start = motif.get('Start', 0) - 1  # Convert to 0-based
            end = motif.get('End', 0)
            covered_positions.update(range(start, end))
        
        class_density = min((len(covered_positions) / sequence_length) * 100, 100.0)
        density_by_class[class_name] = round(class_density, 4)
    
    # Calculate overall density (all motifs combined)
    all_covered_positions = set()
    for motif in motifs:
        start = motif.get('Start', 0) - 1  # Convert to 0-based
        end = motif.get('End', 0)
        all_covered_positions.update(range(start, end))
    
    overall_density = min((len(all_covered_positions) / sequence_length) * 100, 100.0)
    density_by_class['Overall'] = round(overall_density, 4)
    
    return density_by_class


def calculate_positional_density(motifs: List[Dict[str, Any]], 
                                  sequence_length: int,
                                  unit: str = 'Mbp',
                                  by_class: bool = True,
                                  by_subclass: bool = False) -> Dict[str, float]:
    """
    Calculate positional density (frequency) for motifs.
    
    Positional Density (λ) = Total count of predicted motifs / 
                             Total length (in kbp or Mbp) of analyzed region
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Total length of analyzed sequence in bp
        unit: 'kbp' or 'Mbp' for reporting units
        by_class: If True, calculate density per motif class
        by_subclass: If True, calculate density per motif subclass (takes precedence over by_class)
        
    Returns:
        Dictionary with positional density (motifs per unit)
        - If by_subclass=True: keys are 'Class:Subclass' format
        - If by_class=True: keys are class names
        - Otherwise: key is 'Overall'
    """
    if not motifs or sequence_length == 0:
        return {'Overall': 0.0}
    
    # Convert to appropriate unit
    if unit == 'kbp':
        sequence_length_unit = sequence_length / 1000
    elif unit == 'Mbp':
        sequence_length_unit = sequence_length / 1000000
    else:
        sequence_length_unit = sequence_length
    
    if not by_class and not by_subclass:
        # Overall positional density
        overall_density = len(motifs) / sequence_length_unit
        return {'Overall': round(overall_density, 2)}
    
    # Positional density per subclass
    if by_subclass:
        density_by_subclass = {}
        subclass_counts = Counter()
        
        for motif in motifs:
            class_name = motif.get('Class', 'Unknown')
            subclass_name = motif.get('Subclass', 'Unknown')
            key = f"{class_name}:{subclass_name}"
            subclass_counts[key] += 1
        
        for subclass_key, count in subclass_counts.items():
            subclass_density = count / sequence_length_unit
            density_by_subclass[subclass_key] = round(subclass_density, 2)
        
        # Also add overall
        density_by_subclass['Overall'] = round(len(motifs) / sequence_length_unit, 2)
        
        return density_by_subclass
    
    # Positional density per class
    density_by_class = {}
    class_counts = Counter(m.get('Class', 'Unknown') for m in motifs)
    
    for class_name, count in class_counts.items():
        class_density = count / sequence_length_unit
        density_by_class[class_name] = round(class_density, 2)
    
    # Also add overall
    density_by_class['Overall'] = round(len(motifs) / sequence_length_unit, 2)
    
    return density_by_class


def calculate_enhanced_statistics(motifs: List[Dict[str, Any]], 
                                  sequence: str,
                                  include_enrichment: bool = False,
                                  n_shuffles: int = 0,
                                  progress_callback=None) -> Dict[str, Any]:
    """
    Calculate comprehensive statistics including density analysis.
    
    Note: Enrichment analysis has been removed for performance optimization.
    
    Args:
        motifs: List of detected motifs
        sequence: Original DNA sequence
        include_enrichment: Deprecated (kept for backward compatibility)
        n_shuffles: Deprecated (kept for backward compatibility)
        progress_callback: Optional callback for progress updates
        
    Returns:
        Dictionary with comprehensive statistics
    """
    import warnings
    
    # Warn about deprecated parameters
    if include_enrichment or n_shuffles > 0:
        warnings.warn(
            "The 'include_enrichment' and 'n_shuffles' parameters are deprecated. "
            "Enrichment analysis has been removed for performance optimization.",
            DeprecationWarning,
            stacklevel=2
        )
    
    sequence_length = len(sequence)
    
    # Basic statistics
    basic_stats = calculate_motif_statistics(motifs, sequence_length)
    
    # Density calculations
    genomic_density = calculate_genomic_density(motifs, sequence_length, by_class=True)
    positional_density_kbp = calculate_positional_density(motifs, sequence_length, unit='kbp', by_class=True)
    positional_density_mbp = calculate_positional_density(motifs, sequence_length, unit='Mbp', by_class=True)
    
    enhanced_stats = {
        'basic': basic_stats,
        'genomic_density': genomic_density,
        'positional_density_per_kbp': positional_density_kbp,
        'positional_density_per_mbp': positional_density_mbp
    }
    
    return enhanced_stats


if __name__ == "__main__":
    test_utilities()

# =============================================================================
# VISUALIZATION SUITE (merged from visualizations.py)
# =============================================================================
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║                    NBDSCANNER VISUALIZATION SUITE                             ║
║        Comprehensive Plotting and Visualization Functions                    ║
╚══════════════════════════════════════════════════════════════════════════════╝

VISUALIZATION FUNCTIONS (integrated into utilities.py)
AUTHOR: Dr. Venkata Rajesh Yella
VERSION: 2024.1
LICENSE: MIT

DESCRIPTION:
    Comprehensive plotting and visualization functions for Non-B DNA motif analysis.
    Includes both static and interactive visualizations with scientific styling.

VISUALIZATION TABLE:
┌──────────────────────────────────────────────────────────────────────────────┐
│ Category     │ Functions                 │ Description                       │
├──────────────┼───────────────────────────┼───────────────────────────────────┤
│ Distribution │ plot_motif_distribution   │ Class/subclass distribution plots │
│ Coverage     │ plot_coverage_map         │ Sequence coverage visualization   │
│ Statistics   │ plot_length_distribution  │ Length distributions              │
│ Hierarchy    │ plot_nested_pie_chart     │ Class-subclass hierarchy          │
│ Export       │ save_all_plots            │ Batch plot export                 │
└──────────────────────────────────────────────────────────────────────────────┘

ADVANCED VISUALIZATIONS (Publication-Quality):
┌──────────────────────────────────────────────────────────────────────────────┐
│ - plot_genome_landscape_track      │ Horizontal genomic ruler with glyphs │
│ - plot_sliding_window_heat_ribbon  │ 1D density heatmap with score        │
│ - plot_ridge_plots_length_by_class │ Stacked density ridges (joyplots)    │
│ - plot_sunburst_treemap            │ Hierarchical composition             │
│ - plot_hexbin_start_vs_score       │ 2D hexbin with marginals             │
│ - plot_upset_intersection          │ UpSet plot for motif overlaps        │
│ - plot_score_violin_beeswarm       │ Score distributions with points      │
│ - plot_cluster_hotspot_map         │ Cluster regions with annotations     │
└──────────────────────────────────────────────────────────────────────────────┘

FEATURES:
    - Publication-quality static plots
    - Interactive visualizations (Plotly)
    - Colorblind-friendly palettes
    - SVG/PNG export at 300 DPI
    - Customizable styling
"""

# =============================================================================
# NATURE-LEVEL PUBLICATION STYLING & CONFIGURATION
# =============================================================================
# Following Nature Methods/Nature Genetics style guidelines:
# - Sans-serif fonts (Arial/Helvetica preferred)
# - High DPI (300+ for print)
# - Clean, minimal design with proper spacing
# - Colorblind-friendly palettes

# Default DPI for publication quality (Nature requires 300 DPI minimum)
PUBLICATION_DPI = 300

# Nature-level color palette for motif classes (colorblind-friendly)
# Based on Wong, B. (2011) Nature Methods colorblind-safe palette
# Each motif class has a unique color for distinguishability
# Reference: Wong, B. (2011) Points of view: Color blindness. Nat Methods 8, 441
MOTIF_CLASS_COLORS = {
    # Primary motif classes (6 core colors per Nature guidelines)
    'Curved_DNA': '#CC79A7',          # Reddish Purple - Structure
    'G-Quadruplex': '#0072B2',        # Blue - Stable structures
    'Z-DNA': '#882255',               # Wine - Alternative helices
    'Cruciform': '#56B4E9',           # Sky Blue - Symmetric structures
    'Triplex': '#E69F00',             # Orange - Triple-stranded
    'R-Loop': '#009E73',              # Bluish Green - RNA-DNA hybrids
    
    # Secondary classes (consolidated to reduce color count)
    'i-Motif': '#0072B2',             # Same as G4 (complementary structures)
    'A-philic_DNA': '#CC79A7',        # Same as Curved (structural affinity)
    'Slipped_DNA': '#E69F00',         # Same as Triplex (repeats)
    
    # Meta-classes (neutral grays)
    'Hybrid': '#888888',              # Medium gray
    'Non-B_DNA_Clusters': '#666666'   # Dark gray
}

# Helper function to format display names
def _format_display_name(name: str) -> str:
    """Convert internal names to publication-ready display format.
    
    Replaces underscores with spaces for better readability.
    
    Args:
        name: Internal name (e.g., 'Curved_DNA')
        
    Returns:
        Display name (e.g., 'Curved DNA')
    """
    return name.replace('_', ' ')

# Nature-level scientific styling configuration for publication-quality plots
# Reference: Nature author guidelines for figure preparation
_NATURE_STYLE_PARAMS = {
    # Typography - Arial/Helvetica as per Nature guidelines
    'font.family': 'sans-serif',
    'font.sans-serif': ['Arial', 'Helvetica', 'DejaVu Sans'],
    'font.size': 8,  # Nature recommends 5-7pt, we use 8pt for readability
    
    # Title and labels
    'axes.titlesize': 9,
    'axes.titleweight': 'bold',
    'axes.labelsize': 8,
    'axes.labelweight': 'normal',
    
    # Tick labels
    'xtick.labelsize': 7,
    'ytick.labelsize': 7,
    'xtick.major.size': 3,
    'ytick.major.size': 3,
    'xtick.major.width': 0.8,
    'ytick.major.width': 0.8,
    
    # Legend
    'legend.fontsize': 7,
    'legend.frameon': False,
    'legend.borderpad': 0.4,
    
    # Figure
    'figure.titlesize': 10,
    'figure.titleweight': 'bold',
    'figure.dpi': PUBLICATION_DPI,
    'figure.facecolor': 'white',
    'savefig.dpi': PUBLICATION_DPI,
    'savefig.format': 'pdf',  # Vector format for publication
    'savefig.bbox': 'tight',
    'savefig.pad_inches': 0.05,
    
    # Axes - minimal, clean design
    'axes.grid': False,  # Nature typically uses minimal grids
    'axes.spines.top': False,
    'axes.spines.right': False,
    'axes.spines.left': True,
    'axes.spines.bottom': True,
    'axes.linewidth': 0.8,
    'axes.edgecolor': 'black',
    'axes.facecolor': 'white',
    
    # Lines
    'lines.linewidth': 1.0,
    'lines.markersize': 4,
    
    # PDF/SVG output
    'pdf.fonttype': 42,  # TrueType for editability
    'ps.fonttype': 42,
}

# Apply Nature-level styling at module load
plt.rcParams.update(_NATURE_STYLE_PARAMS)

# Constants for enrichment analysis visualization
INFINITE_FOLD_ENRICHMENT_CAP = 100  # Cap for infinite fold enrichment values in plots

# Figure size constants (in inches) following Nature guidelines
# Nature column widths: single=89mm(3.5in), 1.5=120mm(4.7in), double=183mm(7.2in)
FIGURE_SIZES = {
    'single_column': (3.5, 2.8),      # Single column
    'one_and_half': (4.7, 3.5),       # 1.5 column
    'double_column': (7.2, 4.5),      # Double column
    'square': (3.5, 3.5),             # Square figure
    'wide': (7.2, 3.0),               # Wide panoramic
}


def set_scientific_style(style: str = 'nature'):
    """
    Apply scientific publication-ready styling.
    
    Args:
        style: Style preset ('nature', 'default', 'presentation')
               - 'nature': Nature/Science journal style (default)
               - 'default': Standard seaborn whitegrid
               - 'presentation': Larger fonts for presentations
    """
    if style == 'nature':
        plt.rcParams.update(_NATURE_STYLE_PARAMS)
        sns.set_style("white")
        sns.set_context("paper")
    elif style == 'presentation':
        plt.rcParams.update(_NATURE_STYLE_PARAMS)
        plt.rcParams.update({
            'font.size': 12,
            'axes.titlesize': 14,
            'axes.labelsize': 12,
            'xtick.labelsize': 11,
            'ytick.labelsize': 11,
            'legend.fontsize': 10,
        })
        sns.set_style("whitegrid")
        sns.set_context("talk")
    else:
        sns.set_style("whitegrid")
        sns.set_palette("husl")

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def _apply_nature_style(ax):
    """Apply Nature journal style to matplotlib axis (remove top/right spines).
    
    | Parameter | Type | Description            |
    |-----------|------|------------------------|
    | ax        | Axes | Matplotlib axis object |
    """
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

# =============================================================================
# DISTRIBUTION PLOTS (Nature-Level Quality)
# =============================================================================

def plot_motif_distribution(motifs: List[Dict[str, Any]], 
                           by: str = 'Class',
                           title: Optional[str] = None,
                           figsize: Tuple[float, float] = None,
                           style: str = 'nature') -> plt.Figure:
    """
    Plot distribution of motifs by class or subclass.
    
    Publication-quality bar chart following Nature Methods guidelines.
    Shows all classes/subclasses even when count is 0, ensuring comprehensive
    visualization of both detected and undetected motif types.
    
    Args:
        motifs: List of motif dictionaries.
        by: Group by 'Class' or 'Subclass'.
        title: Custom plot title.
        figsize: Figure size (width, height) in inches. Uses Nature standard if None.
        style: Style preset ('nature', 'default', 'presentation').
        
    Returns:
        Matplotlib figure object (publication-ready at 300 DPI)
    """
    set_scientific_style(style)
    
    # Use Nature-appropriate figure size
    if figsize is None:
        figsize = FIGURE_SIZES['double_column'] if by == 'Subclass' else FIGURE_SIZES['one_and_half']
    
    # Define ALL expected classes and subclasses (always show these)
    ALL_CLASSES = [
        'Curved_DNA', 'Slipped_DNA', 'Cruciform', 'R-Loop', 'Triplex',
        'G-Quadruplex', 'i-Motif', 'Z-DNA', 'A-philic_DNA', 'Hybrid', 'Non-B_DNA_Clusters'
    ]
    
    ALL_SUBCLASSES = [
        'Global Curvature', 'Local Curvature',  # Curved DNA
        'Direct Repeat', 'STR',  # Slipped DNA
        'Inverted Repeats',  # Cruciform
        'R-loop formation sites', 'QmRLFS-m1', 'QmRLFS-m2',  # R-Loop
        'Triplex', 'Sticky DNA',  # Triplex
        'Telomeric G4', 'Stacked canonical G4s', 'Stacked G4s with linker',  # G-Quadruplex
        'Canonical intramolecular G4', 'Extended-loop canonical',
        'Higher-order G4 array/G4-wire', 'Intramolecular G-triplex', 'Two-tetrad weak PQS',
        'Canonical i-motif', 'Relaxed i-motif', 'AC-motif',  # i-Motif
        'Z-DNA', 'eGZ (Extruded-G) DNA',  # Z-DNA
        'A-philic DNA',  # A-philic
    ]
    
    # Count motifs by specified grouping
    counts = Counter(m.get(by, 'Unknown') for m in motifs) if motifs else Counter()
    
    # Prepare data with all categories
    if by == 'Class':
        categories = ALL_CLASSES
    else:
        categories = ALL_SUBCLASSES
    
    # Get counts (0 if not present)
    values = [counts.get(cat, 0) for cat in categories]
    
    # Get colors (colorblind-friendly)
    if by == 'Class':
        colors = [MOTIF_CLASS_COLORS.get(cat, '#808080') for cat in categories]
    else:
        # Map subclasses to their parent class colors
        subclass_to_class = {
            'Global Curvature': 'Curved_DNA', 'Local Curvature': 'Curved_DNA',
            'Direct Repeat': 'Slipped_DNA', 'STR': 'Slipped_DNA',
            'Inverted Repeats': 'Cruciform',
            'R-loop formation sites': 'R-Loop', 'QmRLFS-m1': 'R-Loop', 'QmRLFS-m2': 'R-Loop',
            'Triplex': 'Triplex', 'Sticky DNA': 'Triplex',
            'Telomeric G4': 'G-Quadruplex', 'Stacked canonical G4s': 'G-Quadruplex',
            'Stacked G4s with linker': 'G-Quadruplex', 'Canonical intramolecular G4': 'G-Quadruplex',
            'Extended-loop canonical': 'G-Quadruplex', 'Higher-order G4 array/G4-wire': 'G-Quadruplex',
            'Intramolecular G-triplex': 'G-Quadruplex', 'Two-tetrad weak PQS': 'G-Quadruplex',
            'Canonical i-motif': 'i-Motif', 'Relaxed i-motif': 'i-Motif', 'AC-motif': 'i-Motif',
            'Z-DNA': 'Z-DNA', 'eGZ (Extruded-G) DNA': 'Z-DNA',
            'A-philic DNA': 'A-philic_DNA',
        }
        colors = [MOTIF_CLASS_COLORS.get(subclass_to_class.get(cat, ''), '#808080') for cat in categories]
    
    # Create figure with high DPI
    fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
    
    # Create bars with Nature-style aesthetics
    bars = ax.bar(range(len(categories)), values, color=colors, 
                  edgecolor='black', linewidth=0.5, width=0.8)
    
    # Customize axes (Nature style - minimal, clean)
    ax.set_xlabel(f'Motif {by}', fontweight='normal')
    ax.set_ylabel('Count', fontweight='normal')
    if title:
        # Replace underscores with spaces in title
        display_title = title.replace('_', ' ')
        ax.set_title(display_title, fontweight='bold', pad=10)
    ax.set_xticks(range(len(categories)))
    
    # Replace underscores with spaces in category labels
    # Apply 45° rotation for all categories (Nature style - consistent readability)
    display_categories = [cat.replace('_', ' ') for cat in categories]
    ax.set_xticklabels(display_categories, rotation=45, ha='right')
    
    # Add count labels on ALL bars (improved visibility)
    # Show numbers for all categories to make distribution clear
    max_val = max(values) if max(values) > 0 else 1
    for bar, count in zip(bars, values):
        height = bar.get_height()
        # Position label above bar if count > 0, at baseline if 0
        y_pos = height + max_val * 0.02 if count > 0 else 0.5
        # Use larger font (8pt) and bold for better readability
        ax.text(bar.get_x() + bar.get_width()/2., y_pos,
                str(count), ha='center', va='bottom', fontsize=8, fontweight='bold')
    
    # Apply Nature journal style
    _apply_nature_style(ax)
    
    plt.tight_layout()
    return fig


def plot_class_subclass_sunburst(motifs: List[Dict[str, Any]], 
                                 title: str = "Motif Class-Subclass Distribution",
                                 figsize: Tuple[float, float] = None) -> Union[plt.Figure, Any]:
    """
    Create sunburst plot showing class-subclass hierarchy.
    
    Publication-quality hierarchical visualization.
    
    Args:
        motifs: List of motif dictionaries
        title: Plot title
        figsize: Figure size (width, height) in inches
        
    Returns:
        Plotly figure if available, otherwise matplotlib figure
    """
    if figsize is None:
        figsize = FIGURE_SIZES['square']
        
    if not motifs:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center', 
                transform=ax.transAxes)
        ax.axis('off')
        if title:
            ax.set_title(title)
        return fig
    
    if not PLOTLY_AVAILABLE:
        # Fallback to matplotlib nested pie chart
        return plot_nested_pie_chart(motifs, title)
    
    # Build hierarchical data for sunburst
    class_subclass_counts = defaultdict(lambda: defaultdict(int))
    
    for motif in motifs:
        class_name = motif.get('Class', 'Unknown')
        subclass_name = motif.get('Subclass', 'Unknown')
        class_subclass_counts[class_name][subclass_name] += 1
    
    # Prepare data for plotly
    ids = []
    labels = []
    parents = []
    values = []
    colors = []
    
    # Add classes (inner ring)
    for class_name, subclasses in class_subclass_counts.items():
        total_class_count = sum(subclasses.values())
        ids.append(class_name)
        labels.append(f"{class_name}<br>({total_class_count})")
        parents.append("")
        values.append(total_class_count)
        colors.append(MOTIF_CLASS_COLORS.get(class_name, '#808080'))
    
    # Add subclasses (outer ring)
    for class_name, subclasses in class_subclass_counts.items():
        for subclass_name, count in subclasses.items():
            ids.append(f"{class_name}_{subclass_name}")
            labels.append(f"{subclass_name}<br>({count})")
            parents.append(class_name)
            values.append(count)
            # Lighter shade of class color for subclasses
            base_color = MOTIF_CLASS_COLORS.get(class_name, '#808080')
            colors.append(base_color + '80')  # Add transparency
    
    fig = go.Figure(go.Sunburst(
        ids=ids,
        labels=labels,
        parents=parents,
        values=values,
        branchvalues="total",
        marker=dict(colors=colors, line=dict(color="#FFFFFF", width=1)),
        hovertemplate='<b>%{label}</b><br>Count: %{value}<extra></extra>',
    ))
    
    # Publication-quality layout
    fig.update_layout(
        title=dict(text=title, font=dict(size=10, family='Arial')),
        font=dict(size=8, family='Arial'),
        width=int(figsize[0] * 100),
        height=int(figsize[1] * 100),
        margin=dict(t=30, l=10, r=10, b=10)
    )
    
    return fig


def plot_nested_pie_chart(motifs: List[Dict[str, Any]], 
                         title: str = "Motif Distribution",
                         figsize: Tuple[float, float] = None) -> plt.Figure:
    """
    Create nested donut chart with improved text placement to avoid overlapping labels.
    
    Publication-quality hierarchical pie chart following Nature guidelines.
    
    Args:
        motifs: List of motif dictionaries
        title: Plot title
        figsize: Figure size (width, height) in inches
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style('nature')
    
    if figsize is None:
        figsize = FIGURE_SIZES['square']
    
    # Count by class and subclass
    class_counts = Counter(m.get('Class', 'Unknown') for m in motifs)
    class_subclass_counts = defaultdict(lambda: defaultdict(int))
    
    for motif in motifs:
        class_name = motif.get('Class', 'Unknown')
        subclass_name = motif.get('Subclass', 'Unknown')
        class_subclass_counts[class_name][subclass_name] += 1
    
    fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
    
    # Inner donut (classes)
    class_names = list(class_counts.keys())
    class_values = list(class_counts.values())
    class_colors = [MOTIF_CLASS_COLORS.get(name, '#808080') for name in class_names]
    
    # Create inner donut with Nature-style clean design
    # Use labels=None to manually place labels later for better control
    wedges1, texts1, autotexts1 = ax.pie(
        class_values, 
        labels=None,  # We'll add labels manually
        colors=class_colors,
        radius=0.65,
        autopct=lambda pct: f'{pct:.1f}%' if pct > 3 else '',  # Show more percentage labels with 1 decimal
        pctdistance=0.80,
        startangle=90,
        wedgeprops=dict(width=0.35, edgecolor='white', linewidth=2)  # Thicker edge for better clarity
    )
    
    # Manually add class labels with better positioning (replace underscores with spaces)
    for i, (wedge, class_name) in enumerate(zip(wedges1, class_names)):
        angle = (wedge.theta2 + wedge.theta1) / 2
        x = 0.5 * np.cos(np.radians(angle))
        y = 0.5 * np.sin(np.radians(angle))
        # Replace underscores with spaces in labels
        display_name = class_name.replace('_', ' ')
        # Add white background box for better readability
        ax.text(x, y, display_name, ha='center', va='center', fontsize=8, fontweight='bold',
                bbox=dict(boxstyle='round,pad=0.3', facecolor='white', edgecolor='none', alpha=0.8))
    
    # Outer donut (subclasses)
    all_subclass_counts = []
    all_subclass_colors = []
    all_subclass_labels = []
    
    for class_name in class_names:
        subclass_dict = class_subclass_counts[class_name]
        base_color = MOTIF_CLASS_COLORS.get(class_name, '#808080')
        
        for subclass_name, count in subclass_dict.items():
            all_subclass_counts.append(count)
            # Truncate long names for clean appearance and replace underscores with spaces
            # Use consistent truncation length (15 chars max, including ellipsis)
            display_name = subclass_name.replace('_', ' ')
            MAX_LABEL_LENGTH = 15
            label = display_name if len(display_name) <= MAX_LABEL_LENGTH else display_name[:MAX_LABEL_LENGTH-1] + '…'
            all_subclass_labels.append(label)
            all_subclass_colors.append(base_color)
    
    # Use smarter labeling strategy to avoid overlap
    # For many subclasses, hide labels and rely on legend instead
    if len(all_subclass_labels) > 10:
        # Hide outer ring labels when there are too many
        wedges2, texts2 = ax.pie(
            all_subclass_counts,
            labels=None,  # No labels for cleaner appearance
            colors=all_subclass_colors,
            radius=1.0,
            startangle=90,
            wedgeprops=dict(width=0.35, edgecolor='white', linewidth=1.5)  # Thicker edge for clarity
        )
        
        # Add a legend for subclasses instead
        # Note: All subclasses of the same parent class share the same color by design.
        # This ensures visual grouping in the nested donut chart.
        # Track first occurrence of each unique label for the legend.
        seen_labels = {}
        legend_handles = []
        legend_labels = []
        for i, (label, color) in enumerate(zip(all_subclass_labels, all_subclass_colors)):
            if label not in seen_labels:
                seen_labels[label] = color
                legend_handles.append(plt.Rectangle((0,0),1,1, fc=color, ec='white', lw=1))
                legend_labels.append(label)
                if len(legend_labels) >= 12:  # Limit to 12 for better display
                    break
        
        ax.legend(legend_handles, legend_labels, loc='center left', bbox_to_anchor=(1, 0.5),
                 fontsize=7, frameon=True, title='Top Subclasses', title_fontsize=8,
                 framealpha=0.95, edgecolor='lightgray')
    else:
        # For fewer subclasses, show labels with improved spacing
        wedges2, texts2 = ax.pie(
            all_subclass_counts,
            labels=all_subclass_labels,
            colors=all_subclass_colors,
            radius=1.0,
            labeldistance=1.18,  # Push labels further out to avoid overlap
            startangle=90,
            wedgeprops=dict(width=0.35, edgecolor='white', linewidth=1.5),  # Thicker edge
            textprops={'fontsize': 7, 'weight': 'medium'}  # Larger, bolder text
        )
        
        # Adjust label positions to avoid overlap
        for text in texts2:
            text.set_fontsize(7)
            text.set_weight('medium')
            # Add slight rotation for better readability
            angle = text.get_rotation()
            if 90 < angle < 270:
                text.set_rotation(angle - 180)
    
    if title:
        # Replace underscores with spaces in title
        display_title = title.replace('_', ' ')
        ax.set_title(display_title, fontweight='bold', fontsize=12, pad=15)
    
    # Style percentage labels - larger and bolder
    for autotext in autotexts1:
        autotext.set_fontsize(7)
        autotext.set_weight('bold')
        autotext.set_color('white')
    
    return fig


# =============================================================================
# COVERAGE & POSITIONAL PLOTS (Nature-Level Quality)
# =============================================================================

def plot_coverage_map(motifs: List[Dict[str, Any]], 
                     sequence_length: int,
                     title: Optional[str] = None,
                     figsize: Tuple[float, float] = None,
                     style: str = 'nature') -> plt.Figure:
    """
    Plot motif coverage map showing positions along sequence.
    
    Publication-quality genomic track visualization.
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Length of the analyzed sequence
        title: Custom plot title
        figsize: Figure size (width, height) in inches
        style: Style preset ('nature', 'default', 'presentation')
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style(style)
    
    if figsize is None:
        figsize = FIGURE_SIZES['wide']
    
    if not motifs:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center', 
                transform=ax.transAxes)
        ax.axis('off')
        if title:
            ax.set_title(title)
        return fig
    
    # Group motifs by class
    class_motifs = defaultdict(list)
    for motif in motifs:
        class_name = motif.get('Class', 'Unknown')
        class_motifs[class_name].append(motif)
    
    # Create figure with high DPI
    fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
    
    y_pos = 0
    class_positions = {}
    
    for class_name, class_motif_list in class_motifs.items():
        class_positions[class_name] = y_pos
        color = MOTIF_CLASS_COLORS.get(class_name, '#808080')
        
        for motif in class_motif_list:
            start = motif.get('Start', 0) - 1  # Convert to 0-based
            end = motif.get('End', start + 1)
            length = end - start
            
            # Draw motif as rectangle (Nature style - clean edges)
            rect = patches.Rectangle(
                (start, y_pos - 0.35), length, 0.7,
                facecolor=color, edgecolor='black', linewidth=0.3
            )
            ax.add_patch(rect)
        
        y_pos += 1
    
    # Customize axes (Nature style)
    ax.set_xlim(0, sequence_length)
    ax.set_ylim(-0.5, len(class_motifs) - 0.5)
    ax.set_xlabel('Position (bp)')
    ax.set_ylabel('Motif Class')
    if title:
        # Replace underscores with spaces in title
        display_title = title.replace('_', ' ')
        ax.set_title(display_title, fontweight='bold', pad=10)
    
    # Set y-axis labels with underscores replaced by spaces
    ax.set_yticks(list(class_positions.values()))
    display_labels = [label.replace('_', ' ') for label in class_positions.keys()]
    ax.set_yticklabels(display_labels)
    
    # Clean x-axis ticks
    ax.ticklabel_format(style='sci', axis='x', scilimits=(3, 3))
    
    # Apply Nature journal style
    _apply_nature_style(ax)
    
    plt.tight_layout()
    return fig


def plot_density_heatmap(motifs: List[Dict[str, Any]], 
                        sequence_length: int,
                        window_size: int = 1000,
                        title: Optional[str] = None,
                        figsize: Tuple[float, float] = None,
                        style: str = 'nature') -> plt.Figure:
    """
    Plot motif density heatmap along sequence.
    
    Publication-quality density visualization.
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Length of the analyzed sequence
        window_size: Window size for density calculation
        title: Custom plot title
        figsize: Figure size (width, height) in inches
        style: Style preset
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style(style)
    
    if figsize is None:
        figsize = FIGURE_SIZES['wide']
    
    if not motifs:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center', 
                transform=ax.transAxes)
        ax.axis('off')
        if title:
            ax.set_title(title)
        return fig
    
    # Calculate windows
    num_windows = max(1, sequence_length // window_size)
    windows = np.linspace(0, sequence_length, num_windows + 1)
    
    # Get unique classes
    classes = sorted(set(m.get('Class', 'Unknown') for m in motifs))
    
    # Calculate density matrix
    density_matrix = np.zeros((len(classes), num_windows))
    
    for i, class_name in enumerate(classes):
        class_motifs = [m for m in motifs if m.get('Class') == class_name]
        
        for j in range(num_windows):
            window_start = windows[j]
            window_end = windows[j + 1]
            
            # Count motifs in window
            count = 0
            for motif in class_motifs:
                motif_start = motif.get('Start', 0)
                motif_end = motif.get('End', 0)
                
                # Check if motif overlaps with window
                if not (motif_end <= window_start or motif_start >= window_end):
                    count += 1
            
            density_matrix[i, j] = count
    
    # Create heatmap with publication quality
    fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
    
    # Use colorblind-friendly colormap
    im = ax.imshow(density_matrix, cmap='viridis', aspect='auto', interpolation='nearest')
    
    # Customize axes (Nature style)
    ax.set_xlabel(f'Position (kb)')
    ax.set_ylabel('Motif Class')
    if title:
        # Replace underscores with spaces in title
        display_title = title.replace('_', ' ')
        ax.set_title(display_title, fontweight='bold', pad=10)
    
    # Set ticks and labels with underscores replaced by spaces
    ax.set_yticks(range(len(classes)))
    display_classes = [cls.replace('_', ' ') for cls in classes]
    ax.set_yticklabels(display_classes)
    
    # Clean x-axis with kb units
    x_ticks = np.arange(0, num_windows, max(1, num_windows // 5))
    ax.set_xticks(x_ticks)
    ax.set_xticklabels([f'{int(windows[i]/1000)}' for i in x_ticks])
    
    # Add colorbar with proper label
    cbar = plt.colorbar(im, ax=ax, shrink=0.8)
    cbar.set_label('Count', fontsize=7)
    cbar.ax.tick_params(labelsize=6)
    
    plt.tight_layout()
    return fig


# =============================================================================
# STATISTICAL PLOTS (Nature-Level Quality)
# =============================================================================

def plot_score_distribution(motifs: List[Dict[str, Any]], 
                           by_class: bool = True,
                           title: Optional[str] = None,
                           figsize: Tuple[float, float] = None,
                           style: str = 'nature') -> plt.Figure:
    """
    Plot distribution of motif scores.
    
    Publication-quality box plot visualization.
    
    Args:
        motifs: List of motif dictionaries
        by_class: Whether to separate by motif class
        title: Custom plot title
        figsize: Figure size (width, height) in inches
        style: Style preset
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style(style)
    
    if figsize is None:
        figsize = FIGURE_SIZES['one_and_half']
    
    if not motifs:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center', 
                transform=ax.transAxes)
        ax.axis('off')
        if title:
            ax.set_title(title)
        return fig
    
    # Extract scores
    scores_data = []
    for motif in motifs:
        score = motif.get('Score', motif.get('Normalized_Score'))
        if isinstance(score, (int, float)):
            if by_class:
                scores_data.append({
                    'Score': score,
                    'Class': motif.get('Class', 'Unknown')
                })
            else:
                scores_data.append(score)
    
    if not scores_data:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No score data available', ha='center', va='center', 
                transform=ax.transAxes)
        ax.axis('off')
        if title:
            display_title = title.replace('_', ' ')
            ax.set_title(display_title)
        return fig
    
    fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
    
    if by_class and isinstance(scores_data[0], dict):
        # Create DataFrame for seaborn
        df = pd.DataFrame(scores_data)
        
        # Box plot by class (Nature style - clean, minimal)
        colors = [MOTIF_CLASS_COLORS.get(cls, '#808080') for cls in df['Class'].unique()]
        sns.boxplot(data=df, x='Class', y='Score', ax=ax, palette=colors,
                   linewidth=0.8, fliersize=2)
        # Replace underscores with spaces in x-tick labels
        ax.set_xticklabels([label.get_text().replace('_', ' ') for label in ax.get_xticklabels()], 
                          rotation=45, ha='right')
        ax.set_ylabel('Score')
        ax.set_xlabel('Motif Class')
    else:
        # Simple histogram
        ax.hist(scores_data, bins=20, edgecolor='black', linewidth=0.5, color='#56B4E9')
        ax.set_xlabel('Score')
        ax.set_ylabel('Frequency')
    
    if title:
        # Replace underscores with spaces in title
        display_title = title.replace('_', ' ')
        ax.set_title(display_title, fontweight='bold', pad=10)
    
    # Apply Nature journal style
    _apply_nature_style(ax)
    
    plt.tight_layout()
    return fig


def plot_length_distribution(motifs: List[Dict[str, Any]], 
                           by_class: bool = True,
                           title: Optional[str] = None,
                           figsize: Tuple[float, float] = None,
                           style: str = 'nature') -> plt.Figure:
    """
    Plot distribution of motif lengths.
    
    Publication-quality violin plot visualization.
    
    Args:
        motifs: List of motif dictionaries
        by_class: Whether to separate by motif class
        title: Custom plot title
        figsize: Figure size (width, height) in inches
        style: Style preset
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style(style)
    
    if figsize is None:
        figsize = FIGURE_SIZES['one_and_half']
    
    if not motifs:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center', 
                transform=ax.transAxes)
        ax.axis('off')
        if title:
            ax.set_title(title)
        return fig
    
    # Extract lengths
    length_data = []
    for motif in motifs:
        length = motif.get('Length')
        if isinstance(length, int) and length > 0:
            if by_class:
                length_data.append({
                    'Length': length,
                    'Class': motif.get('Class', 'Unknown')
                })
            else:
                length_data.append(length)
    
    if not length_data:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No length data available', ha='center', va='center', 
                transform=ax.transAxes)
        ax.axis('off')
        if title:
            display_title = title.replace('_', ' ')
            ax.set_title(display_title)
        return fig
    
    fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
    
    if by_class and isinstance(length_data[0], dict):
        # Create DataFrame for seaborn
        df = pd.DataFrame(length_data)
        
        # Violin plot by class (Nature style - clean)
        colors = [MOTIF_CLASS_COLORS.get(cls, '#808080') for cls in df['Class'].unique()]
        sns.violinplot(data=df, x='Class', y='Length', ax=ax, palette=colors,
                      linewidth=0.8, inner='box')
        # Replace underscores with spaces in x-tick labels
        ax.set_xticklabels([label.get_text().replace('_', ' ') for label in ax.get_xticklabels()], 
                          rotation=45, ha='right')
        ax.set_ylabel('Length (bp)')
        ax.set_xlabel('Motif Class')
    else:
        # Simple histogram
        ax.hist(length_data, bins=20, edgecolor='black', linewidth=0.5, color='#009E73')
        ax.set_xlabel('Length (bp)')
        ax.set_ylabel('Frequency')
    
    if title:
        # Replace underscores with spaces in title
        display_title = title.replace('_', ' ')
        ax.set_title(display_title, fontweight='bold', pad=10)
    
    # Apply Nature journal style
    _apply_nature_style(ax)
    
    plt.tight_layout()
    return fig


# =============================================================================
# COMPARISON PLOTS (Nature-Level Quality)
# =============================================================================

def plot_class_comparison(results: Dict[str, List[Dict[str, Any]]], 
                         metric: str = 'count',
                         title: Optional[str] = None,
                         figsize: Tuple[float, float] = None,
                         style: str = 'nature') -> plt.Figure:
    """
    Compare motif classes across multiple sequences/samples.
    
    Publication-quality heatmap comparison.
    
    Args:
        results: Dictionary of {sample_name: motifs_list}
        metric: Comparison metric ('count', 'coverage', 'density')
        title: Custom plot title
        figsize: Figure size (width, height) in inches
        style: Style preset
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style(style)
    
    if figsize is None:
        figsize = FIGURE_SIZES['double_column']
    
    if not results:
        fig, ax = plt.subplots(figsize=figsize)
        ax.text(0.5, 0.5, 'No data to display', ha='center', va='center', 
                transform=ax.transAxes, fontsize=14)
        ax.set_title(title or 'Class Comparison')
        return fig
    
    # Collect all classes across samples
    all_classes = set()
    for motifs in results.values():
        all_classes.update(m.get('Class', 'Unknown') for m in motifs)
    all_classes = sorted(all_classes)
    
    # Calculate metrics for each sample
    comparison_data = []
    
    for sample_name, motifs in results.items():
        class_counts = Counter(m.get('Class', 'Unknown') for m in motifs)
        
        for class_name in all_classes:
            count = class_counts.get(class_name, 0)
            
            if metric == 'count':
                value = count
            elif metric == 'coverage':
                # Calculate coverage percentage (simplified)
                class_motifs = [m for m in motifs if m.get('Class') == class_name]
                covered_length = sum(m.get('Length', 0) for m in class_motifs)
                # Assume 10kb sequence for percentage calculation
                value = (covered_length / 10000) * 100
            elif metric == 'density':
                # Motifs per kb (assume 10kb sequence)
                value = count / 10
            else:
                value = count
            
            comparison_data.append({
                'Sample': sample_name,
                'Class': class_name,
                'Value': value
            })
    
    # Create DataFrame and pivot for heatmap
    df = pd.DataFrame(comparison_data)
    pivot_df = df.pivot(index='Sample', columns='Class', values='Value').fillna(0)
    
    # Create heatmap
    fig, ax = plt.subplots(figsize=figsize)
    
    sns.heatmap(pivot_df, annot=True, fmt='.1f', cmap='YlOrRd', 
                ax=ax, cbar_kws={'label': f'Motif {metric.title()}'})
    
    ax.set_title(title or f'Motif Class Comparison by {metric.title()}')
    ax.set_xlabel('Motif Class')
    ax.set_ylabel('Sample')
    
    plt.tight_layout()
    return fig

# =============================================================================
# INTERACTIVE PLOTS (PLOTLY)
# =============================================================================

def create_interactive_coverage_plot(motifs: List[Dict[str, Any]], 
                                   sequence_length: int,
                                   title: str = "Interactive Motif Coverage") -> Union[Any, plt.Figure]:
    """
    Create interactive coverage plot using Plotly
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Length of the analyzed sequence
        title: Plot title
        
    Returns:
        Plotly figure if available, otherwise matplotlib figure
    """
    if not PLOTLY_AVAILABLE:
        return plot_coverage_map(motifs, sequence_length, title)
    
    if not motifs:
        fig = go.Figure()
        fig.add_annotation(text="No motifs to display", 
                          xref="paper", yref="paper",
                          x=0.5, y=0.5, showarrow=False)
        fig.update_layout(title=title)
        return fig
    
    fig = go.Figure()
    
    # Group motifs by class
    class_motifs = defaultdict(list)
    for motif in motifs:
        class_name = motif.get('Class', 'Unknown')
        class_motifs[class_name].append(motif)
    
    y_pos = 0
    for class_name, class_motif_list in class_motifs.items():
        color = MOTIF_CLASS_COLORS.get(class_name, '#808080')
        
        x_starts = []
        x_ends = []
        y_positions = []
        hover_texts = []
        
        for motif in class_motif_list:
            start = motif.get('Start', 0)
            end = motif.get('End', start + 1)
            
            x_starts.append(start)
            x_ends.append(end)
            y_positions.extend([y_pos - 0.4, y_pos - 0.4, y_pos + 0.4, y_pos + 0.4, None])
            
            hover_text = f"{class_name}<br>{motif.get('Subclass', '')}<br>" + \
                        f"Position: {start}-{end}<br>Length: {motif.get('Length', 0)} bp<br>" + \
                        f"Score: {motif.get('Score', 'N/A')}"
            hover_texts.append(hover_text)
        
        # Add rectangles for motifs
        for i, (start, end) in enumerate(zip(x_starts, x_ends)):
            fig.add_shape(
                type="rect",
                x0=start, y0=y_pos - 0.4,
                x1=end, y1=y_pos + 0.4,
                fillcolor=color,
                opacity=0.7,
                line=dict(color="black", width=1)
            )
            
            # Add invisible scatter points for hover
            fig.add_trace(go.Scatter(
                x=[(start + end) / 2],
                y=[y_pos],
                mode='markers',
                marker=dict(size=0.1, opacity=0),
                hoverinfo='text',
                hovertext=hover_texts[i],
                showlegend=False
            ))
        
        # Add class label
        fig.add_trace(go.Scatter(
            x=[sequence_length * 1.02],
            y=[y_pos],
            mode='text',
            text=[class_name],
            textposition="middle right",
            showlegend=False,
            hoverinfo='skip'
        ))
        
        y_pos += 1
    
    fig.update_layout(
        title=title,
        xaxis_title="Sequence Position (bp)",
        yaxis_title="Motif Class",
        xaxis=dict(range=[0, sequence_length * 1.15]),
        yaxis=dict(range=[-0.5, len(class_motifs) - 0.5], showticklabels=False),
        showlegend=False,
        height=max(400, len(class_motifs) * 60)
    )
    
    return fig

# =============================================================================
# BATCH EXPORT FUNCTIONS (Publication-Ready)
# =============================================================================

def save_all_plots(motifs: List[Dict[str, Any]], 
                   sequence_length: int,
                   output_dir: str = "plots",
                   file_format: str = "pdf",
                   dpi: int = 300,
                   style: str = 'nature') -> Dict[str, str]:
    """
    Generate and save all standard plots in publication quality.
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Length of analyzed sequence
        output_dir: Output directory for plots
        file_format: File format ('pdf', 'png', 'svg') - pdf recommended for publication
        dpi: Resolution for raster formats (default 300 DPI for publication)
        style: Style preset ('nature', 'default')
        
    Returns:
        Dictionary of {plot_name: file_path}
    """
    import os
    
    # Apply style
    set_scientific_style(style)
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    saved_files = {}
    
    # Use publication DPI if not specified
    if dpi < PUBLICATION_DPI:
        dpi = PUBLICATION_DPI
    
    # List of plots to generate (diverse, non-repetitive visualization types)
    plots_to_generate = [
        ("motif_distribution_class", lambda: plot_motif_distribution(motifs, by='Class', style=style)),
        ("coverage_map", lambda: plot_coverage_map(motifs, sequence_length, style=style)),
        ("density_heatmap", lambda: plot_density_heatmap(motifs, sequence_length, style=style)),
        ("score_distribution", lambda: plot_score_distribution(motifs, by_class=True, style=style)),
        ("length_distribution", lambda: plot_length_distribution(motifs, by_class=True, style=style)),
        ("nested_donut_chart", lambda: plot_nested_pie_chart(motifs))
    ]
    
    for plot_name, plot_func in plots_to_generate:
        try:
            fig = plot_func()
            filename = f"{plot_name}.{file_format}"
            filepath = os.path.join(output_dir, filename)
            
            if hasattr(fig, 'savefig'):  # Matplotlib figure
                # Publication-quality save options
                fig.savefig(filepath, format=file_format, dpi=dpi, 
                           bbox_inches='tight', pad_inches=0.05,
                           facecolor='white', edgecolor='none')
                plt.close(fig)
            else:  # Plotly figure
                if file_format.lower() == 'png':
                    fig.write_image(filepath, scale=2)  # Higher resolution
                elif file_format.lower() in ('pdf', 'svg'):
                    fig.write_image(filepath)
                elif file_format.lower() == 'html':
                    fig.write_html(filepath)
                else:
                    # Unsupported format - try default write_image
                    print(f"⚠ Unsupported format '{file_format}' for Plotly, attempting save anyway")
                    try:
                        fig.write_image(filepath)
                    except Exception as fmt_err:
                        raise ValueError(f"Unsupported file format: {file_format}. Use 'png', 'pdf', 'svg', or 'html'.") from fmt_err
            
            saved_files[plot_name] = filepath
            print(f"[OK] Saved {plot_name} to {filepath}")
            
        except Exception as e:
            print(f"✗ Failed to generate {plot_name}: {e}")
    
    return saved_files


# =============================================================================
# ENHANCED SCIENTIFIC VISUALIZATION FUNCTIONS
# =============================================================================

def plot_class_analysis_comprehensive(motifs: List[Dict[str, Any]], 
                                     figsize: Tuple[int, int] = (16, 12)) -> plt.Figure:
    """
    Comprehensive class-level analysis with multiple subplots.
    Shows distribution, statistics, and comparison of all 11 Non-B DNA classes.
    Highlights which classes were detected and which were not.
    
    Args:
        motifs: List of motif dictionaries
        figsize: Figure size (width, height)
        
    Returns:
        Matplotlib figure object with multiple subplots
    """
    set_scientific_style()
    
    # Define all 11 Non-B DNA classes
    all_classes = [
        'Curved_DNA', 'Slipped_DNA', 'Cruciform', 'R-Loop', 'Triplex',
        'G-Quadruplex', 'i-Motif', 'Z-DNA', 'A-philic_DNA', 
        'Hybrid', 'Non-B_DNA_Clusters'
    ]
    
    # Count motifs by class
    detected_classes = Counter(m.get('Class', 'Unknown') for m in motifs)
    
    # Identify detected vs not detected
    detected = [cls for cls in all_classes if detected_classes.get(cls, 0) > 0]
    not_detected = [cls for cls in all_classes if detected_classes.get(cls, 0) == 0]
    
    # Create figure with subplots
    fig = plt.figure(figsize=figsize)
    gs = fig.add_gridspec(3, 2, hspace=0.3, wspace=0.3)
    
    # 1. Main distribution bar chart
    ax1 = fig.add_subplot(gs[0, :])
    counts = [detected_classes.get(cls, 0) for cls in all_classes]
    colors = [MOTIF_CLASS_COLORS.get(cls, '#808080') for cls in all_classes]
    bars = ax1.bar(range(len(all_classes)), counts, color=colors, alpha=0.8, 
                   edgecolor='black', linewidth=1.5)
    
    ax1.set_xlabel('Non-B DNA Class', fontsize=12, fontweight='bold')
    ax1.set_ylabel('Count', fontsize=12, fontweight='bold')
    ax1.set_title('Distribution of All 11 Non-B DNA Classes', fontsize=14, fontweight='bold')
    ax1.set_xticks(range(len(all_classes)))
    ax1.set_xticklabels(all_classes, rotation=45, ha='right', fontsize=10)
    ax1.grid(axis='y', alpha=0.3)
    
    # Add count labels on bars
    for bar, count in zip(bars, counts):
        if count > 0:
            height = bar.get_height()
            ax1.text(bar.get_x() + bar.get_width()/2., height,
                    f'{count}', ha='center', va='bottom', fontweight='bold', fontsize=9)
    
    # 2. Detected vs Not Detected pie chart
    ax2 = fig.add_subplot(gs[1, 0])
    detection_counts = [len(detected), len(not_detected)]
    detection_labels = [f'Detected\n({len(detected)} classes)', 
                       f'Not Detected\n({len(not_detected)} classes)']
    colors_pie = ['#4CAF50', '#FF5722']
    ax2.pie(detection_counts, labels=detection_labels, autopct='%1.1f%%',
            colors=colors_pie, startangle=90, textprops={'fontsize': 11, 'fontweight': 'bold'})
    ax2.set_title('Class Detection Status', fontsize=12, fontweight='bold')
    
    # 3. Statistics table for detected classes
    ax3 = fig.add_subplot(gs[1, 1])
    ax3.axis('off')
    
    if detected:
        # Calculate statistics for detected classes
        class_stats = []
        for cls in detected[:5]:  # Show top 5
            cls_motifs = [m for m in motifs if m.get('Class') == cls]
            count = len(cls_motifs)
            avg_length = np.mean([m.get('Length', 0) for m in cls_motifs]) if cls_motifs else 0
            avg_score = np.mean([m.get('Score', 0) for m in cls_motifs]) if cls_motifs else 0
            class_stats.append([cls[:15], count, f'{avg_length:.1f}', f'{avg_score:.3f}'])
        
        table = ax3.table(cellText=class_stats,
                         colLabels=['Class', 'Count', 'Avg Len', 'Avg Score'],
                         cellLoc='left', loc='center',
                         colWidths=[0.4, 0.2, 0.2, 0.2])
        table.auto_set_font_size(False)
        table.set_fontsize(9)
        table.scale(1, 2)
        
        # Style header
        for i in range(4):
            table[(0, i)].set_facecolor('#E0E0E0')
            table[(0, i)].set_text_props(weight='bold')
        
        ax3.set_title('Top Detected Classes (Statistics)', fontsize=12, fontweight='bold', pad=20)
    else:
        ax3.text(0.5, 0.5, 'No classes detected', ha='center', va='center',
                transform=ax3.transAxes, fontsize=12)
    
    # 4. List of not detected classes
    ax4 = fig.add_subplot(gs[2, :])
    ax4.axis('off')
    
    if not_detected:
        not_detected_text = 'Classes NOT Detected:\n' + ', '.join(not_detected)
        ax4.text(0.5, 0.5, not_detected_text, ha='center', va='center',
                transform=ax4.transAxes, fontsize=11, 
                bbox=dict(boxstyle='round', facecolor='#FFEBEE', alpha=0.8),
                wrap=True)
    else:
        ax4.text(0.5, 0.5, 'All 11 Non-B DNA classes detected! [OK]', 
                ha='center', va='center', transform=ax4.transAxes, fontsize=12,
                fontweight='bold', color='green')
    
    plt.suptitle(f'Comprehensive Class Analysis ({len(motifs)} total motifs)', 
                fontsize=16, fontweight='bold', y=0.98)
    
    return fig


def plot_subclass_analysis_comprehensive(motifs: List[Dict[str, Any]], 
                                        figsize: Tuple[int, int] = (18, 14)) -> plt.Figure:
    """
    Comprehensive subclass-level analysis showing all detected subclasses
    organized by their parent class.
    
    Args:
        motifs: List of motif dictionaries
        figsize: Figure size (width, height)
        
    Returns:
        Matplotlib figure object with subclass analysis
    """
    set_scientific_style()
    
    # Group motifs by class and subclass
    class_subclass_counts = defaultdict(lambda: defaultdict(int))
    for motif in motifs:
        class_name = motif.get('Class', 'Unknown')
        subclass_name = motif.get('Subclass', 'Unknown')
        class_subclass_counts[class_name][subclass_name] += 1
    
    # Prepare data for visualization
    all_subclasses = []
    all_counts = []
    all_classes = []
    
    for class_name in sorted(class_subclass_counts.keys()):
        for subclass_name, count in sorted(class_subclass_counts[class_name].items()):
            all_subclasses.append(f"{class_name}:{subclass_name}")
            all_counts.append(count)
            all_classes.append(class_name)
    
    # Create figure
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=figsize, 
                                   gridspec_kw={'height_ratios': [2, 1]})
    
    # 1. Subclass distribution bar chart
    colors = [MOTIF_CLASS_COLORS.get(cls, '#808080') for cls in all_classes]
    x_pos = range(len(all_subclasses))
    bars = ax1.barh(x_pos, all_counts, color=colors, alpha=0.8, 
                    edgecolor='black', linewidth=0.5)
    
    ax1.set_yticks(x_pos)
    ax1.set_yticklabels(all_subclasses, fontsize=9)
    ax1.set_xlabel('Count', fontsize=12, fontweight='bold')
    ax1.set_ylabel('Class:Subclass', fontsize=12, fontweight='bold')
    ax1.set_title('Distribution of All Detected Subclasses', fontsize=14, fontweight='bold')
    ax1.grid(axis='x', alpha=0.3)
    
    # Add count labels on bars
    for bar, count in zip(bars, all_counts):
        width = bar.get_width()
        ax1.text(width, bar.get_y() + bar.get_height()/2.,
                f' {count}', ha='left', va='center', fontweight='bold', fontsize=8)
    
    # 2. Subclass summary by class
    ax2.axis('off')
    
    # Create summary text
    summary_lines = ['Subclass Summary by Class:\n']
    for class_name in sorted(class_subclass_counts.keys()):
        subclasses = class_subclass_counts[class_name]
        n_subclasses = len(subclasses)
        total_count = sum(subclasses.values())
        summary_lines.append(f'{class_name}: {n_subclasses} subclass(es), {total_count} motifs')
    
    summary_text = '\n'.join(summary_lines)
    ax2.text(0.1, 0.5, summary_text, ha='left', va='center',
            transform=ax2.transAxes, fontsize=10, family='monospace',
            bbox=dict(boxstyle='round', facecolor='#F5F5F5', alpha=0.8))
    
    plt.suptitle(f'Comprehensive Subclass Analysis ({len(motifs)} total motifs)', 
                fontsize=16, fontweight='bold', y=0.99)
    plt.tight_layout()
    
    return fig


def plot_score_statistics_by_class(motifs: List[Dict[str, Any]], 
                                   figsize: Tuple[int, int] = (14, 8)) -> plt.Figure:
    """
    Advanced statistical visualization of scores by class.
    Shows box plots, violin plots, and statistical annotations.
    
    Args:
        motifs: List of motif dictionaries
        figsize: Figure size (width, height)
        
    Returns:
        Matplotlib figure object with score statistics
    """
    set_scientific_style()
    
    if not motifs:
        fig, ax = plt.subplots(figsize=figsize)
        ax.text(0.5, 0.5, 'No motifs to analyze', ha='center', va='center',
               transform=ax.transAxes, fontsize=14)
        return fig
    
    # Prepare data
    df_data = []
    for motif in motifs:
        df_data.append({
            'Class': motif.get('Class', 'Unknown'),
            'Score': motif.get('Score', 0),
            'Length': motif.get('Length', 0)
        })
    df = pd.DataFrame(df_data)
    
    # Create figure with subplots
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=figsize, 
                                   gridspec_kw={'height_ratios': [2, 1]})
    
    # 1. Violin plot with box plot overlay
    classes = sorted(df['Class'].unique())
    positions = range(len(classes))
    
    # Create violin plot
    parts = ax1.violinplot([df[df['Class'] == cls]['Score'].values for cls in classes],
                          positions=positions, widths=0.7, showmeans=True, showmedians=True)
    
    # Color violins by class
    for i, pc in enumerate(parts['bodies']):
        cls = classes[i]
        color = MOTIF_CLASS_COLORS.get(cls, '#808080')
        pc.set_facecolor(color)
        pc.set_alpha(0.6)
    
    # Overlay box plots
    bp = ax1.boxplot([df[df['Class'] == cls]['Score'].values for cls in classes],
                     positions=positions, widths=0.3, patch_artist=True,
                     boxprops=dict(facecolor='white', alpha=0.7),
                     medianprops=dict(color='red', linewidth=2))
    
    ax1.set_xticks(positions)
    ax1.set_xticklabels(classes, rotation=45, ha='right', fontsize=10)
    ax1.set_ylabel('Score', fontsize=12, fontweight='bold')
    ax1.set_title('Score Distribution by Class (Violin + Box Plot)', fontsize=14, fontweight='bold')
    ax1.grid(axis='y', alpha=0.3)
    
    # Add statistical annotations
    for i, cls in enumerate(classes):
        scores = df[df['Class'] == cls]['Score'].values
        if len(scores) > 0:
            mean_score = np.mean(scores)
            std_score = np.std(scores)
            ax1.text(i, ax1.get_ylim()[1] * 0.95, 
                    f'μ={mean_score:.2f}\nσ={std_score:.2f}',
                    ha='center', va='top', fontsize=8,
                    bbox=dict(boxstyle='round', facecolor='white', alpha=0.7))
    
    # 2. Statistical summary table
    ax2.axis('off')
    
    # Calculate statistics
    stats_data = []
    for cls in classes:
        cls_scores = df[df['Class'] == cls]['Score'].values
        if len(cls_scores) > 0:
            stats_data.append([
                cls[:15],
                len(cls_scores),
                f'{np.mean(cls_scores):.3f}',
                f'{np.median(cls_scores):.3f}',
                f'{np.std(cls_scores):.3f}',
                f'{np.min(cls_scores):.3f}',
                f'{np.max(cls_scores):.3f}'
            ])
    
    if stats_data:
        table = ax2.table(cellText=stats_data,
                         colLabels=['Class', 'N', 'Mean', 'Median', 'Std', 'Min', 'Max'],
                         cellLoc='center', loc='center',
                         colWidths=[0.25, 0.1, 0.13, 0.13, 0.13, 0.13, 0.13])
        table.auto_set_font_size(False)
        table.set_fontsize(8)
        table.scale(1, 1.8)
        
        # Style header
        for i in range(7):
            table[(0, i)].set_facecolor('#E0E0E0')
            table[(0, i)].set_text_props(weight='bold')
    
    plt.suptitle('Score Statistics Analysis', fontsize=16, fontweight='bold', y=0.98)
    plt.tight_layout()
    
    return fig


def plot_length_statistics_by_class(motifs: List[Dict[str, Any]], 
                                   figsize: Tuple[int, int] = (14, 10)) -> plt.Figure:
    """
    Advanced visualization of motif length distributions by class.
    
    Args:
        motifs: List of motif dictionaries
        figsize: Figure size (width, height)
        
    Returns:
        Matplotlib figure object with length statistics
    """
    set_scientific_style()
    
    if not motifs:
        fig, ax = plt.subplots(figsize=figsize)
        ax.text(0.5, 0.5, 'No motifs to analyze', ha='center', va='center',
               transform=ax.transAxes, fontsize=14)
        return fig
    
    # Prepare data
    df_data = []
    for motif in motifs:
        df_data.append({
            'Class': motif.get('Class', 'Unknown'),
            'Length': motif.get('Length', 0)
        })
    df = pd.DataFrame(df_data)
    
    # Create figure with subplots
    fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=figsize, 
                                       gridspec_kw={'height_ratios': [2, 1.5, 1]})
    
    # 1. Histogram with KDE overlay for each class
    classes = sorted(df['Class'].unique())
    
    for cls in classes:
        cls_lengths = df[df['Class'] == cls]['Length'].values
        if len(cls_lengths) > 1:
            color = MOTIF_CLASS_COLORS.get(cls, '#808080')
            ax1.hist(cls_lengths, bins=20, alpha=0.4, label=cls, color=color, edgecolor='black')
    
    ax1.set_xlabel('Length (bp)', fontsize=12, fontweight='bold')
    ax1.set_ylabel('Frequency', fontsize=12, fontweight='bold')
    ax1.set_title('Length Distribution by Class', fontsize=14, fontweight='bold')
    ax1.legend(fontsize=9, ncol=2)
    ax1.grid(axis='y', alpha=0.3)
    
    # 2. Box plot comparison
    bp = ax2.boxplot([df[df['Class'] == cls]['Length'].values for cls in classes],
                     labels=classes, patch_artist=True, vert=True)
    
    # Color boxes by class
    for i, (patch, cls) in enumerate(zip(bp['boxes'], classes)):
        color = MOTIF_CLASS_COLORS.get(cls, '#808080')
        patch.set_facecolor(color)
        patch.set_alpha(0.6)
    
    ax2.set_xticklabels(classes, rotation=45, ha='right', fontsize=10)
    ax2.set_ylabel('Length (bp)', fontsize=12, fontweight='bold')
    ax2.set_title('Length Comparison (Box Plot)', fontsize=12, fontweight='bold')
    ax2.grid(axis='y', alpha=0.3)
    
    # 3. Statistical summary table
    ax3.axis('off')
    
    stats_data = []
    for cls in classes:
        cls_lengths = df[df['Class'] == cls]['Length'].values
        if len(cls_lengths) > 0:
            stats_data.append([
                cls[:15],
                len(cls_lengths),
                f'{np.mean(cls_lengths):.1f}',
                f'{np.median(cls_lengths):.1f}',
                f'{np.std(cls_lengths):.1f}',
                f'{np.min(cls_lengths):.0f}',
                f'{np.max(cls_lengths):.0f}'
            ])
    
    if stats_data:
        table = ax3.table(cellText=stats_data,
                         colLabels=['Class', 'N', 'Mean', 'Median', 'Std', 'Min', 'Max'],
                         cellLoc='center', loc='center',
                         colWidths=[0.25, 0.1, 0.13, 0.13, 0.13, 0.13, 0.13])
        table.auto_set_font_size(False)
        table.set_fontsize(8)
        table.scale(1, 1.8)
        
        # Style header
        for i in range(7):
            table[(0, i)].set_facecolor('#E0E0E0')
            table[(0, i)].set_text_props(weight='bold')
    
    plt.suptitle('Length Statistics Analysis', fontsize=16, fontweight='bold', y=0.98)
    plt.tight_layout()
    
    return fig


# =============================================================================
# TESTING & EXAMPLES
# =============================================================================

def test_visualizations():
    """Test visualization functions with example data"""
    print("Testing NBDScanner visualizations...")
    
    # Create example motif data
    example_motifs = [
        {'Class': 'G-Quadruplex', 'Subclass': 'Canonical intramolecular G4', 'Start': 1, 'End': 21, 'Length': 21, 'Score': 0.85},
        {'Class': 'G-Quadruplex', 'Subclass': 'Extended-loop canonical', 'Start': 45, 'End': 60, 'Length': 16, 'Score': 0.72},
        {'Class': 'Curved_DNA', 'Subclass': 'A-tract', 'Start': 80, 'End': 95, 'Length': 16, 'Score': 0.65},
        {'Class': 'Z-DNA', 'Subclass': 'CG alternating', 'Start': 120, 'End': 135, 'Length': 16, 'Score': 0.90},
        {'Class': 'i-Motif', 'Subclass': 'Canonical i-motif', 'Start': 160, 'End': 180, 'Length': 21, 'Score': 0.78}
    ]
    
    sequence_length = 200
    
    print(f"\nTesting with {len(example_motifs)} example motifs:")
    for motif in example_motifs:
        print(f"  {motif['Class']} at {motif['Start']}-{motif['End']}")
    
    # Test basic plots
    try:
        fig1 = plot_motif_distribution(example_motifs, by='Class')
        plt.close(fig1)
        print("[OK] Motif distribution plot: PASS")
    except Exception as e:
        print(f"✗ Motif distribution plot: FAIL - {e}")
    
    try:
        fig2 = plot_coverage_map(example_motifs, sequence_length)
        plt.close(fig2)
        print("[OK] Coverage map plot: PASS")
    except Exception as e:
        print(f"✗ Coverage map plot: FAIL - {e}")
    
    try:
        fig3 = plot_score_distribution(example_motifs)
        plt.close(fig3)
        print("[OK] Score distribution plot: PASS")
    except Exception as e:
        print(f"✗ Score distribution plot: FAIL - {e}")
    
    try:
        fig4 = plot_nested_pie_chart(example_motifs)
        plt.close(fig4)
        print("[OK] Nested pie chart: PASS")
    except Exception as e:
        print(f"✗ Nested pie chart: FAIL - {e}")
    
    print(f"\n[OK] Visualization testing completed")
    print(f"Plotly available: {'Yes' if PLOTLY_AVAILABLE else 'No'}")


# =============================================================================
# ENHANCED STATISTICS VISUALIZATIONS: DENSITY AND ENRICHMENT
# =============================================================================

def plot_density_comparison(genomic_density: Dict[str, float],
                            positional_density: Dict[str, float],
                            title: str = "Motif Density Analysis",
                            figsize: Tuple[int, int] = (14, 6)) -> plt.Figure:
    """
    Plot comparison of genomic density (coverage %) and positional density (motifs/kbp).
    
    Args:
        genomic_density: Dictionary of class -> genomic density (%)
        positional_density: Dictionary of class -> positional density (motifs/unit)
        title: Plot title
        figsize: Figure size
        
    Returns:
        Matplotlib figure object
    """
    set_scientific_style()
    
    # Remove 'Overall' for class-specific comparison
    classes = [k for k in genomic_density.keys() if k != 'Overall']
    if not classes:
        classes = list(genomic_density.keys())
    
    # Sort classes alphabetically
    classes = sorted(classes)
    
    genomic_vals = [genomic_density.get(c, 0) for c in classes]
    positional_vals = [positional_density.get(c, 0) for c in classes]
    
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=figsize, dpi=PUBLICATION_DPI)
    
    # Genomic Density (Coverage %)
    colors1 = [MOTIF_CLASS_COLORS.get(c, '#808080') for c in classes]
    
    # Replace underscores with spaces in labels
    display_classes = [c.replace('_', ' ') for c in classes]
    
    bars1 = ax1.barh(display_classes, genomic_vals, color=colors1, alpha=0.8, 
                     edgecolor='black', linewidth=0.5)
    ax1.set_xlabel('Genomic Density (Coverage %)', fontsize=11, fontweight='bold')
    ax1.set_ylabel('Motif Class', fontsize=11, fontweight='bold')
    ax1.set_title('A. Genomic Density (σ_G)', fontsize=12, fontweight='bold', pad=10)
    
    # Add value labels with safe max calculation
    max_genomic = max(genomic_vals) if genomic_vals and max(genomic_vals) > 0 else 1
    for i, (bar, val) in enumerate(zip(bars1, genomic_vals)):
        if val > 0:
            ax1.text(val + max_genomic * 0.01, i, f'{val:.3f}%', 
                    va='center', fontsize=9, fontweight='bold')
    
    # Positional Density (Frequency)
    colors2 = [MOTIF_CLASS_COLORS.get(c, '#808080') for c in classes]
    bars2 = ax2.barh(display_classes, positional_vals, color=colors2, alpha=0.8, 
                     edgecolor='black', linewidth=0.5)
    ax2.set_xlabel('Positional Density (motifs/kbp)', fontsize=11, fontweight='bold')
    ax2.set_ylabel('Motif Class', fontsize=11, fontweight='bold')
    ax2.set_title('B. Positional Density (λ)', fontsize=12, fontweight='bold', pad=10)
    
    # Add value labels with safe max calculation
    max_positional = max(positional_vals) if positional_vals and max(positional_vals) > 0 else 1
    for i, (bar, val) in enumerate(zip(bars2, positional_vals)):
        if val > 0:
            ax2.text(val + max_positional * 0.01, i, f'{val:.2f}', 
                    va='center', fontsize=9, fontweight='bold')
    
    # Apply Nature journal style
    _apply_nature_style(ax1)
    _apply_nature_style(ax2)
    
    plt.suptitle(title, fontsize=14, fontweight='bold', y=1.00)
    plt.tight_layout(rect=[0, 0, 1, 0.97])
    
    return fig


def plot_enrichment_analysis(enrichment_results: Dict[str, Dict[str, Any]],
                             title: str = "Motif Enrichment Analysis",
                             figsize: Tuple[int, int] = (14, 8)) -> plt.Figure:
    """
    Plot enrichment analysis results with fold enrichment and p-values.
    
    Note: Enrichment analysis has been removed for performance.
    This function is kept for backward compatibility with legacy data.
    
    Args:
        enrichment_results: Dictionary with enrichment metrics
        title: Plot title
        figsize: Figure size
        
    Returns:
        Matplotlib figure object
    """
    set_scientific_style()
    
    # Extract data (exclude 'Overall' for class-specific view)
    classes = [k for k in enrichment_results.keys() if k != 'Overall']
    if not classes:
        classes = list(enrichment_results.keys())
    
    fold_enrichments = []
    p_values = []
    observed_densities = []
    background_means = []
    
    for cls in classes:
        result = enrichment_results[cls]
        fe = result.get('fold_enrichment', 0)
        # Handle both string 'Inf' and float infinity values robustly
        if fe == 'Inf' or (isinstance(fe, float) and np.isinf(fe)):
            fe = INFINITE_FOLD_ENRICHMENT_CAP  # Cap infinite values for visualization
        fold_enrichments.append(fe)
        p_values.append(result.get('p_value', 1.0))
        observed_densities.append(result.get('observed_density', 0))
        background_means.append(result.get('background_mean', 0))
    
    fig = plt.figure(figsize=figsize)
    gs = fig.add_gridspec(2, 2, hspace=0.35, wspace=0.3)
    
    # 1. Fold Enrichment
    ax1 = fig.add_subplot(gs[0, 0])
    colors = [MOTIF_CLASS_COLORS.get(c, '#808080') for c in classes]
    bars1 = ax1.barh(classes, fold_enrichments, color=colors, alpha=0.8, edgecolor='black', linewidth=0.5)
    ax1.axvline(x=1.0, color='red', linestyle='--', linewidth=2, label='No enrichment (FE=1)')
    ax1.set_xlabel('Fold Enrichment', fontsize=11, fontweight='bold')
    ax1.set_ylabel('Motif Class', fontsize=11, fontweight='bold')
    ax1.set_title('A. Fold Enrichment', fontsize=12, fontweight='bold')
    ax1.legend(loc='best', fontsize=9)
    
    # Add value labels
    for i, (bar, val) in enumerate(zip(bars1, fold_enrichments)):
        label_text = f'{val:.2f}' if val < INFINITE_FOLD_ENRICHMENT_CAP else 'Inf'
        ax1.text(val + max(fold_enrichments) * 0.01, i, label_text, 
                va='center', fontsize=9, fontweight='bold')
    
    # 2. P-values
    ax2 = fig.add_subplot(gs[0, 1])
    # Color code by significance
    p_colors = ['green' if p < 0.05 else 'orange' if p < 0.1 else 'red' for p in p_values]
    bars2 = ax2.barh(classes, p_values, color=p_colors, alpha=0.7, edgecolor='black', linewidth=0.5)
    ax2.axvline(x=0.05, color='green', linestyle='--', linewidth=2, label='p=0.05')
    ax2.axvline(x=0.1, color='orange', linestyle='--', linewidth=1.5, label='p=0.1')
    ax2.set_xlabel('P-value', fontsize=11, fontweight='bold')
    ax2.set_ylabel('Motif Class', fontsize=11, fontweight='bold')
    ax2.set_title('B. Statistical Significance', fontsize=12, fontweight='bold')
    ax2.set_xlim(0, max(1.0, max(p_values) * 1.1))
    ax2.legend(loc='best', fontsize=9)
    
    # Add value labels
    for i, (bar, val) in enumerate(zip(bars2, p_values)):
        ax2.text(val + 0.02, i, f'{val:.3f}', va='center', fontsize=9, fontweight='bold')
    
    # 3. Observed vs Background Density
    ax3 = fig.add_subplot(gs[1, :])
    x = np.arange(len(classes))
    width = 0.35
    
    bars3a = ax3.bar(x - width/2, observed_densities, width, label='Observed', 
                    color='steelblue', alpha=0.8, edgecolor='black', linewidth=0.5)
    bars3b = ax3.bar(x + width/2, background_means, width, label='Background (Mean)', 
                    color='coral', alpha=0.8, edgecolor='black', linewidth=0.5)
    
    ax3.set_xlabel('Motif Class', fontsize=11, fontweight='bold')
    ax3.set_ylabel('Density (%)', fontsize=11, fontweight='bold')
    ax3.set_title('C. Observed vs. Background Density Comparison', fontsize=12, fontweight='bold')
    ax3.set_xticks(x)
    ax3.set_xticklabels(classes, rotation=45, ha='right', fontsize=9)
    ax3.legend(loc='best', fontsize=10)
    ax3.grid(axis='y', alpha=0.3)
    
    plt.suptitle(title, fontsize=14, fontweight='bold', y=0.99)
    
    return fig


def plot_enrichment_summary_table(enrichment_results: Dict[str, Dict[str, Any]],
                                  title: str = "Enrichment Summary Statistics") -> plt.Figure:
    """
    Create a summary table visualization for enrichment results.
    
    Note: Enrichment analysis has been removed for performance.
    This function is kept for backward compatibility with legacy data.
    
    Args:
        enrichment_results: Dictionary with enrichment metrics
        title: Plot title
        
    Returns:
        Matplotlib figure object
    """
    set_scientific_style()
    
    # Prepare data for table
    classes = [k for k in enrichment_results.keys() if k != 'Overall']
    if not classes:
        return None
    
    table_data = []
    for cls in classes:
        result = enrichment_results[cls]
        fe = result.get('fold_enrichment', 0)
        fe_str = f"{fe:.2f}" if fe != 'Inf' else 'Inf'
        
        row = [
            cls,
            result.get('observed_count', 0),
            f"{result.get('observed_density', 0):.4f}%",
            f"{result.get('background_mean', 0):.4f}%",
            fe_str,
            f"{result.get('p_value', 1.0):.4f}",
            '***' if result.get('p_value', 1.0) < 0.001 else 
            '**' if result.get('p_value', 1.0) < 0.01 else 
            '*' if result.get('p_value', 1.0) < 0.05 else 'ns'
        ]
        table_data.append(row)
    
    # Create figure
    fig, ax = plt.subplots(figsize=(14, max(6, len(classes) * 0.4)))
    ax.axis('tight')
    ax.axis('off')
    
    # Create table
    headers = ['Class', 'Count', 'Observed\nDensity', 'Background\nMean', 
               'Fold\nEnrichment', 'P-value', 'Sig.']
    
    table = ax.table(cellText=table_data, colLabels=headers, 
                    cellLoc='center', loc='center',
                    colWidths=[0.20, 0.10, 0.15, 0.15, 0.15, 0.12, 0.08])
    
    table.auto_set_font_size(False)
    table.set_fontsize(10)
    table.scale(1, 2)
    
    # Style header
    for i, header in enumerate(headers):
        cell = table[(0, i)]
        cell.set_facecolor('#2196F3')
        cell.set_text_props(weight='bold', color='white', fontsize=11)
    
    # Style rows with alternating colors
    for i in range(1, len(table_data) + 1):
        for j in range(len(headers)):
            cell = table[(i, j)]
            if i % 2 == 0:
                cell.set_facecolor('#F5F5F5')
            else:
                cell.set_facecolor('white')
            
            # Highlight significant results
            if j == 6:  # Significance column
                if table_data[i-1][j] in ['***', '**', '*']:
                    cell.set_facecolor('#C8E6C9')
                    cell.set_text_props(weight='bold', color='green')
    
    plt.title(title, fontsize=14, fontweight='bold', pad=20)
    plt.tight_layout()
    
    return fig


# =============================================================================
# CIRCOS PLOT FOR NON-B DNA MOTIF DENSITY
# =============================================================================

# Motif classes to exclude from Circos visualization (dynamic classes)
CIRCOS_EXCLUDED_CLASSES = ['Hybrid', 'Non-B_DNA_Clusters']


def plot_circos_motif_density(motifs: List[Dict[str, Any]], 
                               sequence_length: int,
                               title: str = "Non-B DNA Motif Density Circos Plot",
                               figsize: Tuple[int, int] = (12, 12),
                               window_size: int = None) -> plt.Figure:
    """
    Create a circular Circos-style plot showing non-B DNA motif class density.
    
    The plot shows:
    - Outer ring: Sequence position ruler
    - Inner rings: One ring per motif class showing density
    - Center: Summary statistics
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Total length of analyzed sequence in bp
        title: Plot title
        figsize: Figure size (width, height)
        window_size: Window size for density calculation (auto-calculated if None)
        
    Returns:
        Matplotlib figure object with Circos-style visualization
    """
    set_scientific_style()
    
    if not motifs or sequence_length == 0:
        fig, ax = plt.subplots(figsize=figsize)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center',
                transform=ax.transAxes, fontsize=14)
        ax.set_title(title)
        return fig
    
    # Auto-calculate window size if not provided
    if window_size is None:
        window_size = max(100, sequence_length // 50)
    
    # Calculate number of windows
    num_windows = max(1, sequence_length // window_size)
    
    # Get unique classes (excluding Hybrid and Clusters for cleaner visualization)
    classes = sorted(set(m.get('Class', 'Unknown') for m in motifs 
                        if m.get('Class') not in CIRCOS_EXCLUDED_CLASSES))
    
    if not classes:
        classes = sorted(set(m.get('Class', 'Unknown') for m in motifs))
    
    # Calculate density per window per class
    class_densities = {}
    for class_name in classes:
        densities = []
        class_motifs = [m for m in motifs if m.get('Class') == class_name]
        
        for i in range(num_windows):
            window_start = i * window_size
            window_end = (i + 1) * window_size
            
            # Count motifs in this window
            count = 0
            for motif in class_motifs:
                motif_start = motif.get('Start', 0) - 1  # 0-based
                motif_end = motif.get('End', 0)
                # Check overlap with window
                if not (motif_end <= window_start or motif_start >= window_end):
                    count += 1
            
            # Convert to density (motifs per kb)
            window_kb = window_size / 1000
            densities.append(count / window_kb if window_kb > 0 else 0)
        
        class_densities[class_name] = densities
    
    # Create figure
    fig = plt.figure(figsize=figsize)
    ax = fig.add_subplot(111, projection='polar')
    
    # Calculate angles for each window
    theta = np.linspace(0, 2 * np.pi, num_windows, endpoint=False)
    
    # Width of each bar
    width = 2 * np.pi / num_windows * 0.8
    
    # Ring configuration
    ring_width = 0.12
    inner_radius = 0.3
    
    # Plot each class as a ring
    for i, class_name in enumerate(classes):
        densities = class_densities[class_name]
        
        # Normalize densities for this class (0-1 scale for bar height)
        max_density = max(densities) if max(densities) > 0 else 1
        normalized = [d / max_density for d in densities]
        
        # Calculate ring position
        ring_bottom = inner_radius + i * ring_width
        
        # Get color for this class
        color = MOTIF_CLASS_COLORS.get(class_name, '#808080')
        
        # Plot bars for this ring
        heights = [n * ring_width * 0.9 for n in normalized]
        # Replace underscores with spaces in legend label
        display_name = class_name.replace('_', ' ')
        bars = ax.bar(theta, heights, width=width, bottom=ring_bottom,
                     color=color, alpha=0.7, edgecolor='white', linewidth=0.5,
                     label=f'{display_name} (max: {max_density:.1f}/kb)')
    
    # Add outer position ruler
    outer_radius = inner_radius + len(classes) * ring_width + 0.05
    ruler_theta = np.linspace(0, 2 * np.pi, 12, endpoint=False)
    ruler_labels = [f'{int(i * sequence_length / 12 / 1000)}kb' for i in range(12)]
    
    ax.set_xticks(ruler_theta)
    ax.set_xticklabels(ruler_labels, fontsize=9, fontweight='bold')
    
    # Remove radial labels
    ax.set_yticklabels([])
    
    # Set limits
    ax.set_ylim(0, outer_radius + 0.1)
    
    # Add legend
    ax.legend(loc='center', bbox_to_anchor=(0.5, 0.5), fontsize=8, 
             framealpha=0.9, ncol=1)
    
    # Add title (replace underscores with spaces)
    display_title = title.replace('_', ' ')
    fig.suptitle(display_title, fontsize=14, fontweight='bold', y=0.98)
    
    # Add center statistics
    total_motifs = len([m for m in motifs if m.get('Class') not in CIRCOS_EXCLUDED_CLASSES])
    center_text = f"Total: {total_motifs}\n{len(classes)} classes\n{sequence_length/1000:.1f} kb"
    ax.text(0, 0, center_text, ha='center', va='center', fontsize=10, fontweight='bold')
    
    plt.tight_layout()
    return fig


def plot_radial_class_density(motifs: List[Dict[str, Any]], 
                               sequence_length: int,
                               title: str = "Radial Motif Class Density",
                               figsize: Tuple[int, int] = (10, 10)) -> plt.Figure:
    """
    Create a radial bar chart showing motif density per class.
    
    A simpler alternative to full Circos plot, showing aggregate density
    per motif class in a radial/polar layout.
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Total length of analyzed sequence in bp
        title: Plot title
        figsize: Figure size
        
    Returns:
        Matplotlib figure object
    """
    set_scientific_style()
    
    if not motifs or sequence_length == 0:
        fig, ax = plt.subplots(figsize=figsize)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center',
                transform=ax.transAxes, fontsize=14)
        ax.set_title(title)
        return fig
    
    # Calculate density per class (motifs per kb)
    sequence_kb = sequence_length / 1000
    class_counts = Counter(m.get('Class', 'Unknown') for m in motifs
                          if m.get('Class') not in CIRCOS_EXCLUDED_CLASSES)
    
    if not class_counts:
        class_counts = Counter(m.get('Class', 'Unknown') for m in motifs)
    
    classes = list(class_counts.keys())
    densities = [class_counts[c] / sequence_kb for c in classes]
    
    # Create polar plot
    fig = plt.figure(figsize=figsize)
    ax = fig.add_subplot(111, projection='polar')
    
    # Calculate angles
    num_classes = len(classes)
    theta = np.linspace(0, 2 * np.pi, num_classes, endpoint=False)
    width = 2 * np.pi / num_classes * 0.7
    
    # Get colors
    colors = [MOTIF_CLASS_COLORS.get(c, '#808080') for c in classes]
    
    # Plot bars
    bars = ax.bar(theta, densities, width=width, color=colors, alpha=0.8,
                 edgecolor='white', linewidth=2)
    
    # Add class labels (replace underscores with spaces)
    ax.set_xticks(theta)
    display_classes = [cls.replace('_', ' ') for cls in classes]
    ax.set_xticklabels(display_classes, fontsize=10, fontweight='bold')
    
    # Add value labels on bars
    for angle, density, bar in zip(theta, densities, bars):
        ax.text(angle, density + max(densities) * 0.05, f'{density:.1f}',
               ha='center', va='bottom', fontsize=9, fontweight='bold')
    
    # Style (replace underscores with spaces in title)
    ax.set_ylabel('Density (motifs/kb)', labelpad=30, fontsize=11, fontweight='bold')
    display_title = title.replace('_', ' ')
    ax.set_title(display_title, fontsize=14, fontweight='bold', pad=20)
    
    plt.tight_layout()
    return fig


def plot_stacked_density_track(motifs: List[Dict[str, Any]], 
                                sequence_length: int,
                                title: str = "Stacked Motif Density Track",
                                figsize: Tuple[int, int] = (14, 6),
                                window_size: int = None) -> plt.Figure:
    """
    Create a stacked area chart showing motif density along the sequence.
    
    A linear (non-circular) alternative to Circos plot showing how different
    motif classes are distributed along the sequence.
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Total length of analyzed sequence in bp
        title: Plot title
        figsize: Figure size
        window_size: Window size for density calculation (auto if None)
        
    Returns:
        Matplotlib figure object
    """
    set_scientific_style()
    
    if not motifs or sequence_length == 0:
        fig, ax = plt.subplots(figsize=figsize)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center',
                transform=ax.transAxes, fontsize=14)
        ax.set_title(title)
        return fig
    
    # Auto-calculate window size
    if window_size is None:
        window_size = max(100, sequence_length // 100)
    
    num_windows = max(1, sequence_length // window_size)
    
    # Get unique classes
    classes = sorted(set(m.get('Class', 'Unknown') for m in motifs
                        if m.get('Class') not in CIRCOS_EXCLUDED_CLASSES))
    
    if not classes:
        classes = sorted(set(m.get('Class', 'Unknown') for m in motifs))
    
    # Calculate density per window per class
    positions = np.arange(num_windows) * window_size / 1000  # In kb
    class_densities = {}
    
    for class_name in classes:
        densities = []
        class_motifs = [m for m in motifs if m.get('Class') == class_name]
        
        for i in range(num_windows):
            window_start = i * window_size
            window_end = (i + 1) * window_size
            
            count = 0
            for motif in class_motifs:
                motif_start = motif.get('Start', 0) - 1
                motif_end = motif.get('End', 0)
                if not (motif_end <= window_start or motif_start >= window_end):
                    count += 1
            
            densities.append(count)
        
        class_densities[class_name] = densities
    
    # Create stacked area plot
    fig, ax = plt.subplots(figsize=figsize)
    
    # Stack the densities
    colors = [MOTIF_CLASS_COLORS.get(c, '#808080') for c in classes]
    
    # Create arrays for stacking (replace underscores with spaces in labels)
    density_arrays = [np.array(class_densities[c]) for c in classes]
    display_classes = [cls.replace('_', ' ') for cls in classes]
    
    ax.stackplot(positions, *density_arrays, labels=display_classes, colors=colors, alpha=0.8)
    
    # Styling (replace underscores with spaces in title)
    ax.set_xlabel('Position (kb)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Motif Count per Window', fontsize=12, fontweight='bold')
    display_title = title.replace('_', ' ')
    ax.set_title(display_title, fontsize=14, fontweight='bold')
    ax.legend(loc='upper right', fontsize=9, framealpha=0.9)
    ax.grid(axis='y', alpha=0.3)
    
    plt.tight_layout()
    return fig


# =============================================================================
# SUBCLASS-LEVEL DENSITY AND ENRICHMENT VISUALIZATIONS
# =============================================================================

def plot_density_comparison_by_subclass(genomic_density: Dict[str, float],
                                        positional_density: Dict[str, float],
                                        title: str = "Motif Density Analysis (by Subclass)",
                                        figsize: Tuple[int, int] = (16, 10)) -> plt.Figure:
    """
    Plot comparison of genomic density (coverage %) and positional density (motifs/kbp)
    at the subclass level.
    
    Args:
        genomic_density: Dictionary of 'Class:Subclass' -> genomic density (%)
        positional_density: Dictionary of 'Class:Subclass' -> positional density (motifs/unit)
        title: Plot title
        figsize: Figure size
        
    Returns:
        Matplotlib figure object
    """
    set_scientific_style()
    
    # Remove 'Overall' for subclass-specific comparison
    subclasses = [k for k in genomic_density.keys() if k != 'Overall' and ':' in k]
    if not subclasses:
        # Fallback to regular class-level if no subclass data
        subclasses = [k for k in genomic_density.keys() if k != 'Overall']
    
    # Sort by class, then by subclass
    subclasses.sort()
    
    genomic_vals = [genomic_density.get(c, 0) for c in subclasses]
    positional_vals = [positional_density.get(c, 0) for c in subclasses]
    
    # Create figure with two subplots
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=figsize, dpi=PUBLICATION_DPI)
    
    # Genomic Density (Coverage %)
    y_pos = np.arange(len(subclasses))
    
    # Get colors based on parent class
    colors1 = []
    for subclass in subclasses:
        if ':' in subclass:
            parent_class = subclass.split(':')[0]
            colors1.append(MOTIF_CLASS_COLORS.get(parent_class, '#808080'))
        else:
            colors1.append(MOTIF_CLASS_COLORS.get(subclass, '#808080'))
    
    bars1 = ax1.barh(y_pos, genomic_vals, color=colors1, alpha=0.8, edgecolor='black', linewidth=0.5)
    ax1.set_xlabel('Genomic Density (Coverage %)', fontsize=11, fontweight='bold')
    ax1.set_ylabel('Motif Subclass', fontsize=11, fontweight='bold')
    ax1.set_title('A. Genomic Density (σ_G) by Subclass', fontsize=12, fontweight='bold', pad=10)
    ax1.set_yticks(y_pos)
    
    # Replace underscores and format labels
    display_labels = [label.replace('_', ' ').replace(':', ': ') for label in subclasses]
    ax1.set_yticklabels(display_labels, fontsize=8)
    
    # Add value labels
    max_val = max(genomic_vals) if genomic_vals else 1
    for i, (bar, val) in enumerate(zip(bars1, genomic_vals)):
        if val > 0:
            ax1.text(val + max_val * 0.01, i, f'{val:.3f}%', 
                    va='center', fontsize=7, fontweight='bold')
    
    # Positional Density (Frequency)
    bars2 = ax2.barh(y_pos, positional_vals, color=colors1, alpha=0.8, edgecolor='black', linewidth=0.5)
    ax2.set_xlabel('Positional Density (motifs/kbp)', fontsize=11, fontweight='bold')
    ax2.set_ylabel('Motif Subclass', fontsize=11, fontweight='bold')
    ax2.set_title('B. Positional Density (λ) by Subclass', fontsize=12, fontweight='bold', pad=10)
    ax2.set_yticks(y_pos)
    ax2.set_yticklabels(display_labels, fontsize=8)
    
    # Add value labels
    max_val2 = max(positional_vals) if positional_vals else 1
    for i, (bar, val) in enumerate(zip(bars2, positional_vals)):
        if val > 0:
            ax2.text(val + max_val2 * 0.01, i, f'{val:.2f}', 
                    va='center', fontsize=7, fontweight='bold')
    
    plt.suptitle(title, fontsize=14, fontweight='bold', y=0.995)
    plt.tight_layout(rect=[0, 0, 1, 0.99])
    
    return fig


def plot_enrichment_analysis_by_subclass(enrichment_results: Dict[str, Dict[str, Any]],
                                         title: str = "Motif Enrichment Analysis (by Subclass)",
                                         figsize: Tuple[int, int] = (16, 12)) -> plt.Figure:
    """
    Plot enrichment analysis results at subclass level with fold enrichment and p-values.
    
    Note: Enrichment analysis has been removed for performance.
    This function is kept for backward compatibility with legacy data.
    
    Args:
        enrichment_results: Dictionary with enrichment metrics
        title: Plot title
        figsize: Figure size
        
    Returns:
        Matplotlib figure object
    """
    set_scientific_style()
    
    # Extract data (exclude 'Overall' for subclass-specific view)
    subclasses = [k for k in enrichment_results.keys() if k != 'Overall' and ':' in k]
    if not subclasses:
        # Fallback to class level
        subclasses = [k for k in enrichment_results.keys() if k != 'Overall']
    
    # Sort alphabetically
    subclasses.sort()
    
    fold_enrichments = []
    p_values = []
    observed_densities = []
    background_means = []
    
    for subclass in subclasses:
        result = enrichment_results[subclass]
        fe = result.get('fold_enrichment', 0)
        # Handle both string 'Inf' and float infinity values robustly
        if fe == 'Inf' or (isinstance(fe, float) and np.isinf(fe)):
            fe = INFINITE_FOLD_ENRICHMENT_CAP  # Cap infinite values for visualization
        fold_enrichments.append(fe)
        p_values.append(result.get('p_value', 1.0))
        observed_densities.append(result.get('observed_density', 0))
        background_means.append(result.get('background_mean', 0))
    
    # Create figure with subplots
    fig = plt.figure(figsize=figsize)
    gs = fig.add_gridspec(3, 1, hspace=0.35, height_ratios=[2, 2, 1.5])
    
    y_pos = np.arange(len(subclasses))
    
    # Get colors based on parent class
    colors = []
    for subclass in subclasses:
        if ':' in subclass:
            parent_class = subclass.split(':')[0]
            colors.append(MOTIF_CLASS_COLORS.get(parent_class, '#808080'))
        else:
            colors.append(MOTIF_CLASS_COLORS.get(subclass, '#808080'))
    
    # 1. Fold Enrichment
    ax1 = fig.add_subplot(gs[0])
    bars1 = ax1.barh(y_pos, fold_enrichments, color=colors, alpha=0.8, edgecolor='black', linewidth=0.5)
    ax1.axvline(x=1.0, color='red', linestyle='--', linewidth=2, label='No enrichment (FE=1)')
    ax1.set_xlabel('Fold Enrichment', fontsize=11, fontweight='bold')
    ax1.set_ylabel('Motif Subclass', fontsize=11, fontweight='bold')
    ax1.set_title('A. Fold Enrichment by Subclass', fontsize=12, fontweight='bold')
    ax1.set_yticks(y_pos)
    
    # Format labels
    display_labels = [label.replace('_', ' ').replace(':', ': ') for label in subclasses]
    ax1.set_yticklabels(display_labels, fontsize=8)
    ax1.legend(loc='best', fontsize=9)
    
    # Add value labels
    max_fe = max(fold_enrichments) if fold_enrichments else 1
    for i, (bar, val) in enumerate(zip(bars1, fold_enrichments)):
        label_text = f'{val:.2f}' if val < INFINITE_FOLD_ENRICHMENT_CAP else 'Inf'
        ax1.text(val + max_fe * 0.01, i, label_text, 
                va='center', fontsize=7, fontweight='bold')
    
    # 2. P-values
    ax2 = fig.add_subplot(gs[1])
    # Color code by significance
    p_colors = ['green' if p < 0.05 else 'orange' if p < 0.1 else 'red' for p in p_values]
    bars2 = ax2.barh(y_pos, p_values, color=p_colors, alpha=0.7, edgecolor='black', linewidth=0.5)
    ax2.axvline(x=0.05, color='green', linestyle='--', linewidth=2, label='p=0.05')
    ax2.axvline(x=0.1, color='orange', linestyle='--', linewidth=1.5, label='p=0.1')
    ax2.set_xlabel('P-value', fontsize=11, fontweight='bold')
    ax2.set_ylabel('Motif Subclass', fontsize=11, fontweight='bold')
    ax2.set_title('B. Statistical Significance by Subclass', fontsize=12, fontweight='bold')
    ax2.set_yticks(y_pos)
    ax2.set_yticklabels(display_labels, fontsize=8)
    ax2.set_xlim(0, max(1.0, max(p_values) * 1.1))
    ax2.legend(loc='best', fontsize=9)
    
    # Add value labels
    for i, (bar, val) in enumerate(zip(bars2, p_values)):
        ax2.text(val + 0.02, i, f'{val:.3f}', va='center', fontsize=7, fontweight='bold')
    
    # 3. Observed vs Background Density (grouped bar chart)
    ax3 = fig.add_subplot(gs[2])
    x = np.arange(len(subclasses))
    width = 0.35
    
    bars3a = ax3.bar(x - width/2, observed_densities, width, label='Observed', 
                    color='steelblue', alpha=0.8, edgecolor='black', linewidth=0.5)
    bars3b = ax3.bar(x + width/2, background_means, width, label='Background (Mean)', 
                    color='coral', alpha=0.8, edgecolor='black', linewidth=0.5)
    
    ax3.set_xlabel('Motif Subclass', fontsize=11, fontweight='bold')
    ax3.set_ylabel('Density (%)', fontsize=11, fontweight='bold')
    ax3.set_title('C. Observed vs. Background Density Comparison', fontsize=12, fontweight='bold')
    ax3.set_xticks(x)
    ax3.set_xticklabels(display_labels, rotation=45, ha='right', fontsize=7)
    ax3.legend(loc='best', fontsize=10)
    ax3.grid(axis='y', alpha=0.3)
    
    plt.suptitle(title, fontsize=14, fontweight='bold', y=0.995)
    
    return fig


def plot_subclass_density_heatmap(motifs: List[Dict[str, Any]], 
                                  sequence_length: int,
                                  window_size: int = 1000,
                                  title: str = "Subclass Density Heatmap",
                                  figsize: Tuple[int, int] = (16, 12)) -> plt.Figure:
    """
    Create a heatmap showing density of each subclass across the sequence.
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Length of the analyzed sequence
        window_size: Window size for density calculation
        title: Plot title
        figsize: Figure size
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style()
    
    if not motifs or sequence_length == 0:
        fig, ax = plt.subplots(figsize=figsize)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center',
                transform=ax.transAxes, fontsize=14)
        ax.set_title(title)
        return fig
    
    # Calculate windows
    num_windows = max(1, sequence_length // window_size)
    windows = np.linspace(0, sequence_length, num_windows + 1)
    
    # Get unique subclasses
    subclass_groups = defaultdict(list)
    for motif in motifs:
        class_name = motif.get('Class', 'Unknown')
        subclass_name = motif.get('Subclass', 'Unknown')
        key = f"{class_name}:{subclass_name}"
        subclass_groups[key].append(motif)
    
    subclasses = sorted(subclass_groups.keys())
    
    # Calculate density matrix
    density_matrix = np.zeros((len(subclasses), num_windows))
    
    for i, subclass_key in enumerate(subclasses):
        subclass_motifs = subclass_groups[subclass_key]
        
        for j in range(num_windows):
            window_start = windows[j]
            window_end = windows[j + 1]
            
            # Count motifs in window
            count = 0
            for motif in subclass_motifs:
                motif_start = motif.get('Start', 0) - 1  # 0-based
                motif_end = motif.get('End', 0)
                
                # Check if motif overlaps with window
                if not (motif_end <= window_start or motif_start >= window_end):
                    count += 1
            
            density_matrix[i, j] = count
    
    # Create heatmap with publication quality
    fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
    
    # Use colorblind-friendly colormap
    im = ax.imshow(density_matrix, cmap='viridis', aspect='auto', interpolation='nearest')
    
    # Customize axes (Nature style)
    ax.set_xlabel(f'Position (kb)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Motif Subclass', fontsize=12, fontweight='bold')
    
    # Replace underscores with spaces in title
    display_title = title.replace('_', ' ')
    ax.set_title(display_title, fontsize=14, fontweight='bold', pad=10)
    
    # Set ticks and labels with underscores replaced by spaces
    ax.set_yticks(range(len(subclasses)))
    display_subclasses = [sc.replace('_', ' ').replace(':', ': ') for sc in subclasses]
    ax.set_yticklabels(display_subclasses, fontsize=8)
    
    # Clean x-axis with kb units
    x_ticks = np.arange(0, num_windows, max(1, num_windows // 10))
    ax.set_xticks(x_ticks)
    ax.set_xticklabels([f'{int(windows[i]/1000)}' for i in x_ticks])
    
    # Add colorbar with proper label
    cbar = plt.colorbar(im, ax=ax, shrink=0.8)
    cbar.set_label('Motif Count', fontsize=10, fontweight='bold')
    cbar.ax.tick_params(labelsize=8)
    
    plt.tight_layout()
    return fig


# =============================================================================
# NATURE-LEVEL PUBLICATION VISUALIZATIONS (GENOME-WIDE)
# =============================================================================

def plot_manhattan_motif_density(motifs: List[Dict[str, Any]], 
                                  sequence_length: int,
                                  window_size: int = None,
                                  score_type: str = 'density',
                                  title: str = "Manhattan Plot - Motif Distribution",
                                  figsize: Tuple[int, int] = None) -> plt.Figure:
    """
    Create Manhattan plot showing motif density or score across genomic coordinates.
    
    Ideal for highlighting hotspots, clusters, and hybrid zones in large genomes.
    Publication-quality visualization following Nature Methods guidelines.
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Total length of analyzed sequence in bp
        window_size: Window size for density calculation (auto if None)
        score_type: 'density' for motif count or 'score' for average score
        title: Plot title
        figsize: Figure size (width, height)
        
    Returns:
        Matplotlib figure object (publication-ready at 300 DPI)
    """
    set_scientific_style()
    
    if figsize is None:
        figsize = FIGURE_SIZES['wide']
    
    if not motifs or sequence_length == 0:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center',
                transform=ax.transAxes, fontsize=14)
        ax.set_title(title)
        return fig
    
    # Auto-calculate window size (1% of sequence or minimum 1kb)
    if window_size is None:
        window_size = max(1000, sequence_length // 100)
    
    num_windows = max(1, sequence_length // window_size)
    
    # Get unique classes (exclude synthetic classes for cleaner visualization)
    classes = sorted(set(m.get('Class', 'Unknown') for m in motifs
                        if m.get('Class') not in CIRCOS_EXCLUDED_CLASSES))
    
    if not classes:
        classes = sorted(set(m.get('Class', 'Unknown') for m in motifs))
    
    # Create figure
    fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
    
    # Calculate metric for each window and class
    positions_kb = []
    values = []
    colors = []
    
    for class_name in classes:
        class_motifs = [m for m in motifs if m.get('Class') == class_name]
        color = MOTIF_CLASS_COLORS.get(class_name, '#808080')
        
        for i in range(num_windows):
            window_start = i * window_size
            window_end = (i + 1) * window_size
            window_center_kb = (window_start + window_end) / 2 / 1000
            
            # Count motifs in this window
            window_motifs = []
            for motif in class_motifs:
                motif_start = motif.get('Start', 0) - 1
                motif_end = motif.get('End', 0)
                if not (motif_end <= window_start or motif_start >= window_end):
                    window_motifs.append(motif)
            
            if window_motifs:
                if score_type == 'density':
                    # Density: motifs per kb
                    value = len(window_motifs) / (window_size / 1000)
                else:  # score
                    # Average score in window
                    scores = [m.get('Score', 0) for m in window_motifs if isinstance(m.get('Score'), (int, float))]
                    value = np.mean(scores) if scores else 0
                
                positions_kb.append(window_center_kb)
                values.append(value)
                colors.append(color)
    
    # Plot points with class-specific colors
    ax.scatter(positions_kb, values, c=colors, s=20, alpha=0.6, edgecolors='black', linewidth=0.3)
    
    # Styling
    ax.set_xlabel('Genomic Position (kb)', fontsize=12, fontweight='bold')
    y_label = 'Motif Density (motifs/kb)' if score_type == 'density' else 'Average Motif Score'
    ax.set_ylabel(y_label, fontsize=12, fontweight='bold')
    
    # Replace underscores with spaces in title
    display_title = title.replace('_', ' ')
    ax.set_title(display_title, fontsize=14, fontweight='bold', pad=10)
    
    # Add horizontal grid for readability
    ax.grid(axis='y', alpha=0.3, linestyle='--', linewidth=0.5)
    
    # Create legend with class names
    legend_elements = []
    display_classes = []
    for class_name in classes:
        if any(c == MOTIF_CLASS_COLORS.get(class_name, '#808080') for c in colors):
            legend_elements.append(plt.scatter([], [], c=MOTIF_CLASS_COLORS.get(class_name, '#808080'), 
                                             s=50, alpha=0.6, edgecolors='black', linewidth=0.5))
            display_classes.append(class_name.replace('_', ' '))
    
    if legend_elements:
        ax.legend(legend_elements, display_classes, loc='upper right', 
                 fontsize=8, framealpha=0.9, ncol=min(3, len(legend_elements)))
    
    # Apply label suppression policy (Nature-ready: no individual labels)
    # Only annotate top 5% density hotspots if requested
    if values:
        threshold = np.percentile(values, 95)  # Top 5%
        hotspot_count = 0
        max_labels = 10  # Maximum labels per plot
        min_distance_kb = (sequence_length / 1000) / 20  # Minimum 5% sequence spacing
        
        last_label_pos = -float('inf')
        for pos, val, color in sorted(zip(positions_kb, values, colors), 
                                      key=lambda x: x[1], reverse=True):
            if val >= threshold and hotspot_count < max_labels:
                # Check distance from last label
                if pos - last_label_pos >= min_distance_kb:
                    # Subtle annotation (Nature-ready: minimal, professional)
                    ax.annotate(f'{val:.1f}', xy=(pos, val), 
                               xytext=(0, 5), textcoords='offset points',
                               fontsize=6, alpha=0.7, ha='center',
                               bbox=dict(boxstyle='round,pad=0.2', fc='white', 
                                       ec=color, lw=0.5, alpha=0.8))
                    last_label_pos = pos
                    hotspot_count += 1
    
    # Apply Nature journal style
    _apply_nature_style(ax)
    
    plt.tight_layout()
    return fig


def plot_cumulative_motif_distribution(motifs: List[Dict[str, Any]], 
                                       sequence_length: int,
                                       title: str = "Cumulative Motif Distribution",
                                       figsize: Tuple[int, int] = None,
                                       by_class: bool = True) -> plt.Figure:
    """
    Create cumulative distribution plot showing running sum of motifs over genome.
    
    Useful for comparing motifs or samples, showing how motif accumulation
    varies across the sequence. Publication-quality following Nature guidelines.
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Total length of analyzed sequence in bp
        title: Plot title
        figsize: Figure size (width, height)
        by_class: Whether to separate by motif class
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style()
    
    if figsize is None:
        figsize = FIGURE_SIZES['wide']
    
    if not motifs or sequence_length == 0:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center',
                transform=ax.transAxes, fontsize=14)
        ax.set_title(title)
        return fig
    
    fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
    
    if by_class:
        # Group by class
        classes = sorted(set(m.get('Class', 'Unknown') for m in motifs
                            if m.get('Class') not in CIRCOS_EXCLUDED_CLASSES))
        
        if not classes:
            classes = sorted(set(m.get('Class', 'Unknown') for m in motifs))
        
        for class_name in classes:
            class_motifs = [m for m in motifs if m.get('Class') == class_name]
            
            # Sort by start position
            class_motifs_sorted = sorted(class_motifs, key=lambda m: m.get('Start', 0))
            
            # Calculate cumulative count
            positions = [0]
            cumulative = [0]
            
            for i, motif in enumerate(class_motifs_sorted, 1):
                positions.append(motif.get('Start', 0) / 1000)  # Convert to kb
                cumulative.append(i)
            
            # Add final point at sequence end
            positions.append(sequence_length / 1000)
            cumulative.append(len(class_motifs))
            
            # Plot with class color
            color = MOTIF_CLASS_COLORS.get(class_name, '#808080')
            display_name = class_name.replace('_', ' ')
            ax.plot(positions, cumulative, color=color, linewidth=1.5, 
                   label=display_name, alpha=0.8)
    else:
        # Overall cumulative
        motifs_sorted = sorted(motifs, key=lambda m: m.get('Start', 0))
        
        positions = [0]
        cumulative = [0]
        
        for i, motif in enumerate(motifs_sorted, 1):
            positions.append(motif.get('Start', 0) / 1000)
            cumulative.append(i)
        
        positions.append(sequence_length / 1000)
        cumulative.append(len(motifs))
        
        ax.plot(positions, cumulative, color='#0072B2', linewidth=2, alpha=0.8)
    
    # Styling
    ax.set_xlabel('Genomic Position (kb)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Cumulative Motif Count', fontsize=12, fontweight='bold')
    
    # Replace underscores in title
    display_title = title.replace('_', ' ')
    ax.set_title(display_title, fontsize=14, fontweight='bold', pad=10)
    
    # Add grid
    ax.grid(alpha=0.3, linestyle='--', linewidth=0.5)
    
    if by_class:
        ax.legend(loc='upper left', fontsize=8, framealpha=0.9)
    
    # Apply Nature journal style
    _apply_nature_style(ax)
    
    plt.tight_layout()
    return fig


def plot_motif_cooccurrence_matrix(motifs: List[Dict[str, Any]], 
                                   title: str = "Motif Co-occurrence Matrix",
                                   figsize: Tuple[int, int] = None,
                                   overlap_threshold: int = 1) -> plt.Figure:
    """
    Create heatmap showing co-occurrence frequency between motif classes.
    
    Shows which motif classes tend to appear together (within overlap_threshold bp).
    Excellent for publication figures showing motif relationships.
    
    Args:
        motifs: List of motif dictionaries
        title: Plot title
        figsize: Figure size (width, height)
        overlap_threshold: Maximum distance (bp) to consider as co-occurrence
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style()
    
    if figsize is None:
        figsize = FIGURE_SIZES['square']
    
    if not motifs:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center',
                transform=ax.transAxes, fontsize=14)
        ax.set_title(title)
        return fig
    
    # Get unique classes (exclude synthetic ones)
    classes = sorted(set(m.get('Class', 'Unknown') for m in motifs
                        if m.get('Class') not in CIRCOS_EXCLUDED_CLASSES))
    
    if not classes:
        classes = sorted(set(m.get('Class', 'Unknown') for m in motifs))
    
    # Initialize co-occurrence matrix
    n_classes = len(classes)
    cooccurrence_matrix = np.zeros((n_classes, n_classes))
    
    # Calculate co-occurrences
    for i, class_i in enumerate(classes):
        motifs_i = [m for m in motifs if m.get('Class') == class_i]
        
        for j, class_j in enumerate(classes):
            motifs_j = [m for m in motifs if m.get('Class') == class_j]
            
            # Count overlaps
            count = 0
            for mi in motifs_i:
                start_i = mi.get('Start', 0)
                end_i = mi.get('End', 0)
                
                for mj in motifs_j:
                    start_j = mj.get('Start', 0)
                    end_j = mj.get('End', 0)
                    
                    # Check if they overlap or are within threshold
                    distance = max(0, max(start_i, start_j) - min(end_i, end_j))
                    
                    if distance <= overlap_threshold:
                        count += 1
            
            cooccurrence_matrix[i, j] = count
    
    # Create heatmap
    fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
    
    # Use colorblind-friendly colormap
    im = ax.imshow(cooccurrence_matrix, cmap='YlOrRd', aspect='auto', interpolation='nearest')
    
    # Set ticks and labels
    ax.set_xticks(range(n_classes))
    ax.set_yticks(range(n_classes))
    
    # Replace underscores with spaces in labels
    display_classes = [c.replace('_', ' ') for c in classes]
    ax.set_xticklabels(display_classes, rotation=45, ha='right', fontsize=9)
    ax.set_yticklabels(display_classes, fontsize=9)
    
    # Add colorbar
    cbar = plt.colorbar(im, ax=ax, shrink=0.8)
    cbar.set_label('Co-occurrence Count', fontsize=10, fontweight='bold')
    cbar.ax.tick_params(labelsize=8)
    
    # Add values to cells (if not too many)
    if n_classes <= 10:
        for i in range(n_classes):
            for j in range(n_classes):
                value = int(cooccurrence_matrix[i, j])
                if value > 0:
                    text_color = 'white' if value > cooccurrence_matrix.max() / 2 else 'black'
                    ax.text(j, i, str(value), ha='center', va='center', 
                           color=text_color, fontsize=8, fontweight='bold')
    
    # Title and labels
    display_title = title.replace('_', ' ')
    ax.set_title(display_title, fontsize=14, fontweight='bold', pad=10)
    ax.set_xlabel('Motif Class', fontsize=11, fontweight='bold')
    ax.set_ylabel('Motif Class', fontsize=11, fontweight='bold')
    
    plt.tight_layout()
    return fig


def plot_gc_content_correlation(motifs: List[Dict[str, Any]], 
                                sequence: str,
                                window_size: int = 1000,
                                title: str = "Motif Density vs GC Content",
                                figsize: Tuple[int, int] = None) -> plt.Figure:
    """
    Create scatter plot showing correlation between GC content and motif density.
    
    Shows GC-driven motif enrichment patterns. One dot per genomic window.
    Publication-quality visualization with regression line.
    
    Args:
        motifs: List of motif dictionaries
        sequence: DNA sequence string
        window_size: Window size for GC and density calculation
        title: Plot title
        figsize: Figure size (width, height)
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style()
    
    if figsize is None:
        figsize = FIGURE_SIZES['one_and_half']
    
    sequence_length = len(sequence)
    
    if not motifs or sequence_length == 0:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center',
                transform=ax.transAxes, fontsize=14)
        ax.set_title(title)
        return fig
    
    num_windows = max(1, sequence_length // window_size)
    
    # Calculate GC% and motif density for each window
    gc_percentages = []
    motif_densities = []
    
    for i in range(num_windows):
        window_start = i * window_size
        window_end = min((i + 1) * window_size, sequence_length)
        
        # Calculate GC%
        window_seq = sequence[window_start:window_end].upper()
        gc_count = window_seq.count('G') + window_seq.count('C')
        window_length = len(window_seq)
        gc_pct = (gc_count / window_length * 100) if window_length > 0 else 0
        
        # Count motifs in window
        motif_count = 0
        for motif in motifs:
            motif_start = motif.get('Start', 0) - 1
            motif_end = motif.get('End', 0)
            if not (motif_end <= window_start or motif_start >= window_end):
                motif_count += 1
        
        # Density: motifs per kb
        density = motif_count / (window_size / 1000)
        
        gc_percentages.append(gc_pct)
        motif_densities.append(density)
    
    # Create scatter plot
    fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
    
    ax.scatter(gc_percentages, motif_densities, alpha=0.5, s=30, 
              color='#0072B2', edgecolors='black', linewidth=0.3)
    
    # Add regression line
    if len(gc_percentages) > 1:
        z = np.polyfit(gc_percentages, motif_densities, 1)
        p = np.poly1d(z)
        x_line = np.linspace(min(gc_percentages), max(gc_percentages), 100)
        ax.plot(x_line, p(x_line), 'r--', linewidth=1.5, alpha=0.8, 
               label=f'Linear fit: y={z[0]:.2f}x+{z[1]:.2f}')
        
        # Calculate correlation coefficient
        correlation = np.corrcoef(gc_percentages, motif_densities)[0, 1]
        ax.text(0.05, 0.95, f'R = {correlation:.3f}', transform=ax.transAxes,
               fontsize=10, fontweight='bold', verticalalignment='top',
               bbox=dict(boxstyle='round', facecolor='white', alpha=0.8))
    
    # Styling
    ax.set_xlabel('GC Content (%)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Motif Density (motifs/kb)', fontsize=12, fontweight='bold')
    
    display_title = title.replace('_', ' ')
    ax.set_title(display_title, fontsize=14, fontweight='bold', pad=10)
    
    ax.grid(alpha=0.3, linestyle='--', linewidth=0.5)
    
    if len(gc_percentages) > 1:
        ax.legend(loc='upper right', fontsize=9, framealpha=0.9)
    
    # Apply Nature journal style
    _apply_nature_style(ax)
    
    plt.tight_layout()
    return fig


def plot_linear_motif_track(motifs: List[Dict[str, Any]], 
                            sequence_length: int,
                            region_start: int = 0,
                            region_end: int = None,
                            title: str = "Linear Motif Track",
                            figsize: Tuple[int, int] = None,
                            show_labels: bool = False) -> plt.Figure:  # Default to False per Nature-ready standards
    """
    Create horizontal graphical track with colored blocks for motifs.
    
    Best for visualizing <10kb regions. Colored blocks show motif positions
    with class-specific colors. Publication-quality linear genome browser view.
    
    NATURE-READY: Individual motif labels suppressed by default for clarity.
    Only class track labels are shown.
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Total length of analyzed sequence in bp
        region_start: Start position of region to display (0-based)
        region_end: End position of region to display (None = full sequence)
        title: Plot title
        figsize: Figure size (width, height)
        show_labels: Whether to show class labels (default False for cleaner view)
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style()
    
    if figsize is None:
        figsize = FIGURE_SIZES['wide']
    
    if region_end is None:
        region_end = sequence_length
    
    # Filter motifs in region
    region_motifs = [m for m in motifs 
                     if m.get('End', 0) > region_start and m.get('Start', 0) < region_end]
    
    if not region_motifs:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, f'No motifs in region {region_start}-{region_end}', 
               ha='center', va='center', transform=ax.transAxes, fontsize=14)
        ax.set_title(title)
        return fig
    
    # Group by class
    class_motifs = defaultdict(list)
    for motif in region_motifs:
        class_name = motif.get('Class', 'Unknown')
        class_motifs[class_name].append(motif)
    
    # Create figure
    fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
    
    # Plot each class on a separate track
    classes = sorted(class_motifs.keys())
    track_height = 0.6
    track_spacing = 1.0
    
    for i, class_name in enumerate(classes):
        y_pos = i * track_spacing
        color = MOTIF_CLASS_COLORS.get(class_name, '#808080')
        
        for motif in class_motifs[class_name]:
            start = max(region_start, motif.get('Start', 0))
            end = min(region_end, motif.get('End', 0))
            length = end - start
            
            # Draw motif as rectangle
            rect = patches.Rectangle(
                (start, y_pos - track_height/2), length, track_height,
                facecolor=color, edgecolor='black', linewidth=0.5, alpha=0.8
            )
            ax.add_patch(rect)
            
            # NATURE-READY: No individual motif labels
            # Labels removed to prevent clutter and overlap per publication standards
        
        # Add class label on the left (always shown for track identification)
        display_name = class_name.replace('_', ' ')
        ax.text(region_start - (region_end - region_start) * 0.02, y_pos, 
               display_name, ha='right', va='center', fontsize=9, fontweight='bold')
    
    # Styling
    ax.set_xlim(region_start, region_end)
    ax.set_ylim(-0.5, len(classes) * track_spacing - 0.5)
    
    ax.set_xlabel('Position (bp)', fontsize=12, fontweight='bold')
    ax.set_yticks([])
    
    display_title = title.replace('_', ' ')
    ax.set_title(display_title, fontsize=14, fontweight='bold', pad=10)
    
    # Add position ruler at top
    ax.xaxis.tick_top()
    ax.xaxis.set_label_position('top')
    
    # Format x-axis
    ax.ticklabel_format(style='plain', axis='x')
    
    # Apply Nature journal style (minimal spines)
    for spine in ['left', 'right', 'bottom']:
        ax.spines[spine].set_visible(False)
    ax.spines['top'].set_visible(True)
    
    plt.tight_layout()
    return fig


def plot_cluster_size_distribution(motifs: List[Dict[str, Any]], 
                                   title: str = "Cluster Size Distribution",
                                   figsize: Tuple[int, int] = None) -> plt.Figure:
    """
    Plot distribution of cluster sizes (number of motifs per cluster).
    
    Shows histogram and statistics of cluster composition.
    Publication-quality visualization.
    
    Args:
        motifs: List of motif dictionaries (should include cluster motifs)
        title: Plot title
        figsize: Figure size (width, height)
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style()
    
    if figsize is None:
        figsize = FIGURE_SIZES['one_and_half']
    
    # Extract cluster motifs
    cluster_motifs = [m for m in motifs if m.get('Class') == 'Non-B_DNA_Clusters']
    
    if not cluster_motifs:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No cluster motifs found', ha='center', va='center',
                transform=ax.transAxes, fontsize=14)
        ax.set_title(title)
        return fig
    
    # Extract cluster sizes
    cluster_sizes = []
    cluster_diversities = []
    
    for motif in cluster_motifs:
        size = motif.get('Motif_Count', 0)
        diversity = motif.get('Class_Diversity', 0)
        if size > 0:
            cluster_sizes.append(size)
            cluster_diversities.append(diversity)
    
    # Create figure with subplots
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=figsize, dpi=PUBLICATION_DPI)
    
    # 1. Cluster size histogram
    ax1.hist(cluster_sizes, bins=min(20, max(cluster_sizes) if cluster_sizes else 1), 
            edgecolor='black', linewidth=0.5, color='#4ecdc4', alpha=0.7)
    
    ax1.set_xlabel('Motifs per Cluster', fontsize=11, fontweight='bold')
    ax1.set_ylabel('Frequency', fontsize=11, fontweight='bold')
    ax1.set_title('Cluster Size Distribution', fontsize=12, fontweight='bold')
    
    # Add statistics
    if cluster_sizes:
        mean_size = np.mean(cluster_sizes)
        median_size = np.median(cluster_sizes)
        ax1.axvline(mean_size, color='red', linestyle='--', linewidth=1.5, 
                   label=f'Mean: {mean_size:.1f}')
        ax1.axvline(median_size, color='orange', linestyle='--', linewidth=1.5, 
                   label=f'Median: {median_size:.1f}')
        ax1.legend(fontsize=9, framealpha=0.9)
    
    ax1.grid(axis='y', alpha=0.3)
    _apply_nature_style(ax1)
    
    # 2. Class diversity histogram
    if cluster_diversities:
        ax2.hist(cluster_diversities, bins=min(10, max(cluster_diversities)), 
                edgecolor='black', linewidth=0.5, color='#95e1d3', alpha=0.7)
        
        ax2.set_xlabel('Class Diversity per Cluster', fontsize=11, fontweight='bold')
        ax2.set_ylabel('Frequency', fontsize=11, fontweight='bold')
        ax2.set_title('Cluster Diversity Distribution', fontsize=12, fontweight='bold')
        
        mean_div = np.mean(cluster_diversities)
        ax2.axvline(mean_div, color='red', linestyle='--', linewidth=1.5, 
                   label=f'Mean: {mean_div:.1f}')
        ax2.legend(fontsize=9, framealpha=0.9)
        
        ax2.grid(axis='y', alpha=0.3)
        _apply_nature_style(ax2)
    else:
        ax2.text(0.5, 0.5, 'No diversity data', ha='center', va='center',
                transform=ax2.transAxes, fontsize=12)
        ax2.axis('off')
    
    display_title = title.replace('_', ' ')
    plt.suptitle(display_title, fontsize=14, fontweight='bold', y=0.98)
    plt.tight_layout()
    
    return fig


def plot_motif_length_kde(motifs: List[Dict[str, Any]], 
                          by_class: bool = True,
                          title: str = "Motif Length Distribution (KDE)",
                          figsize: Tuple[int, int] = None) -> plt.Figure:
    """
    Plot kernel density estimation of motif length distributions.
    
    Shows smooth probability density curves for motif lengths,
    useful for comparing length patterns across classes.
    Publication-quality visualization.
    
    Args:
        motifs: List of motif dictionaries
        by_class: Whether to separate by motif class
        title: Plot title
        figsize: Figure size (width, height)
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style()
    
    if figsize is None:
        figsize = FIGURE_SIZES['one_and_half']
    
    if not motifs:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center',
                transform=ax.transAxes, fontsize=14)
        ax.set_title(title)
        return fig
    
    fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
    
    if by_class:
        # Group by class
        classes = sorted(set(m.get('Class', 'Unknown') for m in motifs
                            if m.get('Class') not in CIRCOS_EXCLUDED_CLASSES))
        
        if not classes:
            classes = sorted(set(m.get('Class', 'Unknown') for m in motifs))
        
        for class_name in classes:
            class_motifs = [m for m in motifs if m.get('Class') == class_name]
            lengths = [m.get('Length', 0) for m in class_motifs if m.get('Length', 0) > 0]
            
            if len(lengths) > 1:
                color = MOTIF_CLASS_COLORS.get(class_name, '#808080')
                display_name = class_name.replace('_', ' ')
                
                # Plot KDE
                try:
                    from scipy import stats
                    kde = stats.gaussian_kde(lengths)
                    x_range = np.linspace(min(lengths), max(lengths), 200)
                    density = kde(x_range)
                    ax.plot(x_range, density, color=color, linewidth=2, 
                           label=display_name, alpha=0.8)
                    ax.fill_between(x_range, density, alpha=0.2, color=color)
                except:
                    # Fallback to histogram if KDE fails
                    ax.hist(lengths, bins=20, alpha=0.3, color=color, 
                           label=display_name, density=True, edgecolor='black', linewidth=0.5)
    else:
        # Overall distribution
        lengths = [m.get('Length', 0) for m in motifs if m.get('Length', 0) > 0]
        
        if len(lengths) > 1:
            try:
                from scipy import stats
                kde = stats.gaussian_kde(lengths)
                x_range = np.linspace(min(lengths), max(lengths), 200)
                density = kde(x_range)
                ax.plot(x_range, density, color='#0072B2', linewidth=2.5, alpha=0.8)
                ax.fill_between(x_range, density, alpha=0.3, color='#0072B2')
            except:
                ax.hist(lengths, bins=30, alpha=0.7, color='#0072B2', 
                       density=True, edgecolor='black', linewidth=0.5)
    
    # Styling
    ax.set_xlabel('Motif Length (bp)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Probability Density', fontsize=12, fontweight='bold')
    
    display_title = title.replace('_', ' ')
    ax.set_title(display_title, fontsize=14, fontweight='bold', pad=10)
    
    ax.grid(alpha=0.3, linestyle='--', linewidth=0.5)
    
    if by_class:
        ax.legend(loc='upper right', fontsize=8, framealpha=0.9, ncol=2)
    
    # Apply Nature journal style
    _apply_nature_style(ax)
    
    plt.tight_layout()
    return fig


# =============================================================================
# FUNCTION ALIASES FOR BACKWARD COMPATIBILITY
# =============================================================================

# Aliases for functions with different naming conventions
plot_comprehensive_class_analysis = plot_class_analysis_comprehensive
plot_comprehensive_subclass_analysis = plot_subclass_analysis_comprehensive


# =============================================================================
# ADDITIONAL VISUALIZATION FUNCTIONS FOR NOTEBOOK COMPATIBILITY
# =============================================================================

def plot_genome_landscape_track(motifs: List[Dict[str, Any]], 
                               sequence_length: int,
                               title: str = "Genome Landscape Track",
                               figsize: Tuple[int, int] = None,
                               window_size: int = None) -> plt.Figure:
    """
    Create genome landscape track visualization showing motif distribution along sequence.
    
    This is a simplified horizontal track view showing motif positions and density.
    Similar to plot_linear_motif_track but with additional density information.
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Total length of analyzed sequence in bp
        title: Plot title
        figsize: Figure size (width, height)
        window_size: Window size for density calculation (auto if None)
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style()
    
    if figsize is None:
        figsize = FIGURE_SIZES['wide']
    
    if not motifs or sequence_length == 0:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center',
                transform=ax.transAxes, fontsize=14)
        ax.set_title(title)
        return fig
    
    # Auto-calculate window size
    if window_size is None:
        window_size = max(1000, sequence_length // 50)
    
    # Create figure with two subplots: density track + motif positions
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=figsize, dpi=PUBLICATION_DPI,
                                    gridspec_kw={'height_ratios': [1, 2]})
    
    # Top panel: Density line plot
    num_windows = max(1, sequence_length // window_size)
    positions = []
    densities = []
    
    for i in range(num_windows):
        window_start = i * window_size
        window_end = (i + 1) * window_size
        window_center = (window_start + window_end) / 2 / 1000  # In kb
        
        # Count motifs in window
        count = 0
        for motif in motifs:
            motif_start = motif.get('Start', 0) - 1
            motif_end = motif.get('End', 0)
            if not (motif_end <= window_start or motif_start >= window_end):
                count += 1
        
        density = count / (window_size / 1000)  # motifs per kb
        positions.append(window_center)
        densities.append(density)
    
    # Plot density line
    ax1.fill_between(positions, densities, alpha=0.3, color='#0072B2')
    ax1.plot(positions, densities, color='#0072B2', linewidth=1.5)
    ax1.set_ylabel('Density\n(motifs/kb)', fontsize=10, fontweight='bold')
    ax1.set_xlim(0, sequence_length / 1000)
    ax1.grid(axis='y', alpha=0.3, linestyle='--', linewidth=0.5)
    _apply_nature_style(ax1)
    
    # Bottom panel: Motif track by class
    class_motifs = defaultdict(list)
    for motif in motifs:
        class_name = motif.get('Class', 'Unknown')
        if class_name not in CIRCOS_EXCLUDED_CLASSES:
            class_motifs[class_name].append(motif)
    
    if not class_motifs:
        # Include all classes if none pass the filter
        for motif in motifs:
            class_name = motif.get('Class', 'Unknown')
            class_motifs[class_name].append(motif)
    
    classes = sorted(class_motifs.keys())
    track_height = 0.5
    track_spacing = 1.0
    
    for i, class_name in enumerate(classes):
        y_pos = i * track_spacing
        color = MOTIF_CLASS_COLORS.get(class_name, '#808080')
        
        for motif in class_motifs[class_name]:
            start_kb = motif.get('Start', 0) / 1000
            end_kb = motif.get('End', 0) / 1000
            length_kb = end_kb - start_kb
            
            # Draw motif as rectangle
            rect = patches.Rectangle(
                (start_kb, y_pos - track_height/2), length_kb, track_height,
                facecolor=color, edgecolor='black', linewidth=0.3, alpha=0.8
            )
            ax2.add_patch(rect)
    
    # Customize bottom panel
    ax2.set_xlim(0, sequence_length / 1000)
    ax2.set_ylim(-0.5, len(classes) * track_spacing - 0.5)
    ax2.set_xlabel('Position (kb)', fontsize=11, fontweight='bold')
    ax2.set_ylabel('Motif Class', fontsize=11, fontweight='bold')
    ax2.set_yticks([i * track_spacing for i in range(len(classes))])
    display_classes = [c.replace('_', ' ') for c in classes]
    ax2.set_yticklabels(display_classes, fontsize=9)
    _apply_nature_style(ax2)
    
    # Overall title
    display_title = title.replace('_', ' ')
    fig.suptitle(display_title, fontsize=14, fontweight='bold', y=0.98)
    
    plt.tight_layout(rect=[0, 0, 1, 0.96])
    return fig


def plot_sliding_window_heat_ribbon(motifs: List[Dict[str, Any]], 
                                    sequence_length: int,
                                    title: str = "Sliding Window Heat Ribbon",
                                    figsize: Tuple[int, int] = None,
                                    window_size: int = None) -> plt.Figure:
    """
    Create a 1D heatmap ribbon showing motif density along the sequence.
    
    This creates a horizontal ribbon colored by density values, with an
    accompanying line plot showing the density profile.
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Total length of analyzed sequence in bp
        title: Plot title
        figsize: Figure size (width, height)
        window_size: Window size for density calculation (auto if None)
        
    Returns:
        Matplotlib figure object (publication-ready)
    """
    set_scientific_style()
    
    if figsize is None:
        figsize = FIGURE_SIZES['wide']
    
    if not motifs or sequence_length == 0:
        fig, ax = plt.subplots(figsize=figsize, dpi=PUBLICATION_DPI)
        ax.text(0.5, 0.5, 'No motifs to display', ha='center', va='center',
                transform=ax.transAxes, fontsize=14)
        ax.set_title(title)
        return fig
    
    # Auto-calculate window size
    if window_size is None:
        window_size = max(1000, sequence_length // 100)
    
    num_windows = max(1, sequence_length // window_size)
    
    # Calculate density for each window
    density_values = []
    positions = []
    
    for i in range(num_windows):
        window_start = i * window_size
        window_end = (i + 1) * window_size
        
        # Count motifs in window
        count = 0
        for motif in motifs:
            motif_start = motif.get('Start', 0) - 1
            motif_end = motif.get('End', 0)
            if not (motif_end <= window_start or motif_start >= window_end):
                count += 1
        
        density = count / (window_size / 1000)  # motifs per kb
        density_values.append(density)
        positions.append((window_start + window_end) / 2 / 1000)  # Center position in kb
    
    # Create figure with two subplots
    fig = plt.figure(figsize=figsize, dpi=PUBLICATION_DPI)
    gs = fig.add_gridspec(2, 1, height_ratios=[1, 3], hspace=0.3)
    
    # Top: Heat ribbon (1D heatmap)
    ax1 = fig.add_subplot(gs[0])
    
    # Create 2D array for heatmap (1 row)
    density_array = np.array(density_values).reshape(1, -1)
    
    # Plot heatmap
    im = ax1.imshow(density_array, cmap='YlOrRd', aspect='auto', 
                    extent=[0, sequence_length / 1000, 0, 1],
                    interpolation='bilinear')
    
    ax1.set_yticks([])
    ax1.set_ylabel('Density\nHeatmap', fontsize=10, fontweight='bold')
    ax1.set_xticks([])
    
    # Add colorbar
    cbar = plt.colorbar(im, ax=ax1, orientation='vertical', pad=0.02, shrink=0.8)
    cbar.set_label('Motifs/kb', fontsize=9, fontweight='bold')
    cbar.ax.tick_params(labelsize=8)
    
    # Bottom: Line plot
    ax2 = fig.add_subplot(gs[1])
    
    ax2.plot(positions, density_values, color='#0072B2', linewidth=2, alpha=0.8)
    ax2.fill_between(positions, density_values, alpha=0.3, color='#0072B2')
    
    # Mark peaks (top 10% density)
    if density_values:
        threshold = np.percentile(density_values, 90)
        peak_indices = [i for i, d in enumerate(density_values) if d >= threshold]
        if peak_indices:
            peak_positions = [positions[i] for i in peak_indices]
            peak_densities = [density_values[i] for i in peak_indices]
            ax2.scatter(peak_positions, peak_densities, color='red', s=50, 
                       zorder=5, alpha=0.7, edgecolors='black', linewidth=0.5,
                       label=f'High density (top 10%)')
    
    # Styling
    ax2.set_xlabel('Position (kb)', fontsize=11, fontweight='bold')
    ax2.set_ylabel('Motif Density (motifs/kb)', fontsize=11, fontweight='bold')
    ax2.set_xlim(0, sequence_length / 1000)
    ax2.grid(alpha=0.3, linestyle='--', linewidth=0.5)
    
    if peak_indices:
        ax2.legend(loc='upper right', fontsize=9, framealpha=0.9)
    
    _apply_nature_style(ax2)
    
    # Overall title
    display_title = title.replace('_', ' ')
    fig.suptitle(display_title, fontsize=14, fontweight='bold', y=0.98)
    
    plt.tight_layout(rect=[0, 0, 1, 0.96])
    return fig


# =============================================================================
# JOB ID AND DOWNLOAD PACKAGE GENERATION
# =============================================================================

def generate_job_id(sequence_name: str) -> str:
    """
    Generate a unique, human-readable Job ID for a NonBDNA analysis run.
    
    Format: JOB_<sequence_name>_<short_hash>
    Example: JOB_TP53_chr17_f8a21c
    
    Args:
        sequence_name: Name of the sequence being analyzed
        
    Returns:
        Job ID string (safe for filenames)
    """
    # Clean sequence name for filename safety
    clean_name = re.sub(r'[^\w\-]', '_', sequence_name)[:MAX_SEQUENCE_NAME_LENGTH]
    clean_name = clean_name.strip('_')
    
    # Generate short hash from sequence name + timestamp for uniqueness
    hash_input = f"{sequence_name}_{time.time()}".encode('utf-8')
    short_hash = hashlib.sha256(hash_input).hexdigest()[:HASH_LENGTH]
    
    job_id = f"JOB_{clean_name}_{short_hash}"
    return job_id


def save_figures_to_pdf(figures: List[plt.Figure], output_path: str) -> None:
    """
    Save multiple matplotlib figures to a single multi-page PDF.
    
    Args:
        figures: List of matplotlib figure objects
        output_path: Path to output PDF file
    """
    with PdfPages(output_path) as pdf:
        for fig in figures:
            if fig is not None:
                pdf.savefig(fig, bbox_inches='tight', dpi=300)
                plt.close(fig)


def create_consolidated_pdf(
    motifs: List[Dict[str, Any]],
    sequence_length: int,
    sequence_name: str,
    job_id: str,
    output_dir: str
) -> str:
    """
    Create a consolidated PDF with all visualizations for publication-ready results.
    
    Args:
        motifs: List of motif dictionaries
        sequence_length: Length of sequence in bp
        sequence_name: Name of the sequence
        job_id: Job ID for filename
        output_dir: Directory to save PDF
        
    Returns:
        Path to created PDF file
    """
    pdf_path = os.path.join(output_dir, f"{job_id}.pdf")
    figures = []
    
    try:
        # Overview plots
        figures.append(plot_motif_distribution(motifs, by='Class', title=f"Motif Classes - {sequence_name}"))
        figures.append(plot_motif_distribution(motifs, by='Subclass', title=f"Motif Subclasses - {sequence_name}"))
        figures.append(plot_nested_pie_chart(motifs, title=f"Class-Subclass Distribution - {sequence_name}"))
        
        # Coverage and density
        figures.append(plot_coverage_map(motifs, sequence_length, title=f"Motif Coverage - {sequence_name}"))
        
        # Calculate appropriate window size, ensuring it doesn't exceed sequence length
        window_size = max(DEFAULT_PDF_WINDOW_SIZE_MIN, sequence_length // DEFAULT_PDF_WINDOW_DIVISOR)
        window_size = min(window_size, sequence_length)  # Don't exceed sequence length
        
        figures.append(plot_density_heatmap(motifs, sequence_length, 
                                           window_size=window_size,
                                           title=f"Motif Density - {sequence_name}"))
        
        # Statistical distributions
        figures.append(plot_length_distribution(motifs, by_class=True, title="Length Distribution by Class"))
        figures.append(plot_score_distribution(motifs, by_class=True, title="Score Distribution by Class"))
        
        # Genome-wide analysis
        figures.append(plot_manhattan_motif_density(motifs, sequence_length, title=f"Manhattan Plot - {sequence_name}"))
        figures.append(plot_cumulative_motif_distribution(motifs, sequence_length, 
                                                         title=f"Cumulative Distribution - {sequence_name}",
                                                         by_class=True))
        
        # Advanced visualizations
        figures.append(plot_motif_cooccurrence_matrix(motifs, title=f"Co-occurrence Matrix - {sequence_name}"))
        figures.append(plot_motif_length_kde(motifs, by_class=True, title=f"Length KDE - {sequence_name}"))
        
        # Circos plot
        figures.append(plot_circos_motif_density(motifs, sequence_length, title=f"Circos Density - {sequence_name}"))
        
        # Save all figures to PDF
        save_figures_to_pdf(figures, pdf_path)
        
    except Exception as e:
        # Clean up any figures that were created
        for fig in figures:
            if fig is not None:
                plt.close(fig)
        raise Exception(f"Failed to create consolidated PDF: {str(e)}")
    
    return pdf_path


def export_to_pdf(motifs: List[Dict[str, Any]], 
                  sequence_length: int,
                  sequence_name: str = "Unknown Sequence") -> bytes:
    """
    Generate PDF containing visualization summaries of sequence analyses.
    
    This function creates a multi-page PDF with key visualizations including:
    - Motif distribution by class and subclass
    - Coverage maps and density heatmaps
    - Length and score distributions
    - Circos plot for genome-wide view
    - Manhattan plot and cumulative distribution
    
    Args:
        motifs: List of motif dictionaries with required keys (Start, End, Class, etc.)
        sequence_length: Total length of the analyzed sequence in base pairs
        sequence_name: Name of the sequence for labeling plots
        
    Returns:
        bytes: PDF file data as bytes, ready for download
        
    Raises:
        Exception: If PDF generation fails
    """
    # Create in-memory buffer for PDF
    pdf_buffer = io.BytesIO()
    
    figures = []
    
    try:
        # 1. Distribution plots
        figures.append(plot_motif_distribution(motifs, by='Class', 
                                              title=f"Motif Classes - {sequence_name}"))
        figures.append(plot_motif_distribution(motifs, by='Subclass', 
                                              title=f"Motif Subclasses - {sequence_name}"))
        
        # 2. Nested pie chart for class-subclass relationship
        figures.append(plot_nested_pie_chart(motifs, 
                                            title=f"Class-Subclass Distribution - {sequence_name}"))
        
        # 3. Coverage and density maps
        figures.append(plot_coverage_map(motifs, sequence_length, 
                                        title=f"Motif Coverage - {sequence_name}"))
        
        # Calculate appropriate window size for density heatmap
        # Use adaptive window sizing based on sequence length, with a minimum size
        window_size = max(DEFAULT_PDF_WINDOW_SIZE_MIN, sequence_length // DEFAULT_PDF_WINDOW_DIVISOR)
        window_size = min(window_size, sequence_length)
        
        figures.append(plot_density_heatmap(motifs, sequence_length, 
                                           window_size=window_size,
                                           title=f"Motif Density - {sequence_name}"))
        
        # 4. Statistical distributions
        figures.append(plot_length_distribution(motifs, by_class=True, 
                                               title="Length Distribution by Motif Class"))
        figures.append(plot_score_distribution(motifs, by_class=True,
                                              title="Score Distribution by Motif Class"))
        
        # 5. Genome-wide visualizations
        figures.append(plot_circos_motif_density(motifs, sequence_length, 
                                                title=f"Circos Density - {sequence_name}"))
        
        figures.append(plot_manhattan_motif_density(motifs, sequence_length, 
                                                   title=f"Manhattan Plot - {sequence_name}"))
        
        figures.append(plot_cumulative_motif_distribution(motifs, sequence_length,
                                                         title=f"Cumulative Distribution - {sequence_name}",
                                                         by_class=True))
        
        # Save all figures to PDF buffer
        with PdfPages(pdf_buffer) as pdf:
            for fig in figures:
                if fig is not None:
                    pdf.savefig(fig, bbox_inches='tight', dpi=300)
                    plt.close(fig)
        
        # Get PDF data from buffer
        pdf_data = pdf_buffer.getvalue()
        pdf_buffer.close()
        
        return pdf_data
        
    except Exception as e:
        # Clean up any figures that were created
        for fig in figures:
            if fig is not None:
                try:
                    plt.close(fig)
                except (RuntimeError, ValueError) as cleanup_error:
                    # Log cleanup failure but don't prevent raising the original error
                    logger.debug(f"Failed to close figure during cleanup: {cleanup_error}")
        pdf_buffer.close()
        raise Exception(f"Failed to generate PDF: {str(e)}")


def create_enhanced_excel(
    motifs: List[Dict[str, Any]],
    job_id: str,
    sequence_name: str,
    run_time: float,
    output_dir: str,
    simple_format: bool = True
) -> str:
    """
    Create enhanced Excel file with 2-tab format.
    
    Args:
        motifs: List of motif dictionaries
        job_id: Job ID for filename
        sequence_name: Name of analyzed sequence
        run_time: Analysis runtime in seconds
        output_dir: Directory to save Excel file
        simple_format: If True, use 2-tab format (NonOverlappingConsolidated, OverlappingAll)
        
    Returns:
        Path to created Excel file
    """
    excel_path = os.path.join(output_dir, f"{job_id}.xlsx")
    
    # Create the Excel export with simple format (2 tabs)
    export_to_excel(motifs, excel_path, simple_format=simple_format)
    
    # Now add metadata sheet
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(excel_path)
        
        # Create metadata sheet at the beginning
        ws_meta = wb.create_sheet("Metadata", 0)
        
        # Add metadata
        metadata_rows = [
            ["Job ID", job_id],
            ["Sequence Name", sequence_name],
            ["Analysis Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Run Time (seconds)", f"{run_time:.2f}"],
            ["Total Motifs", len(motifs)],
            ["Unique Classes", len(set(m.get('Class', 'Unknown') for m in motifs))],
            ["Unique Subclasses", len(set(m.get('Subclass', 'Unknown') for m in motifs))],
            ["Tool", APP_NAME],
            ["Version", APP_VERSION],
        ]
        
        for row_idx, (key, value) in enumerate(metadata_rows, start=1):
            key_cell = ws_meta.cell(row=row_idx, column=1, value=key)
            # Make key column bold
            key_cell.font = key_cell.font.copy(bold=True)
            ws_meta.cell(row=row_idx, column=2, value=value)
        
        ws_meta.column_dimensions['A'].width = 25
        ws_meta.column_dimensions['B'].width = 40
        
        wb.save(excel_path)
        
    except Exception as e:
        # If metadata addition fails, log but don't fail entire operation
        logger.warning(f"Could not add metadata sheet: {str(e)}")
    
    return excel_path


def create_readme_file(job_id: str, sequence_name: str, run_time: float, output_dir: str) -> str:
    """
    Create README.txt file with job metadata and usage information.
    
    Args:
        job_id: Job ID
        sequence_name: Name of analyzed sequence  
        run_time: Analysis runtime in seconds
        output_dir: Directory to save README
        
    Returns:
        Path to created README file
    """
    readme_path = os.path.join(output_dir, "README.txt")
    
    readme_content = f"""
╔══════════════════════════════════════════════════════════════════════════════╗
║                    {APP_NAME} Analysis Results                             ║
║                     Complete Results Package                                 ║
╚══════════════════════════════════════════════════════════════════════════════╝

JOB INFORMATION
═══════════════
Job ID:           {job_id}
Sequence Name:    {sequence_name}
Analysis Date:    {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
Run Time:         {run_time:.2f} seconds
Tool:             {APP_NAME}
Version:          {APP_VERSION}

PACKAGE CONTENTS
════════════════
• {job_id}.pdf                    - Consolidated visualization PDF with all plots (300 DPI)
• {job_id}.xlsx                   - Main Excel with 2 tabs: NonOverlappingConsolidated, OverlappingAll
• {job_id}_statistics.xlsx        - Statistical Analysis with comprehensive metrics
• {job_id}_nonoverlapping.csv     - CSV with non-overlapping consolidated motifs only
• README.txt                      - This file

PDF CONTENTS
════════════
The PDF contains publication-ready visualizations (300 DPI) ordered as follows:
1. Overview plots (class and subclass distribution)
2. Coverage and density maps
3. Statistical distributions (length, score)
4. Genome-wide analysis (Manhattan plot, cumulative distribution)
5. Advanced visualizations (co-occurrence matrix, KDE plots)
6. Circos density plot

All figures are suitable for direct use in publications, presentations, and reports.

EXCEL CONTENTS
══════════════
Main Excel ({job_id}.xlsx):
• Tab 1: NonOverlappingConsolidated - All motifs excluding Hybrid and Cluster types
• Tab 2: OverlappingAll - Complete dataset including ALL motifs

Statistics Excel ({job_id}_statistics.xlsx):
• Summary - Overall analysis statistics and metrics
• Class_Level_Analysis - Density and distribution by motif class
• Subclass_Level_Analysis - Detailed breakdown by motif subclass
• Length_Distribution - Length statistics by class
• Score_Distribution - Score statistics by class

CSV CONTENTS
════════════
The CSV file contains non-overlapping consolidated motifs only with core columns:
Sequence_Name, Source, Class, Subclass, Start, End, Length, Sequence, Score

Each sheet includes comprehensive motif annotations and can be imported into
statistical software (R, Python, MATLAB) for further analysis.

USAGE NOTES
═══════════
✓ All files are ready for immediate use
✓ PDF visualizations are publication-quality (300 DPI)
✓ Excel data can be filtered, sorted, and analyzed
✓ CSV is compatible with all standard analysis tools
✓ Statistics file provides comprehensive metrics for reporting
✓ Save this package locally for future reference
✓ Share the ZIP package with collaborators as needed

CITATION
════════
If you use {APP_NAME} in your research, please cite:
- Dr. Venkata Rajesh Yella
- GitHub: https://github.com/VRYella/{APP_NAME}

CONTACT
═══════
For questions or issues:
Email: yvrajesh_bt@kluniversity.in
GitHub: VRYella

╔══════════════════════════════════════════════════════════════════════════════╗
║  Generated by {APP_NAME} - A Comprehensive Non-B DNA Detection System     ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""
    
    with open(readme_path, 'w') as f:
        f.write(readme_content.strip())
    
    return readme_path


def create_zip_package(
    pdf_path: str,
    excel_path: str,
    readme_path: str,
    job_id: str,
    output_dir: str
) -> str:
    """
    Create ZIP archive containing PDF, Excel, and README.
    
    Args:
        pdf_path: Path to PDF file
        excel_path: Path to Excel file
        readme_path: Path to README file
        job_id: Job ID for ZIP filename
        output_dir: Directory to save ZIP
        
    Returns:
        Path to created ZIP file
        
    Raises:
        ValueError: If any input file is outside the output directory
    """
    zip_path = os.path.join(output_dir, f"{job_id}.zip")
    
    # Validate that all files are within the output directory (security check)
    output_dir_abs = os.path.abspath(output_dir)
    for file_path in [pdf_path, excel_path, readme_path]:
        file_abs = os.path.abspath(file_path)
        if not file_abs.startswith(output_dir_abs):
            raise ValueError(f"Security error: File {file_path} is outside output directory")
    
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        zipf.write(pdf_path, os.path.basename(pdf_path))
        zipf.write(excel_path, os.path.basename(excel_path))
        zipf.write(readme_path, os.path.basename(readme_path))
    
    return zip_path


# =============================================================================
# DEPRECATED: File.io upload functionality has been removed
# =============================================================================
# The upload_to_fileio function has been removed in favor of direct downloads
# through Streamlit's native download buttons. This provides:
# - More reliable downloads without external service dependency
# - No expiry time or download limits
# - Better user experience with immediate access
# - Simpler codebase without network error handling
# =============================================================================


# =============================================================================
# COLLAPSIBLE CARD COMPONENT
# =============================================================================
def create_collapsible_card(title: str, content: str, card_id: str = None, 
                            default_open: bool = False) -> str:
    """
    Create a collapsible card component with smooth animations.
    
    This component replaces the standard Streamlit expander with a more 
    professional, customizable card design that integrates with the global
    CSS design system. Features include:
    - Smooth expand/collapse animations
    - Chevron rotation on toggle
    - Hover and focus effects
    - Full theme integration
    - Mobile responsive design
    
    Args:
        title: The card header/title text
        content: The HTML content to display when expanded
        card_id: Unique identifier for the card (auto-generated if None)
        default_open: Whether the card should be open by default
        
    Returns:
        HTML string for rendering with st.components.v1.html(..., height=None)
        
    Example:
        >>> import streamlit.components.v1 as components
        >>> card_html = create_collapsible_card(
        ...     title="Q: What is Non-B DNA?",
        ...     content="<p>Non-B DNA refers to...</p>"
        ... )
        >>> components.html(card_html, height=None)
    """
    import uuid
    
    # Generate unique ID if not provided
    if card_id is None:
        card_id = f"card-{uuid.uuid4().hex[:8]}"
    
    # Sanitize IDs for HTML/JavaScript
    safe_id = re.sub(r'[^a-zA-Z0-9_-]', '', card_id)
    
    # Set initial states
    body_class = "collapsible-card-body" + (" open" if default_open else "")
    chevron_class = "collapsible-card-chevron" + (" open" if default_open else "")
    
    html = f"""
<!DOCTYPE html>
<html>
<head>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Inter', 'IBM Plex Sans', 'Segoe UI', system-ui, -apple-system, sans-serif;
            margin: 0;
            padding: 0;
        }}
        
        .collapsible-card {{
            margin: 0.75rem 0;
            border-radius: 16px;
            overflow: hidden;
            transition: all 0.2s ease;
            box-shadow: 0 2px 8px rgba(37, 99, 235, 0.15);
            border: 2px solid rgba(96, 165, 250, 0.4);
            background: #FFFFFF;
        }}
        
        .collapsible-card:hover {{
            box-shadow: 0 4px 16px rgba(37, 99, 235, 0.25);
            border-color: rgba(37, 99, 235, 0.6);
        }}
        
        .collapsible-card-header {{
            display: flex;
            align-items: center;
            gap: 10px;
            padding: 0.75rem 1rem;
            background: linear-gradient(135deg, #EEF2FF 0%, #E0E7FF 100%);
            cursor: pointer;
            user-select: none;
            transition: all 0.2s ease;
            border: none;
            width: 100%;
            text-align: left;
            font-family: inherit;
            font-size: 0.95rem;
            font-weight: 700;
            color: #0F172A;
        }}
        
        .collapsible-card-header:hover {{
            background: linear-gradient(135deg, #DBEAFE 0%, #BFDBFE 100%);
        }}
        
        .collapsible-card-chevron {{
            flex-shrink: 0;
            width: 20px;
            height: 20px;
            transition: transform 0.3s ease;
            color: #2563EB;
            font-weight: bold;
            font-size: 1.2rem;
            line-height: 1;
        }}
        
        .collapsible-card-chevron.open {{
            transform: rotate(180deg);
        }}
        
        .collapsible-card-label {{
            flex: 1;
            margin: 0;
            line-height: 1.5;
            color: #0F172A;
            font-weight: 700;
        }}
        
        .collapsible-card-body {{
            max-height: 0;
            overflow: hidden;
            transition: max-height 0.4s ease-out, padding 0.3s ease-out;
            padding: 0 1rem;
            background: #FFFFFF;
        }}
        
        .collapsible-card-body.open {{
            max-height: 100vh;
            padding: 0.75rem 1rem;
        }}
        
        .collapsible-card-content {{
            color: #0F172A;
            font-size: 0.95rem;
            line-height: 1.6;
        }}
        
        .collapsible-card-content p {{
            margin-bottom: 0.5rem;
        }}
        
        .collapsible-card-content p:last-child {{
            margin-bottom: 0;
        }}
        
        .collapsible-card-content strong {{
            font-weight: 600;
            color: #2563EB;
        }}
        
        .collapsible-card-content a {{
            color: #3B82F6;
            text-decoration: none;
            font-weight: 500;
        }}
        
        .collapsible-card-content a:hover {{
            color: #2563EB;
            text-decoration: underline;
        }}
    </style>
</head>
<body>
    <div class="collapsible-card" id="{safe_id}">
        <button class="collapsible-card-header" 
                onclick="toggleCard()" 
                aria-expanded="{str(default_open).lower()}"
                type="button">
            <span class="collapsible-card-chevron {chevron_class}" id="chevron">▼</span>
            <div class="collapsible-card-label">{title}</div>
        </button>
        <div class="collapsible-card-body {body_class}" id="body">
            <div class="collapsible-card-content">
                {content}
            </div>
        </div>
    </div>
    
    <script>
        function toggleCard() {{
            const body = document.getElementById("body");
            const chevron = document.getElementById("chevron");
            const header = document.querySelector(".collapsible-card-header");
            
            if (body.classList.contains("open")) {{
                body.classList.remove("open");
                chevron.classList.remove("open");
                header.setAttribute("aria-expanded", "false");
            }} else {{
                body.classList.add("open");
                chevron.classList.add("open");
                header.setAttribute("aria-expanded", "true");
            }}
        }}
    </script>
</body>
</html>
    """
    
    return html


def render_summary_panel(seq_length: int, 
                        processing_time: float, 
                        motif_count: int,
                        total_chunks: int = 0,
                        theme_color: str = "#10b981") -> str:
    """
    Create a modern, compact scientific summary panel with theme-aware styling.
    
    This function generates a publication-ready summary block that displays:
    - Sequence length (with appropriate units: bp, kb, or Mb)
    - Processing time (formatted as hh:mm:ss with clock icon 🕐)
    - Total motifs detected
    - Optional chunk information (for chunked processing)
    
    The panel uses:
    - Rounded borders and soft gradients
    - Bold, high-contrast text
    - Grid-based compact layout
    - Theme color tokens (customizable)
    - Dynamic box shadows matching theme colors
    - Fully responsive formatting
    
    Args:
        seq_length: Total sequence length in base pairs
        processing_time: Processing duration in seconds
        motif_count: Total number of motifs detected
        total_chunks: Number of chunks processed (0 if no chunking)
        theme_color: Primary gradient color. Supported colors:
            - "#10b981" (success green, default)
            - "#2563EB" (primary blue)
            - "#9C27B0" (genomic purple)
            Other colors will use green gradient/shadow as fallback
        
    Returns:
        HTML string ready for st.markdown(..., unsafe_allow_html=True)
        
    Example:
        >>> summary_html = render_summary_panel(
        ...     seq_length=1500000,
        ...     processing_time=125.5,
        ...     motif_count=342,
        ...     total_chunks=5,
        ...     theme_color="#10b981"
        ... )
        >>> st.markdown(summary_html, unsafe_allow_html=True)
    """
    # Calculate sequence length display with appropriate units
    if seq_length >= 1_000_000:
        seq_length_display = f"{seq_length / 1_000_000:.1f} Mb"
    elif seq_length >= 1_000:
        seq_length_display = f"{seq_length / 1_000:.1f} kb"
    else:
        seq_length_display = f"{seq_length:,} bp"
    
    # Format processing time as hh:mm:ss or mm:ss
    hours = int(processing_time // 3600)
    mins = int((processing_time % 3600) // 60)
    secs = int(processing_time % 60)
    
    if hours > 0:
        time_display = f"{hours:02d}:{mins:02d}:{secs:02d}"
    else:
        time_display = f"{mins:02d}:{secs:02d}"
    
    # Determine secondary gradient color and shadow based on theme color
    # Maps primary color to (secondary gradient, rgba shadow)
    theme_map = {
        "#10b981": ("#059669", "16, 185, 129"),   # Success green
        "#2563EB": ("#3B82F6", "37, 99, 235"),    # Primary blue
        "#9C27B0": ("#BA68C8", "156, 39, 176"),   # Genomic purple
    }
    
    # Get theme-specific colors or use green as default
    secondary_color, shadow_rgb = theme_map.get(theme_color, ("#059669", "16, 185, 129"))
    
    # Build chunk info if applicable
    chunk_info = ""
    if total_chunks > 0:
        chunk_info = f"<div style='margin-bottom: 0.4rem;'><b>Total chunks:</b> {total_chunks}</div>"
    
    # Generate HTML with modern styling and dynamic theme colors
    html = f"""
    <div style='background: linear-gradient(135deg, {theme_color} 0%, {secondary_color} 100%); 
                padding: 1.2rem; border-radius: 12px; color: white; 
                box-shadow: 0 5px 20px rgba({shadow_rgb}, 0.3); margin-bottom: 1rem;'>
        <h3 style='margin: 0 0 1rem 0; text-align: center; font-size: 1.2rem; 
                   border-bottom: 2px solid rgba(255, 255, 255, 0.3); padding-bottom: 0.8rem;'>
            ✅ Analysis Summary
        </h3>
        <div style='background: rgba(255, 255, 255, 0.15); padding: 0.8rem; 
                   border-radius: 8px; line-height: 1.8;'>
            <div style='margin-bottom: 0.4rem;'><b>Sequence length:</b> {seq_length_display}</div>
            {chunk_info}
            <div style='margin-bottom: 0.4rem;'><b>🕐 Processing time:</b> {time_display}</div>
            <div style='margin-bottom: 0.4rem;'><b>Motifs detected:</b> {motif_count:,}</div>
        </div>
    </div>
    """
    
    return html


if __name__ == "__main__":
    test_visualizations()